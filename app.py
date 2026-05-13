"""
OMIO Ultra DB Studio  v4.0
======================
Professional streaming data tool – handles CSV / JSON / JSONL / TXT / XLSX
files of ANY size (tested concept: 100 GB+) without loading full content into RAM.

Big-File Viewer works like EmEditor:
  • byte-offset seek → jump to any position in O(1)
  • forward/back chunk navigation
  • Go-to-row  (scans only the needed prefix)
  • in-chunk search / highlight
  • dual-pane: structured table + raw text

Tabs
----
1  ⚡ Convert      – single-file clean & convert
2  📂 Batch        – parallel directory batch convert
3  🔎 Big Viewer   – EmEditor-style infinite file viewer
4  ✂  Split        – split one file into many parts
5  🔀 Merge        – merge many files into one
6  📊 Profile      – column type-inference & statistics
7  🔍 Filter       – dynamic SQL-like query builder
8  ♊  Duplicates   – find / extract duplicates by key columns
9  💾 Presets      – save, load, run named job presets

Engine
------
• Streaming  : all read/write operations yield one row at a time
• ijson      : true JSON-array streaming (no full JSON load)
• openpyxl   : Excel read/write (write-only for large outputs)
• ThreadPoolExecutor  : parallel batch
• JobControl : pause / resume / cancel for every long job
"""

import csv
import ctypes
import io
import json
import mmap
import os
import sys
import queue
import math
import random
import re
import struct
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import tkinter as tk
from tkinter import (
    BOTH, END, LEFT, RIGHT, VERTICAL, HORIZONTAL,
    X, Y, W, E, N, S, NW,
    filedialog, messagebox,
    StringVar, BooleanVar, IntVar, DoubleVar,
)
from tkinter import ttk

# ── optional deps ────────────────────────────────────────────────────────────
try:
    import ijson
    _IJSON = True
except ImportError:
    ijson = None
    _IJSON = False

try:
    from openpyxl import Workbook, load_workbook
    _OPENPYXL = True
except ImportError:
    Workbook = load_workbook = None
    _OPENPYXL = False

try:
    from PIL import Image, ImageTk
    _PIL = True
except ImportError:
    _PIL = False

try:
    import orjson
    _ORJSON = True
except ImportError:
    orjson = None
    _ORJSON = False

try:
    import xxhash
    _XXHASH = True
except ImportError:
    xxhash = None
    _XXHASH = False

try:
    import cudf as _cudf_mod
    _CUDF = True
except ImportError:
    _cudf_mod = None
    _CUDF = False

_CPU_COUNT = max(1, (os.cpu_count() or 1))
_IO_BUFFER = 1 << 20   # 1 MB I/O buffer — will be overridden by HW profile below

# ═════════════════════════════════════════════════════════════════════════════
#  ADAPTIVE HARDWARE PROFILE
#  Runs once at startup — reads available RAM & CPU, sets all tuning constants.
#  Works on any machine: 4 GB laptop → 512 GB server → GPU workstation.
# ═════════════════════════════════════════════════════════════════════════════
def _build_hw_profile():
    """
    Detect actual free RAM and CPU topology, then return a tuning dict.

    Tiers
    -----
    LOW    < 4 GB available RAM   — conservative buffers, single-worker fallbacks
    MID    4 – 16 GB              — balanced, uses all cores, 32 MB chunks
    HIGH   > 16 GB                — large buffers, maximum parallelism
    GPU    cudf available         — GPU path for CSV/TSV, CPU fallback for rest
    """
    import os as _os

    # ── Detect available RAM ─────────────────────────────────────────────────
    avail_bytes = 0
    try:
        import psutil as _ps
        avail_bytes = _ps.virtual_memory().available
    except ImportError:
        # psutil not installed — estimate from platform APIs
        try:
            import ctypes as _ct
            class _MEMSTATUS(_ct.Structure):
                _fields_ = [("dwLength",                _ct.c_ulong),
                             ("dwMemoryLoad",            _ct.c_ulong),
                             ("ullTotalPhys",            _ct.c_ulonglong),
                             ("ullAvailPhys",            _ct.c_ulonglong),
                             ("ullTotalPageFile",        _ct.c_ulonglong),
                             ("ullAvailPageFile",        _ct.c_ulonglong),
                             ("ullTotalVirtual",         _ct.c_ulonglong),
                             ("ullAvailVirtual",         _ct.c_ulonglong),
                             ("ullAvailExtendedVirtual", _ct.c_ulonglong)]
            st = _MEMSTATUS()
            st.dwLength = _ct.sizeof(_MEMSTATUS)
            _ct.windll.kernel32.GlobalMemoryStatusEx(_ct.byref(st))
            avail_bytes = st.ullAvailPhys
        except Exception:
            try:
                with open("/proc/meminfo") as _f:
                    for _line in _f:
                        if _line.startswith("MemAvailable:"):
                            avail_bytes = int(_line.split()[1]) * 1024
                            break
            except Exception:
                avail_bytes = 2 * 1024**3   # assume 2 GB if unknown

    avail_gb = avail_bytes / (1024 ** 3)
    cpus     = max(1, _os.cpu_count() or 1)

    # ── GPU detection ────────────────────────────────────────────────────────
    has_gpu = _CUDF   # set earlier in optional-deps block

    # ── Tier selection ───────────────────────────────────────────────────────
    if avail_gb < 3.5:
        tier = "LOW"
    elif avail_gb < 14:
        tier = "MID"
    else:
        tier = "HIGH"

    # ── Tuning constants per tier ────────────────────────────────────────────
    #
    #  io_buffer      — OS read/write buffer for open()
    #  read_chunk     — bytes per mmap/read block in filter/index
    #  write_buf      — output file write buffer
    #  max_anchors    — sparse index anchor points (BigFileEngine)
    #  dense_limit    — file size below which dense (every-line) index is used
    #  workers        — ThreadPoolExecutor max_workers for batch operations
    #  index_workers  — workers for the BigFileEngine indexer
    #  bv_chunk_cap   — max rows loaded at once in Big Viewer
    #
    if tier == "LOW":
        return dict(
            tier         = "LOW",
            avail_gb     = round(avail_gb, 1),
            cpus         = cpus,
            has_gpu      = has_gpu,
            io_buffer    = 256  * 1024,           # 256 KB
            read_chunk   =   4  * 1024 * 1024,    # 4 MB
            write_buf    = 512  * 1024,           # 512 KB
            max_anchors  = 100_000,               # low RAM → fewer anchors
            dense_limit  =  64  * 1024 * 1024,   # 64 MB dense threshold
            workers      = min(cpus, 2),          # at most 2 parallel workers
            index_workers= 1,
            bv_chunk_cap = 200,                   # rows per page
        )
    elif tier == "MID":
        return dict(
            tier         = "MID",
            avail_gb     = round(avail_gb, 1),
            cpus         = cpus,
            has_gpu      = has_gpu,
            io_buffer    =   1  * 1024 * 1024,    # 1 MB
            read_chunk   =  32  * 1024 * 1024,    # 32 MB
            write_buf    =   8  * 1024 * 1024,    # 8 MB
            max_anchors  = 500_000,
            dense_limit  = 256  * 1024 * 1024,   # 256 MB
            workers      = cpus,
            index_workers= max(1, cpus // 2),
            bv_chunk_cap = 500,
        )
    else:  # HIGH
        return dict(
            tier         = "HIGH",
            avail_gb     = round(avail_gb, 1),
            cpus         = cpus,
            has_gpu      = has_gpu,
            io_buffer    =   4  * 1024 * 1024,    # 4 MB
            read_chunk   = 128  * 1024 * 1024,    # 128 MB
            write_buf    =  32  * 1024 * 1024,    # 32 MB
            max_anchors  = 2_000_000,             # more anchors → finer seek
            dense_limit  = 512  * 1024 * 1024,   # 512 MB dense threshold
            workers      = cpus,
            index_workers= cpus,
            bv_chunk_cap = 1000,
        )

_HW = _build_hw_profile()

# Apply adaptive values as module-level constants used throughout the code
_IO_BUFFER    = _HW["io_buffer"]
_READ_CHUNK   = _HW["read_chunk"]
_WRITE_BUF    = _HW["write_buf"]
_N_WORKERS    = _HW["workers"]

# Human-readable engine label shown in the UI
def _engine_label():
    hw = _HW
    gpu = "🟢 GPU (cudf) + " if hw["has_gpu"] else ""
    return (
        f"{gpu}🔵 CPU {hw['cpus']} cores  │  "
        f"RAM tier: {hw['tier']} ({hw['avail_gb']} GB free)  │  "
        f"Chunks: {hw['read_chunk'] // (1024*1024)} MB"
    )

def _fast_json_loads(s):
    """Use orjson when available (10x faster), fallback to json."""
    if _ORJSON:
        return orjson.loads(s)
    return json.loads(s)

def _fast_json_dumps(obj):
    """Use orjson when available (10x faster), fallback to json."""
    if _ORJSON:
        return orjson.dumps(obj, option=orjson.OPT_NON_STR_KEYS).decode()
    return json.dumps(obj, ensure_ascii=False)

def _fast_hash(data):
    """Ultra-fast hash using xxhash (C library), fallback to hash()."""
    if _XXHASH:
        if isinstance(data, str):
            return xxhash.xxh64_hexdigest(data.encode())
        return xxhash.xxh64_hexdigest(data)
    return str(hash(data))

# ── Windows DPI awareness (crisp text on HiDPI screens) ──────────────────────
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(2)   # per-monitor v2
except Exception:
    try: ctypes.windll.user32.SetProcessDPIAware()
    except Exception: pass

# ─────────────────────────────────────────────────────────────────────────────
APP_NAME         = "OMIO Ultra DB Studio"
APP_VERSION      = "4.0"

def _resource_path(rel: str) -> Path:
    """Resolve resource path – works in dev and PyInstaller --onefile bundles."""
    if hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS) / rel
    return Path(__file__).parent / rel

ICON_PATH = _resource_path("omio.ico")
SUPPORTED_IN     = {".csv", ".json", ".jsonl", ".txt", ".xlsx", ".tsv", ".log"}
SUPPORTED_IN_LABEL = "CSV · JSON · JSONL · XLSX · TXT · TSV · LOG"
SUPPORTED_OUT    = ("csv", "jsonl", "txt", "xlsx", "tsv")
CHUNK_SIZE_BYTES = 4 * 1024 * 1024   # 4 MB default view chunk for raw mode

# ── design tokens ─────────────────────────────────────────────────────────────
C = {
    "bg":       "#07021a",   # deep-space purple-black
    "panel":    "#0f0730",   # dark purple panel
    "panel2":   "#130a38",   # slightly lighter panel
    "border":   "#2a1060",   # purple border
    "accent":   "#00c8e8",   # bright cyan
    "accent2":  "#00838f",   # mid teal
    "danger":   "#ff4d5e",
    "warn":     "#ffaa00",
    "ok":       "#00e676",
    "text":     "#e0f7fa",   # light cyan-white
    "muted":    "#7c6fa0",   # muted purple
    "sel":      "#1a0d4a",
    "inp":      "#04011a",
    "btn":      "#120840",   # dark purple button
    "btn_h":    "#1e1060",   # hover lighter purple
    "head":     "#00c8e8",
    "tab_on":   "#130a38",
    "tab_off":  "#07021a",
    "prog":     "#00c8e8",
    "strip":    "#07021a",
}

FT = ("Segoe UI",  16, "bold")    # title
FH = ("Segoe UI",  11, "bold")    # section head
FB = ("Segoe UI",  10)            # body
FS = ("Segoe UI",   9)            # small / muted
FM = ("Consolas",  10)            # mono
FMS= ("Consolas",   9)            # mono small


# ═════════════════════════════════════════════════════════════════════════════
#  JOB CONTROL
# ═════════════════════════════════════════════════════════════════════════════
class JobControl:
    def __init__(self):
        self._pause  = threading.Event()
        self._cancel = threading.Event()
        self._pause.set()

    def pause(self):  self._pause.clear()
    def resume(self): self._pause.set()
    def cancel(self): self._cancel.set();  self._pause.set()
    def reset(self):  self._cancel.clear(); self._pause.set()

    def wait_if_paused(self): self._pause.wait()
    def is_paused(self):      return not self._pause.is_set()
    def is_cancelled(self):   return self._cancel.is_set()


# ═════════════════════════════════════════════════════════════════════════════
#  DATA PROCESSOR  – streaming engine
# ═════════════════════════════════════════════════════════════════════════════
class DataProcessor:

    # ── low-level helpers ────────────────────────────────────────────────────
    @staticmethod
    def _detect_delim(path: Path):
        ext = path.suffix.lower()
        if ext in (".txt", ".tsv", ".log"): return "\t"
        # sniff first line for delimiter
        try:
            with path.open("r", encoding="utf-8", errors="replace") as f:
                sample = f.read(4096)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except Exception:
            return ","

    @staticmethod
    def _clean(row: dict, trim=True) -> dict:
        out = {}
        for k, v in row.items():
            if v is None: v = ""
            if trim and isinstance(v, str): v = v.strip()
            out[str(k)] = v
        return out

    @staticmethod
    def _is_empty(row: dict) -> bool:
        return all(str(v).strip() == "" for v in row.values())

    # ── file threat scanner (CDR — Content Disarm & Reconstruction) ──────────
    # Detects macros, VBA, embedded objects, formulas, scripts, suspicious
    # patterns. Returns (threats_list, severity: "clean"|"low"|"medium"|"high").
    # The data is ALWAYS extracted as plain text rows — all active content stripped.

    _SUSPICIOUS_PATTERNS = [
        (re.compile(r'=\s*(CMD|EXEC|SYSTEM|SHELL)\s*\(', re.I),  "formula_injection",  "high",   "Formula injection (=CMD/EXEC/SHELL)"),
        (re.compile(r'=\s*HYPERLINK\s*\(',               re.I),  "formula_hyperlink",  "medium", "Formula HYPERLINK injection"),
        (re.compile(r'=\s*IMPORTXML\s*\(',               re.I),  "formula_import",     "medium", "Formula IMPORTXML data exfil"),
        (re.compile(r"<script[\s>]",                     re.I),  "script_tag",         "high",   "Embedded <script> tag"),
        (re.compile(r"javascript\s*:",                   re.I),  "javascript_uri",     "high",   "javascript: URI"),
        (re.compile(r"on(load|error|click|mouseover)\s*=", re.I), "event_handler",     "medium", "HTML event handler attribute"),
        (re.compile(r"<iframe[\s>]",                     re.I),  "iframe_tag",         "medium", "Embedded <iframe> tag"),
        (re.compile(r"<object[\s>]",                     re.I),  "object_tag",         "medium", "Embedded <object> tag"),
        (re.compile(r"<embed[\s>]",                      re.I),  "embed_tag",          "medium", "Embedded <embed> tag"),
        (re.compile(r"\bpowershell\b.*-e(nc(odedcommand)?)?", re.I), "powershell",    "high",   "PowerShell encoded command"),
        (re.compile(r"\bcmd\s*/c\b",                     re.I),  "cmd_exec",           "high",   "cmd.exe /c execution"),
        (re.compile(r"\bwscript\b|\bcscript\b",          re.I),  "wscript",            "high",   "WScript/CScript execution"),
        (re.compile(r"\\x[0-9a-f]{2}.*\\x[0-9a-f]{2}",  re.I),  "hex_obfuscation",   "low",    "Hex-encoded obfuscation"),
    ]

    _FORMULA_DANGER = re.compile(r"^[=+\-@]\s*(CMD|EXEC|SYSTEM|SHELL|DDE|HYPERLINK|IMPORTXML)\s*\(", re.I)
    _FORMULA_PREFIX = re.compile(r"^[=+\-@]")

    @staticmethod
    def scan_file(path, log=None, sample_bytes=2*1024*1024):
        """Scan a file for threats. Returns dict with findings."""
        path = Path(path)
        ext = path.suffix.lower()
        sz = path.stat().st_size
        threats = []
        severity = "clean"
        sev_rank = {"clean": 0, "low": 1, "medium": 2, "high": 3}

        def _add(kind, level, detail, location=""):
            nonlocal severity
            threats.append({"type": kind, "severity": level,
                            "detail": detail, "location": location})
            if sev_rank.get(level, 0) > sev_rank.get(severity, 0):
                severity = level

        # ── XLSX-specific: macros, external links, VBA ──────────────────────
        if ext == ".xlsx":
            try:
                import zipfile
                with zipfile.ZipFile(str(path), "r") as zf:
                    names = zf.namelist()
                    # Check for VBA macros
                    vba_files = [n for n in names if "vbaProject" in n or n.endswith(".bin")]
                    if vba_files:
                        _add("vba_macro", "high", f"VBA macro project: {', '.join(vba_files)}", "xlsx_zip")
                    # Check for external links
                    ext_links = [n for n in names if "externalLink" in n]
                    if ext_links:
                        _add("external_link", "medium", f"External data links: {len(ext_links)}", "xlsx_zip")
                    # Check for ActiveX / embedded objects
                    activex = [n for n in names if "activeX" in n.lower()]
                    if activex:
                        _add("activex", "high", f"ActiveX controls: {len(activex)}", "xlsx_zip")
                    embeddings = [n for n in names if "embeddings" in n.lower()]
                    if embeddings:
                        _add("embedded_object", "medium", f"Embedded objects: {len(embeddings)}", "xlsx_zip")
            except Exception:
                pass

        # ── Sample first N bytes for pattern-based detection ────────────────
        try:
            raw_sample = b""
            with path.open("rb") as f:
                raw_sample = f.read(min(sz, sample_bytes))
            text_sample = raw_sample.decode("utf-8", errors="replace")

            for pat, kind, level, detail in DataProcessor._SUSPICIOUS_PATTERNS:
                matches = pat.findall(text_sample)
                if matches:
                    _add(kind, level, f"{detail} ({len(matches)} occurrence{'s' if len(matches)>1 else ''})")

            # Formula injection in cell values (CSV/TXT/TSV)
            if ext in {".csv", ".txt", ".tsv", ".log"}:
                formula_count = 0
                for line in text_sample.split("\n")[:5000]:
                    for cell in line.split(","):
                        cell = cell.strip().strip('"').strip("'")
                        if DataProcessor._FORMULA_DANGER.match(cell):
                            formula_count += 1
                if formula_count:
                    _add("formula_injection", "high",
                         f"Dangerous formula cells: {formula_count}")

        except Exception:
            pass

        # ── Check file magic bytes for disguised files ──────────────────────
        try:
            with path.open("rb") as f:
                magic = f.read(8)
            # EXE/DLL disguised as data file
            if magic[:2] == b"MZ":
                _add("disguised_exe", "high", "File has MZ (executable) magic bytes — disguised PE file")
            # ZIP bomb check (ratio)
            if ext in {".xlsx", ".json", ".csv"} and magic[:2] == b"PK":
                import zipfile
                try:
                    with zipfile.ZipFile(str(path), "r") as zf:
                        total_uncompressed = sum(i.file_size for i in zf.infolist())
                        if sz > 0 and total_uncompressed / sz > 100:
                            _add("zip_bomb", "high",
                                 f"Suspicious compression ratio: {total_uncompressed/sz:.0f}x")
                except Exception:
                    pass
        except Exception:
            pass

        return {
            "path": str(path),
            "size": sz,
            "threats": threats,
            "severity": severity,
            "threat_count": len(threats),
        }

    @staticmethod
    def _sanitize_value(val):
        """Strip dangerous prefixes and content from a cell value."""
        if not isinstance(val, str):
            return val
        # Remove formula injection prefixes
        s = val.strip()
        if s and s[0] in "=+-@" and DataProcessor._FORMULA_PREFIX.match(s):
            s = "'" + s   # prefix with quote to neutralize
        # Remove script tags, javascript URIs
        s = re.sub(r"<script[^>]*>.*?</script>", "", s, flags=re.I | re.S)
        s = re.sub(r"javascript\s*:", "", s, flags=re.I)
        return s

    @staticmethod
    def iter_rows_safe(path, log=None, sanitize=True):
        """Like iter_rows but with sanitization: strips formulas, scripts, macros.
        XLSX loaded with data_only=True (no macros), all cell values sanitized."""
        path = Path(path)
        ext = path.suffix.lower()

        # For XLSX: force data_only + read_only (strips macros/VBA entirely)
        if ext == ".xlsx":
            if not _OPENPYXL:
                raise RuntimeError("openpyxl required")
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(v).strip() if v is not None else f"col_{i}"
                              for i, v in enumerate(row)]
                    continue
                vals = ["" if v is None else v for v in row]
                d = {header[i]: (vals[i] if i < len(vals) else "")
                     for i in range(len(header))}
                if sanitize:
                    d = {k: DataProcessor._sanitize_value(v) for k, v in d.items()}
                yield d
            wb.close()
            return

        # For all other formats: use normal iter_rows + sanitize each value
        for row in DataProcessor.iter_rows(path, log=log):
            if sanitize:
                yield {k: DataProcessor._sanitize_value(v) for k, v in row.items()}
            else:
                yield row

    @staticmethod
    def convert_safe(in_path, out_path, fmt, trim=True, drop_empty=True,
                     sanitize=True, log=None, control=None,
                     on_progress=None, total=None):
        """Convert with CDR: scan → sanitize → write clean output."""
        in_path = Path(in_path)
        t0 = time.time(); proc = kept = 0

        # Phase 1: scan
        if log: log(f"🔍 Scanning: {in_path.name}")
        scan = DataProcessor.scan_file(in_path, log=log)
        sev = scan["severity"]
        sev_icons = {"clean": "✅", "low": "⚠️", "medium": "🔶", "high": "🔴"}
        if log:
            log(f"  Scan result: {sev_icons.get(sev, '?')} {sev.upper()} — "
                f"{scan['threat_count']} threat{'s' if scan['threat_count'] != 1 else ''}")
            for t in scan["threats"]:
                log(f"    [{t['severity'].upper()}] {t['detail']}")

        # Phase 2: stream with sanitization
        if log: log(f"🛡️ Converting with sanitization {'ON' if sanitize else 'OFF'}")

        def _stream():
            nonlocal proc, kept
            for raw in DataProcessor.iter_rows_safe(in_path, log=log, sanitize=sanitize):
                if control and control.is_cancelled(): return
                if control: control.wait_if_paused()
                proc += 1
                row = DataProcessor._clean(raw, trim)
                if drop_empty and DataProcessor._is_empty(row): continue
                kept += 1
                if on_progress and proc % 5000 == 0:
                    on_progress(proc, kept, total, time.time() - t0)
                yield row

        n = DataProcessor.write_rows(_stream(), out_path, fmt, log=log)
        elapsed = time.time() - t0
        if on_progress: on_progress(proc, kept, total, elapsed)
        if log:
            log(f"✓ Safe convert: {proc:,} read → {n:,} written │ "
                f"{scan['threat_count']} threats stripped │ {elapsed:.1f}s")
        return n, scan

    # ── streaming row reader ─────────────────────────────────────────────────
    @staticmethod
    def iter_rows(path, log=None):
        path = Path(path)
        ext  = path.suffix.lower()
        if ext not in SUPPORTED_IN:
            raise ValueError(f"Unsupported: {ext}")

        if ext in {".csv", ".txt", ".tsv", ".log"}:
            delim = DataProcessor._detect_delim(path)
            with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
                yield from csv.DictReader(f, delimiter=delim)
            return

        if ext == ".jsonl":
            with path.open("rb", buffering=_IO_BUFFER) as f:
                for ln, raw in enumerate(f, 1):
                    raw = raw.strip()
                    if not raw: continue
                    try:
                        obj = _fast_json_loads(raw)
                    except (json.JSONDecodeError, ValueError):
                        if log: log(f"[WARN] bad JSONL line {ln}")
                        continue
                    if isinstance(obj, dict): yield obj
            return

        if ext == ".json":
            if _IJSON:
                yielded = False
                try:
                    with path.open("rb") as f:
                        for item in ijson.items(f, "item"):
                            if isinstance(item, dict):
                                yielded = True
                                yield item
                except Exception:
                    pass
                if yielded: return
            with path.open("rb", buffering=_IO_BUFFER) as f:
                try:
                    data = _fast_json_loads(f.read())
                except (json.JSONDecodeError, ValueError):
                    if log: log(f"[ERR] invalid JSON: {path.name}")
                    return
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict): yield item
            elif isinstance(data, dict):
                yield data
            return

        if ext == ".xlsx":
            if not _OPENPYXL:
                raise RuntimeError("openpyxl required")
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(row)]
                    continue
                vals = ["" if v is None else v for v in row]
                yield {header[i]: (vals[i] if i < len(vals) else "") for i in range(len(header))}
            wb.close()

    # ── writer ───────────────────────────────────────────────────────────────
    @staticmethod
    def write_rows(rows, out_path, fmt, log=None):
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        if fmt in ("csv", "txt", "tsv"):
            delim = "\t" if fmt in ("txt", "tsv") else ","
            first = next(rows, None)
            if first is None:
                out_path.write_text("", encoding="utf-8"); return 0
            fields = list(first.keys())
            with out_path.open("w", encoding="utf-8", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fields, delimiter=delim, extrasaction="ignore")
                w.writeheader(); w.writerow(first); count = 1
                for row in rows:
                    w.writerow({k: row.get(k, "") for k in fields}); count += 1
            return count

        if fmt == "jsonl":
            count = 0
            with out_path.open("w", encoding="utf-8") as f:
                for row in rows:
                    f.write(_fast_json_dumps(row) + "\n"); count += 1
            return count

        if fmt == "xlsx":
            if not _OPENPYXL: raise RuntimeError("openpyxl required")
            first = next(rows, None)
            if first is None:
                wb = Workbook(write_only=True); wb.create_sheet("data"); wb.save(str(out_path)); return 0
            headers = list(first.keys())
            wb = Workbook(write_only=True); ws = wb.create_sheet("data")
            ws.append(headers); ws.append([first.get(h, "") for h in headers]); count = 1
            for row in rows:
                ws.append([row.get(h, "") for h in headers]); count += 1
            wb.save(str(out_path)); return count

        raise ValueError(f"Unknown format: {fmt}")

    # ── high-level ops ───────────────────────────────────────────────────────
    @staticmethod
    def count_rows(path, log=None, control=None):
        n = 0
        for _ in DataProcessor.iter_rows(path, log=log):
            if control and control.is_cancelled(): break
            if control: control.wait_if_paused()
            n += 1
        return n

    @staticmethod
    def convert(in_path, out_path, fmt, trim=True, drop_empty=True,
                log=None, control=None, on_progress=None, total=None):
        t0 = time.time(); proc = kept = 0
        def _stream():
            nonlocal proc, kept
            for raw in DataProcessor.iter_rows(in_path, log=log):
                if control and control.is_cancelled(): return
                if control: control.wait_if_paused()
                proc += 1
                row = DataProcessor._clean(raw, trim)
                if drop_empty and DataProcessor._is_empty(row): continue
                kept += 1
                if on_progress and proc % 5000 == 0:
                    on_progress(proc, kept, total, time.time()-t0)
                yield row
        n = DataProcessor.write_rows(_stream(), out_path, fmt, log=log)
        if on_progress: on_progress(proc, kept, total, time.time()-t0)
        return n

    @staticmethod
    def split(in_path, out_dir, rps, fmt="csv", trim=True, drop_empty=True,
              log=None, control=None, on_progress=None, total=None):
        in_path = Path(in_path); out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)
        base = in_path.stem; t0 = time.time()
        part = 1; batch = []; proc = kept = 0; files = []
        if log: log(f"Split: {in_path.name} → {rps:,} rows/file, format={fmt}")
        if log: log(f"Split: Output folder → {out_dir}")
        for raw in DataProcessor.iter_rows(in_path, log=log):
            if control and control.is_cancelled(): break
            if control: control.wait_if_paused()
            proc += 1
            row = DataProcessor._clean(raw, trim)
            if drop_empty and DataProcessor._is_empty(row): continue
            batch.append(row); kept += 1
            if on_progress and proc % 10000 == 0:
                on_progress(proc, kept, total, time.time()-t0)
            if len(batch) >= rps:
                p = out_dir / f"{base}_part{part:04d}.{fmt}"
                n_written = DataProcessor.write_rows(iter(batch), p, fmt, log=None)
                files.append(str(p))
                if log: log(f"  ✓ Part {part}: {p.name} ({n_written:,} rows)")
                batch = []; part += 1
        if batch:
            p = out_dir / f"{base}_part{part:04d}.{fmt}"
            n_written = DataProcessor.write_rows(iter(batch), p, fmt, log=None)
            files.append(str(p))
            if log: log(f"  ✓ Part {part}: {p.name} ({n_written:,} rows — remainder)")
        elapsed = time.time() - t0
        if log:
            log(f"Split complete: {proc:,} rows read │ {kept:,} kept │ "
                f"{len(files)} files created │ {elapsed:.1f}s")
        if on_progress: on_progress(proc, kept, total, elapsed)
        return proc, files

    @staticmethod
    def merge(in_files, out_path, fmt="csv", log=None, control=None,
              on_progress=None, total=None):
        if not in_files: raise ValueError("No files to merge")
        t0 = time.time(); proc = 0
        def _stream():
            nonlocal proc
            for fp in in_files:
                if log: log(f"  Merging: {Path(fp).name}")
                for row in DataProcessor.iter_rows(fp, log=log):
                    if control and control.is_cancelled(): return
                    if control: control.wait_if_paused()
                    proc += 1
                    if on_progress and proc % 5000 == 0:
                        on_progress(proc, proc, total, time.time()-t0)
                    yield row
        n = DataProcessor.write_rows(_stream(), out_path, fmt, log=log)
        if on_progress: on_progress(proc, proc, total, time.time()-t0)
        return n

    # ── analysis ─────────────────────────────────────────────────────────────
    @staticmethod
    def profile_columns(path, sample=10000, log=None):
        stats = {}
        for i, row in enumerate(DataProcessor.iter_rows(path, log=log)):
            if i >= sample: break
            for col, val in row.items():
                if col not in stats:
                    stats[col] = {"count":0,"empty":0,"int":0,"float":0,
                                  "date":0,"bool":0,"text":0,
                                  "min_len":None,"max_len":0,"sample":set()}
                s = stats[col]; s["count"] += 1
                v = str(val).strip()
                if v == "": s["empty"] += 1; continue
                l = len(v)
                s["min_len"] = l if s["min_len"] is None else min(s["min_len"], l)
                s["max_len"] = max(s["max_len"], l)
                if len(s["sample"]) < 5: s["sample"].add(v)
                try: int(v);   s["int"]   += 1; continue
                except ValueError: pass
                try: float(v); s["float"] += 1; continue
                except ValueError: pass
                if v.lower() in {"true","false","yes","no","1","0"}: s["bool"] += 1; continue
                if re.match(r"\d{4}[-/]\d{2}[-/]\d{2}", v): s["date"] += 1; continue
                s["text"] += 1
        return stats

    @staticmethod
    def filter_rows(path, conditions, log=None, control=None):
        OPS = {
            "==":          lambda rv, val: rv.lower() == val.lower(),
            "!=":          lambda rv, val: rv.lower() != val.lower(),
            "contains":    lambda rv, val: val.lower() in rv.lower(),
            "not_contains":lambda rv, val: val.lower() not in rv.lower(),
            "starts":      lambda rv, val: rv.lower().startswith(val.lower()),
            "ends":        lambda rv, val: rv.lower().endswith(val.lower()),
            ">":           lambda rv, val: _num_cmp(rv, val, ">"),
            "<":           lambda rv, val: _num_cmp(rv, val, "<"),
            ">=":          lambda rv, val: _num_cmp(rv, val, ">="),
            "<=":          lambda rv, val: _num_cmp(rv, val, "<="),
            "is_empty":    lambda rv, val: rv.strip() == "",
            "not_empty":   lambda rv, val: rv.strip() != "",
            "regex":       lambda rv, val: bool(re.search(val, rv)),
        }
        def check(row):
            for cond in conditions:
                col = cond.get("col",""); op = cond.get("op","=="); val = str(cond.get("val",""))
                rv  = str(row.get(col,""))
                fn  = OPS.get(op)
                if fn is None: continue
                try:
                    if not fn(rv, val): return False
                except Exception:
                    return False
            return True

        for row in DataProcessor.iter_rows(path, log=log):
            if control and control.is_cancelled(): break
            if control: control.wait_if_paused()
            if check(row): yield row

    @staticmethod
    def find_duplicates(path, key_cols, log=None, control=None, on_progress=None):
        seen = {}; total = 0; t0 = time.time()
        for row in DataProcessor.iter_rows(path, log=log):
            if control and control.is_cancelled(): break
            if control: control.wait_if_paused()
            total += 1
            key = _fast_hash("|".join(str(row.get(c,"")) for c in key_cols))
            seen[key] = seen.get(key, 0) + 1
            if on_progress and total % 10000 == 0:
                dup_cnt = sum(1 for v in seen.values() if v > 1)
                on_progress(total, dup_cnt, None, time.time()-t0)
        dup_keys = {k for k, v in seen.items() if v > 1}
        unique   = sum(1 for v in seen.values() if v == 1)
        return seen, dup_keys, unique, total


def _num_cmp(a, b, op):
    try: a, b = float(a), float(b)
    except ValueError: pass
    return eval(f"a {op} b", {"a": a, "b": b})


def _hex_darken(hex_c: str, factor: float) -> str:
    """Return darker version of hex color (factor 0=unchanged, 1=black)."""
    h = hex_c.lstrip('#')
    r, g, b = int(h[:2], 16), int(h[2:4], 16), int(h[4:], 16)
    return f"#{int(r*(1-factor)):02x}{int(g*(1-factor)):02x}{int(b*(1-factor)):02x}"


def _hex_brighten(hex_c: str, factor: float) -> str:
    """Return brighter version of hex color (factor 0=unchanged, 1=white)."""
    h = hex_c.lstrip('#')
    r, g, b = int(h[:2], 16), int(h[2:4], 16), int(h[4:], 16)
    return (f"#{min(255,int(r+(255-r)*factor)):02x}"
            f"{min(255,int(g+(255-g)*factor)):02x}"
            f"{min(255,int(b+(255-b)*factor)):02x}")


# ═════════════════════════════════════════════════════════════════════════════
#  BIG FILE ENGINE  –  EmEditor-style byte-seek navigation
# ═════════════════════════════════════════════════════════════════════════════
class BigFileEngine:
    """
    Ultra-fast file engine for files up to 1 TB+.
    
    Architecture:
    • Sparse line index — for files >256 MB, only stores every Nth line offset
      (adaptive stride: ~500K anchor points max) to stay under 8 MB RAM.
    • Dense index for files <=256 MB (every line offset stored).
    • mmap-based newline scanning at C speed.
    • O(1) seek-to-anchor + short forward scan for any line.
    • Background indexing with live progress — GUI never blocks.
    • Chunk-based read — never loads more than a few MB at a time.
    • Binary-search offset_to_line for instant line lookup.
    """

    # Class-level constants driven by the adaptive HW profile
    _CACHE_CAP    = 5_000_000
    _INDEX_BUF    = _HW["read_chunk"]          # mmap window = adaptive read chunk
    _DENSE_LIMIT  = _HW["dense_limit"]         # adaptive dense/sparse threshold
    _MAX_ANCHORS  = _HW["max_anchors"]         # adaptive anchor count
    _SEARCH_BLOCK = max(16 * 1024 * 1024,
                        _HW["read_chunk"] // 2)  # search block ≥ 16 MB
    _READ_CHUNK   = _HW["read_chunk"]          # forward-scan read size

    def __init__(self, path: str):
        self.path      = Path(path)
        self.ext       = self.path.suffix.lower()
        self.size      = self.path.stat().st_size
        self.encoding  = "utf-8"
        self._offsets  = []           # byte offset of line starts (dense or sparse)
        self._stride   = 1            # 1 = dense, N = every Nth line stored
        self._indexed  = False
        self._index_lock = threading.Lock()
        self._stop_index = threading.Event()
        self._row_cache_headers: list = []
        self._row_cache_rows:    list = []
        self._row_cache_ready  = False
        self._row_cache_progress: int = 0
        self._stop_cache = threading.Event()
        self._total_lines = 0
        self._index_pct   = 0

    # ── file info ─────────────────────────────────────────────────────────
    def file_info(self):
        sz = self.size
        if sz < 1024: label = f"{sz} B"
        elif sz < 1024**2: label = f"{sz/1024:.1f} KB"
        elif sz < 1024**3: label = f"{sz/1024**2:.1f} MB"
        else:              label = f"{sz/1024**3:.2f} GB"
        return {
            "path": str(self.path),
            "name": self.path.name,
            "size": label,
            "size_bytes": sz,
            "ext": self.ext,
            "indexed": self._indexed,
            "lines_indexed": len(self._offsets),
            "total_lines": self._total_lines,
        }

    # ── fast mmap-based line indexer (sparse for TB-scale) ──────────────
    def start_index(self, on_progress=None):
        """Build byte-offset index using mmap.

        Dense (stride=1) for files <=256 MB — every line offset stored.
        Sparse (stride=N) for larger files — ~500 K anchor points max,
        keeping RAM under ~16 MB even for 1 TB+ files.

        TB-safe algorithm:
          • Sample 1 MB to estimate average line length → pre-compute stride
          • Single-pass mmap scan: store only every stride-th offset
          • No temporary all_lines buffer — O(MAX_ANCHORS) RAM throughout
          • If sample over-estimated (too many anchors collected), re-sample in place
        """
        self._stop_index.clear()
        def _run():
            fsize = self.size
            use_sparse = fsize > self._DENSE_LIMIT
            offsets = [0]
            line_count = 0
            stride = 1

            if fsize == 0:
                with self._index_lock:
                    self._offsets = offsets
                    self._stride = 1
                    self._indexed = True
                    self._total_lines = 1
                if on_progress:
                    on_progress(fsize, fsize, 1)
                return

            # ── Pre-compute stride for sparse mode (sample-based, no OOM) ──
            if use_sparse:
                sample_sz = min(1_048_576, fsize)   # 1 MB sample
                try:
                    with open(self.path, "rb") as _sf:
                        _sample = _sf.read(sample_sz)
                    nl_in_sample = _sample.count(b"\n") or 1
                    avg_line     = sample_sz / nl_in_sample
                    estimated    = max(self._MAX_ANCHORS * 2, int(fsize / avg_line))
                except OSError:
                    estimated    = max(self._MAX_ANCHORS * 2, fsize // 80)
                stride = max(1, estimated // self._MAX_ANCHORS)

            # ── Main index pass (mmap — C-speed newline scanning) ────────────
            try:
                with open(self.path, "rb") as f:
                    fd = f.fileno()
                    with mmap.mmap(fd, 0, access=mmap.ACCESS_READ) as mm:
                        pos = 0
                        while pos < fsize and not self._stop_index.is_set():
                            idx = mm.find(b"\n", pos)
                            if idx == -1:
                                break
                            line_count += 1
                            # Dense: store every line; Sparse: every stride-th
                            if stride == 1 or line_count % stride == 0:
                                offsets.append(idx + 1)
                            pos = idx + 1
                            if line_count % 1_000_000 == 0:
                                self._total_lines  = line_count
                                self._index_pct    = int(pos / max(fsize, 1) * 100)
                                if on_progress:
                                    on_progress(pos, fsize, line_count)

                # Guard: if sample over-estimated (collected too many anchors),
                # re-sample in place to stay under MAX_ANCHORS — O(MAX_ANCHORS)
                if use_sparse and len(offsets) > self._MAX_ANCHORS:
                    step    = max(2, len(offsets) // self._MAX_ANCHORS)
                    offsets = offsets[::step]
                    stride  = stride * step

            except (OSError, ValueError):
                # Fallback: chunked read for network/locked files (no mmap)
                offsets    = [0]
                line_count = 0
                stride     = 1
                buf_size   = 16 * 1024 * 1024
                pos        = 0
                with self.path.open("rb") as f:
                    while not self._stop_index.is_set():
                        buf = f.read(buf_size)
                        if not buf:
                            break
                        chunk_start = 0
                        while True:
                            nl = buf.find(b"\n", chunk_start)
                            if nl == -1:
                                break
                            line_count += 1
                            offsets.append(pos + nl + 1)
                            chunk_start = nl + 1
                        pos += len(buf)
                        if line_count % 500_000 == 0:
                            self._total_lines = line_count
                            if on_progress:
                                on_progress(pos, fsize, line_count)
                # Reduce fallback offsets to MAX_ANCHORS if needed
                if len(offsets) > self._MAX_ANCHORS:
                    step    = max(2, len(offsets) // self._MAX_ANCHORS)
                    offsets = offsets[::step]
                    stride  = step

            with self._index_lock:
                self._offsets     = offsets
                self._stride      = stride
                self._indexed     = True
                self._total_lines = line_count if line_count > 0 else len(offsets)
            if on_progress:
                on_progress(fsize, fsize, self._total_lines)
        threading.Thread(target=_run, daemon=True).start()

    def stop_index(self):
        self._stop_index.set()

    def indexed_lines(self):
        return self._total_lines

    def _seek_to_line(self, target_line: int) -> int:
        """Get byte offset for a given 0-based line number.
        
        Dense index: direct lookup O(1).
        Sparse index: seek to nearest anchor, then scan forward.
        """
        if not self._indexed or not self._offsets:
            return 0
        if self._stride == 1:
            # Dense — direct lookup
            idx = min(target_line, len(self._offsets) - 1)
            return self._offsets[max(0, idx)]
        # Sparse — find nearest anchor before target_line
        anchor_idx = target_line // self._stride
        anchor_idx = min(anchor_idx, len(self._offsets) - 1)
        anchor_offset = self._offsets[anchor_idx]
        lines_to_skip = target_line - (anchor_idx * self._stride)
        if lines_to_skip <= 0:
            return anchor_offset
        # Forward scan from anchor
        with self.path.open("rb") as f:
            f.seek(anchor_offset)
            for _ in range(lines_to_skip):
                line = f.readline()
                if not line:
                    break
            return f.tell()

    # ── row-data cache (for XLSX / JSON) ─────────────────────────────────
    def start_row_cache(self, on_progress=None):
        """Stream ALL rows into memory in a background thread for instant seek."""
        self._stop_cache.clear()
        self._row_cache_ready   = False
        self._row_cache_headers = []
        self._row_cache_rows    = []
        ext = self.ext

        def _build():
            headers = []
            rows    = []
            try:
                if ext == ".xlsx":
                    if not _OPENPYXL: return
                    wb = load_workbook(str(self.path), read_only=True, data_only=True)
                    ws = wb.active
                    hdr = None
                    for ri, row in enumerate(ws.iter_rows(values_only=True)):
                        if self._stop_cache.is_set(): break
                        if hdr is None:
                            hdr = [str(v) if v is not None else f"col_{i}"
                                   for i, v in enumerate(row)]
                            headers = hdr
                            continue
                        vals = ["" if v is None else v for v in row]
                        rows.append({hdr[i]: (vals[i] if i < len(vals) else "")
                                     for i in range(len(hdr))})
                        if len(rows) % 5_000 == 0:
                            self._row_cache_progress = len(rows)
                            if on_progress: on_progress(len(rows), self.size)
                        if len(rows) >= self._CACHE_CAP: break
                    wb.close()

                elif ext == ".json":
                    if _IJSON:
                        try:
                            with self.path.open("rb") as f:
                                for item in ijson.items(f, "item"):
                                    if self._stop_cache.is_set(): break
                                    if isinstance(item, dict):
                                        rows.append(item)
                                        for k in item:
                                            if k not in headers: headers.append(k)
                                    if len(rows) % 5_000 == 0:
                                        self._row_cache_progress = len(rows)
                                        if on_progress: on_progress(len(rows), self.size)
                                    if len(rows) >= self._CACHE_CAP: break
                        except Exception:
                            pass
                    if not rows:  # fallback: load full JSON
                        with self.path.open("r", encoding=self.encoding, errors="replace") as f:
                            try:
                                data = json.load(f)
                            except Exception:
                                data = []
                        if isinstance(data, list):
                            for item in data[:self._CACHE_CAP]:
                                if isinstance(item, dict):
                                    rows.append(item)
                                    for k in item:
                                        if k not in headers: headers.append(k)
                        elif isinstance(data, dict):
                            rows.append(data)
                            headers = list(data.keys())

            except Exception:
                pass

            self._row_cache_headers = headers
            self._row_cache_rows    = rows
            self._row_cache_ready   = True
            self._row_cache_progress = len(rows)
            if on_progress: on_progress(len(rows), self.size)

        threading.Thread(target=_build, daemon=True).start()

    def stop_cache(self):
        self._stop_cache.set()

    # ── raw line access by index ─────────────────────────────────────────
    def read_raw_lines(self, start_line: int, count: int) -> list:
        """Read `count` raw text lines starting at 0-based line index.
        Uses sparse/dense index for O(1) seek (or short forward scan for sparse).
        """
        lines = []
        with self.path.open("rb") as f:
            if self._indexed:
                offset = self._seek_to_line(start_line)
                f.seek(offset)
                for _ in range(count):
                    raw = f.readline()
                    if not raw: break
                    lines.append(raw.decode(self.encoding, errors="replace").rstrip("\n\r"))
            else:
                # fallback: scan from start (slow for large files without index)
                for i, raw in enumerate(f):
                    if i < start_line: continue
                    if i >= start_line + count: break
                    lines.append(raw.decode(self.encoding, errors="replace").rstrip("\n\r"))
        return lines

    # ── structured row access ────────────────────────────────────────────
    def read_rows(self, start_row: int, count: int):
        """Return (headers, rows_list) for `count` data rows starting at start_row (1-based)."""
        ext = self.ext
        start_row = max(1, start_row)

        if ext in {".csv", ".txt", ".tsv", ".log"}:
            return self._read_csv_rows(start_row, count)
        if ext == ".jsonl":
            return self._read_jsonl_rows(start_row, count)
        if ext == ".json":
            return self._read_json_rows(start_row, count)
        if ext == ".xlsx":
            return self._read_xlsx_rows(start_row, count)
        return [], []

    def _read_csv_rows(self, start_row, count):
        delim = DataProcessor._detect_delim(self.path)
        headers = []
        rows = []
        end = start_row + count - 1

        with self.path.open("rb") as raw_f:
            # Seek to start of header line (always line 0)
            raw_f.seek(0)
            text_f = io.TextIOWrapper(raw_f, encoding=self.encoding, errors="replace", newline="")
            reader = csv.DictReader(text_f, delimiter=delim)
            headers = reader.fieldnames or []

            # Use offset index if available for fast seek
            if self._indexed:
                # line 0 = header, line N = data row N
                offset = self._seek_to_line(start_row)
                raw_f.seek(offset)
                text_f = io.TextIOWrapper(raw_f, encoding=self.encoding, errors="replace", newline="")
                sub_reader = csv.DictReader(text_f, fieldnames=headers, delimiter=delim)
                for i, row in enumerate(sub_reader):
                    if i >= count: break
                    rows.append(dict(row))
            else:
                for i, row in enumerate(reader):
                    if i < start_row - 1: continue
                    if i >= end: break
                    rows.append(dict(row))

        return list(headers) if headers else [], rows

    def _read_jsonl_rows(self, start_row, count):
        headers = []
        rows = []
        end = start_row + count - 1
        if self._indexed:
            offset = self._seek_to_line(start_row - 1)
            with self.path.open("rb") as rf:
                rf.seek(offset)
                tf = io.TextIOWrapper(rf, encoding=self.encoding, errors="replace")
                for i in range(count):
                    line = tf.readline()
                    if not line: break
                    line = line.strip()
                    if not line: continue
                    try:
                        obj = _fast_json_loads(line)
                        if isinstance(obj, dict):
                            rows.append(obj)
                            for k in obj:
                                if k not in headers: headers.append(k)
                    except Exception:
                        pass
        else:
            with self.path.open("r", encoding=self.encoding, errors="replace") as f:
                for ln, line in enumerate(f, 1):
                    if ln < start_row: continue
                    if ln > end: break
                    line = line.strip()
                    if not line: continue
                    try:
                        obj = _fast_json_loads(line)
                        if isinstance(obj, dict):
                            rows.append(obj)
                            for k in obj:
                                if k not in headers: headers.append(k)
                    except Exception:
                        pass
        return headers, rows

    def _read_json_rows(self, start_row, count):
        # ── fast path: row cache ready ────────────────────────────────────
        if self._row_cache_ready:
            h = self._row_cache_headers
            r = self._row_cache_rows[start_row - 1 : start_row - 1 + count]
            return h, r
        # ── slow path: ijson forward scan ─────────────────────────────────
        headers = []
        rows = []
        if _IJSON:
            try:
                with self.path.open("rb") as f:
                    gen = ijson.items(f, "item")
                    for i, item in enumerate(gen, 1):
                        if i < start_row: continue
                        if i >= start_row + count: break
                        if isinstance(item, dict):
                            rows.append(item)
                            for k in item:
                                if k not in headers: headers.append(k)
                return headers, rows
            except Exception:
                pass
        # fallback
        with self.path.open("r", encoding=self.encoding, errors="replace") as f:
            try:
                data = json.load(f)
            except Exception:
                return [], []
        if isinstance(data, list):
            slice_ = data[start_row-1 : start_row-1+count]
            for item in slice_:
                if isinstance(item, dict):
                    rows.append(item)
                    for k in item:
                        if k not in headers: headers.append(k)
        elif isinstance(data, dict):
            if start_row == 1:
                rows.append(data)
                headers = list(data.keys())
        return headers, rows

    def _read_xlsx_rows(self, start_row, count):
        # ── fast path: row cache ready ────────────────────────────────────
        if self._row_cache_ready:
            h = self._row_cache_headers
            r = self._row_cache_rows[start_row - 1 : start_row - 1 + count]
            return h, r
        # ── slow path: cache not ready yet, scan forward ──────────────────
        if not _OPENPYXL: return [], []
        headers = []
        rows = []
        wb = load_workbook(str(self.path), read_only=True, data_only=True)
        ws = wb.active
        header = None
        end = start_row + count - 1
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if header is None:
                header = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(row)]
                headers = header
                continue
            data_row = ri
            if data_row < start_row: continue
            if data_row > end: break
            vals = ["" if v is None else v for v in row]
            rows.append({header[i]: (vals[i] if i < len(vals) else "") for i in range(len(header))})
        wb.close()
        return headers, rows

    # ── raw bytes chunk ──────────────────────────────────────────────────
    def read_raw_bytes(self, offset: int, size: int) -> str:
        with self.path.open("rb") as f:
            f.seek(offset)
            data = f.read(size)
        return data.decode(self.encoding, errors="replace")

    def read_hex(self, offset: int, size: int) -> str:
        """Read raw bytes and return hex dump (like hex editors)."""
        with self.path.open("rb") as f:
            f.seek(offset)
            data = f.read(size)
        lines = []
        for i in range(0, len(data), 16):
            chunk = data[i:i+16]
            hex_part = " ".join(f"{b:02x}" for b in chunk)
            ascii_part = "".join(chr(b) if 32 <= b < 127 else "." for b in chunk)
            lines.append(f"{offset+i:08x}  {hex_part:<48}  |{ascii_part}|")
        return "\n".join(lines)

    # ── search ───────────────────────────────────────────────────────────
    def search_forward(self, needle: str, from_offset: int, case_sensitive=True, use_regex=False) -> tuple:
        """Return (byte_offset, line_number) of next match after from_offset, or (-1, -1)."""
        block = self._SEARCH_BLOCK
        if use_regex:
            flags = 0 if case_sensitive else re.IGNORECASE
            try:
                pat = re.compile(needle.encode(self.encoding, errors="replace"), flags)
            except re.error:
                return -1, -1
        else:
            needle_b = needle.encode(self.encoding, errors="replace")
            if not case_sensitive:
                needle_b = needle_b.lower()

        with self.path.open("rb") as f:
            f.seek(max(0, from_offset))
            overlap = b""
            pos = from_offset
            while True:
                buf = f.read(block)
                if not buf:
                    return -1, -1
                data = overlap + buf
                if use_regex:
                    m = pat.search(data)
                    if m:
                        found_off = pos - len(overlap) + m.start()
                        ln = self._offset_to_line(found_off)
                        return found_off, ln
                else:
                    search_data = data.lower() if not case_sensitive else data
                    idx = search_data.find(needle_b)
                    if idx != -1:
                        found_off = pos - len(overlap) + idx
                        ln = self._offset_to_line(found_off)
                        return found_off, ln
                needle_len = 256 if use_regex else len(needle_b)
                overlap = data[-max(needle_len, 256):]
                pos += len(buf)

    def search_backward(self, needle: str, from_offset: int, case_sensitive=True) -> tuple:
        """Return (byte_offset, line_number) of previous match before from_offset, or (-1, -1)."""
        block = self._SEARCH_BLOCK
        needle_b = needle.encode(self.encoding, errors="replace")
        if not case_sensitive:
            needle_b = needle_b.lower()
        pos = max(0, from_offset - block)
        while pos >= 0:
            with self.path.open("rb") as f:
                f.seek(pos)
                data = f.read(min(block + len(needle_b), from_offset - pos))
            search_data = data.lower() if not case_sensitive else data
            idx = search_data.rfind(needle_b)
            if idx != -1:
                found_off = pos + idx
                ln = self._offset_to_line(found_off)
                return found_off, ln
            if pos == 0:
                break
            pos = max(0, pos - block)
        return -1, -1

    def search_count(self, needle: str, case_sensitive=True, use_regex=False) -> int:
        """Count total occurrences of needle in file."""
        count = 0
        block = self._SEARCH_BLOCK * 2  # 32 MB for counting
        if use_regex:
            flags = 0 if case_sensitive else re.IGNORECASE
            try:
                pat = re.compile(needle.encode(self.encoding, errors="replace"), flags)
            except re.error:
                return 0
        else:
            needle_b = needle.encode(self.encoding, errors="replace")
            if not case_sensitive:
                needle_b = needle_b.lower()

        with self.path.open("rb") as f:
            overlap = b""
            while True:
                buf = f.read(block)
                if not buf: break
                data = overlap + buf
                if use_regex:
                    count += len(pat.findall(data))
                else:
                    search_data = data.lower() if not case_sensitive else data
                    start = 0
                    while True:
                        idx = search_data.find(needle_b, start)
                        if idx == -1: break
                        count += 1
                        start = idx + 1
                needle_len = 256 if use_regex else len(needle_b)
                overlap = data[-max(needle_len, 256):]
        return count

    def _offset_to_line(self, offset: int) -> int:
        """Binary search offsets to find approximate line number for a byte offset.
        For sparse index, returns the anchor line * stride (approximate).
        """
        if not self._indexed or not self._offsets:
            return -1
        lo, hi = 0, len(self._offsets) - 1
        while lo <= hi:
            mid = (lo + hi) // 2
            if self._offsets[mid] <= offset:
                lo = mid + 1
            else:
                hi = mid - 1
        # hi is the index into offsets array, multiply by stride for real line num
        return max(0, hi * self._stride)

    def line_to_offset(self, line: int) -> int:
        """Get byte offset for a given 0-based line number. Uses sparse-aware seek."""
        if not self._indexed:
            return -1
        return self._seek_to_line(line)


# ═════════════════════════════════════════════════════════════════════════════
#  PRESETS
# ═════════════════════════════════════════════════════════════════════════════
PRESETS_FILE = Path(__file__).parent / "presets.json"

def load_presets():
    if PRESETS_FILE.exists():
        try: return json.loads(PRESETS_FILE.read_text(encoding="utf-8"))
        except Exception: pass
    return {}

def save_presets(data):
    PRESETS_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


# ═════════════════════════════════════════════════════════════════════════════
#  STYLED WIDGETS
# ═════════════════════════════════════════════════════════════════════════════
def _btn(parent, text, cmd=None, bg=None, fg=None, **kw):
    bg = bg or C["btn"]
    fg = fg or "#ffffff"
    b = tk.Button(parent, text=text, command=cmd,
                  bg=bg, fg=fg,
                  activebackground=C["btn_h"], activeforeground="#ffffff",
                  relief="flat", bd=0,
                  font=("Segoe UI", 10, "bold"),
                  padx=16, pady=7,
                  cursor="hand2",
                  highlightthickness=1,
                  highlightbackground=C["border"], **kw)
    b.bind("<Enter>", lambda e: b.configure(bg=C["btn_h"], highlightbackground=C["accent2"]))
    b.bind("<Leave>", lambda e: b.configure(bg=bg, highlightbackground=C["border"]))
    return b

def abtn(parent, text, cmd=None, **kw):
    b = tk.Button(parent, text=text, command=cmd,
                  bg=C["accent2"], fg="#ffffff",
                  activebackground="#00a0b0", activeforeground="#ffffff",
                  relief="flat", bd=0, font=("Segoe UI", 10, "bold"),
                  padx=16, pady=7, cursor="hand2",
                  highlightthickness=1, highlightbackground=C["accent2"], **kw)
    b.bind("<Enter>", lambda e: b.configure(bg="#00a0b0", highlightbackground=C["accent"]))
    b.bind("<Leave>", lambda e: b.configure(bg=C["accent2"], highlightbackground=C["border"]))
    return b

def dbtn(parent, text, cmd=None, **kw):
    b = tk.Button(parent, text=text, command=cmd,
                  bg="#cc1122", fg="#ffffff",
                  activebackground="#ff2233", activeforeground="#ffffff",
                  disabledforeground="#d0d0d0",
                  relief="flat", bd=0, font=("Segoe UI", 10, "bold"),
                  padx=16, pady=7, cursor="hand2",
                  highlightthickness=1, highlightbackground="#ff2233", **kw)
    b.bind("<Enter>", lambda e: b.configure(bg="#ff2233"))
    b.bind("<Leave>", lambda e: b.configure(bg="#cc1122"))
    return b

def wbtn(parent, text, cmd=None, **kw):
    b = tk.Button(parent, text=text, command=cmd,
                  bg="#b06000", fg="#ffffff",
                  activebackground="#d07800", activeforeground="#ffffff",
                  disabledforeground="#d0d0d0",
                  relief="flat", bd=0, font=("Segoe UI", 10, "bold"),
                  padx=16, pady=7, cursor="hand2",
                  highlightthickness=1, highlightbackground="#d07800", **kw)
    b.bind("<Enter>", lambda e: b.configure(bg="#d07800"))
    b.bind("<Leave>", lambda e: b.configure(bg="#b06000"))
    return b

def _lbl(parent, text, fg=None, font=FB, **kw):
    return tk.Label(parent, text=text, bg=C["panel"], fg=fg or C["text"], font=font, **kw)

def _mlbl(parent, text, fg=None, **kw):
    return tk.Label(parent, text=text, bg=C["bg"], fg=fg or C["text"], font=FB, **kw)

def _ent(parent, var=None, width=None, **kw):
    e = tk.Entry(parent, textvariable=var,
                 bg=C["inp"], fg=C["text"],
                 insertbackground=C["accent"],
                 relief="flat", bd=0, font=FB,
                 highlightthickness=1,
                 highlightcolor=C["accent"],
                 highlightbackground=C["border"], **kw)
    if width: e.configure(width=width)
    return e

def _combo(parent, var, values, width=16, **kw):
    return ttk.Combobox(parent, textvariable=var, values=values,
                        state="readonly", width=width, font=FB, **kw)

def _scroltext(parent, height=8, mono=False, **kw):
    f = tk.Frame(parent, bg=C["inp"])
    t = tk.Text(f, bg=C["inp"], fg=C["text"],
                insertbackground=C["accent"],
                selectbackground=C["sel"],
                relief="flat", bd=0,
                font=FM if mono else FB,
                height=height, wrap="none",
                **kw)
    sy = ttk.Scrollbar(f, orient=VERTICAL,   command=t.yview)
    sx = ttk.Scrollbar(f, orient=HORIZONTAL, command=t.xview)
    t.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
    sy.pack(side=RIGHT, fill=Y)
    sx.pack(side=tk.BOTTOM, fill=X)
    t.pack(side=LEFT, fill=BOTH, expand=True)
    f._text = t
    return f

def _pathrow(parent, label, var, pick_file=True, pick_multi=False, bg=None):
    bg = bg or C["panel"]
    row = tk.Frame(parent, bg=bg)
    tk.Label(row, text=label, bg=bg, fg=C["muted"], font=FS, width=22, anchor=W).pack(side=LEFT)
    _ent(row, var=var).pack(side=LEFT, fill=X, expand=True, padx=(0,5))
    def browse():
        if pick_multi:
            p = filedialog.askopenfilenames(filetypes=[("Supported","*.csv *.json *.jsonl *.txt *.tsv *.xlsx *.log"),("All","*.*")])
            if p: var.set(";".join(p))
        elif pick_file:
            p = filedialog.askopenfilename(filetypes=[("Supported","*.csv *.json *.jsonl *.txt *.tsv *.xlsx *.log"),("All","*.*")])
            if p: var.set(p)
        else:
            p = filedialog.askdirectory()
            if p: var.set(p)
    _btn(row, "Browse", cmd=browse).pack(side=LEFT)
    row.pack(fill=X, padx=10, pady=3)

def _saverow(parent, label, var, bg=None):
    bg = bg or C["panel"]
    row = tk.Frame(parent, bg=bg)
    tk.Label(row, text=label, bg=bg, fg=C["muted"], font=FS, width=22, anchor=W).pack(side=LEFT)
    _ent(row, var=var).pack(side=LEFT, fill=X, expand=True, padx=(0,5))
    def browse():
        p = filedialog.asksaveasfilename(defaultextension=".csv",
            filetypes=[("CSV","*.csv"),("JSONL","*.jsonl"),("TXT","*.txt"),("Excel","*.xlsx"),("All","*.*")])
        if p: var.set(p)
    _btn(row, "Save As", cmd=browse).pack(side=LEFT)
    row.pack(fill=X, padx=10, pady=3)

def _fmtrow(parent, var, bg=None):
    bg = bg or C["panel"]
    row = tk.Frame(parent, bg=bg)
    tk.Label(row, text="Output format", bg=bg, fg=C["muted"], font=FS, width=22, anchor=W).pack(side=LEFT)
    _combo(row, var, SUPPORTED_OUT, width=10).pack(side=LEFT)
    row.pack(fill=X, padx=10, pady=3)

def _checkrow(parent, *items, bg=None):
    bg = bg or C["panel"]
    row = tk.Frame(parent, bg=bg)
    for text, var in items:
        ttk.Checkbutton(row, text=text, variable=var, style="TCheckbutton").pack(side=LEFT, padx=(0,16))
    row.pack(fill=X, padx=10, pady=4)

def _sep(parent, bg=None):
    tk.Frame(parent, bg=bg or C["border"], height=1).pack(fill=X, padx=8, pady=4)

def _sec_head(parent, text, bg=None):
    bg = bg or C["panel"]
    f = tk.Frame(parent, bg=bg)
    tk.Label(f, text=text, bg=bg, fg=C["accent"], font=FH).pack(side=LEFT, padx=12, pady=(10,4))
    tk.Frame(parent, bg=C["border"], height=1).pack(fill=X, padx=8, pady=(0,6))
    return f


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═════════════════════════════════════════════════════════════════════════════
class UltraCSVStudio(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME}  v{APP_VERSION}")
        # Responsive geometry: 90% of screen, min 1100×700
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w  = max(1100, int(sw * 0.88))
        h  = max(700,  int(sh * 0.88))
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.minsize(1000, 640)
        self.configure(bg=C["bg"])
        self.state("zoomed")   # start maximised (Windows)
        # ── icon (title bar + taskbar + task manager) ────────────────────
        if ICON_PATH.exists():
            try:
                self.iconbitmap(str(ICON_PATH))
            except Exception:
                pass
            # Also set via PhotoImage for taskbar on some WM
            if _PIL:
                try:
                    pil_img  = Image.open(str(ICON_PATH))
                    self._tk_icon = ImageTk.PhotoImage(pil_img.resize((64,64), Image.LANCZOS))
                    self.iconphoto(True, self._tk_icon)
                except Exception:
                    pass

        # state
        self._log_q      = queue.Queue()
        self._job_ctrl   = JobControl()
        self._job_running= False
        self._presets    = load_presets()
        self._big_engine : BigFileEngine | None = None

        # progress vars
        self._pval   = DoubleVar(value=0)
        self._pstat  = StringVar(value="Ready")
        self._peta   = StringVar(value="")
        self._pmeta  = StringVar(value="")
        self._prescan= BooleanVar(value=True)

        # big viewer state
        self._big_row    = IntVar(value=1)
        self._big_chunk  = IntVar(value=_HW["bv_chunk_cap"])  # adaptive rows/page
        self._big_search = StringVar()
        self._big_mode   = StringVar(value="Table")
        self._big_offset = 0    # current byte offset for raw mode
        self._big_replace = StringVar()
        self._big_case    = BooleanVar(value=False)
        self._big_regex   = BooleanVar(value=False)
        self._big_whole   = BooleanVar(value=False)
        self._big_search_offset = 0   # track search position in file
        self._big_match_count = StringVar(value="")
        self._big_goto_line = IntVar(value=1)
        self._big_line_count = StringVar(value="")
        self._bv_chunk_cache = {}     # LRU chunk cache: (start_row, count) -> (headers, rows, raw)
        self._bv_cache_order = []     # LRU order tracking
        self._bv_cache_max   = 10     # max cached chunks
        self._bv_prefetching = False  # prevent duplicate prefetch

        self._apply_style()
        self._build_ui()
        self._start_root_bg()
        self._start_hdr_waves()
        self.after(100, self._poll_log)

    # ── wave animation (header) + gradient bg (root) ─────────────────────────
    # Matches the CSS CodePen wave: period ~176px, 4 parallax layers
    # Each layer: (fill_color, y_base_px, amplitude_px, speed_rad/frame, direction)
    _WAVE_LAYERS = [
        ("#0d0035",  36, 4, 0.010,  1),   # slowest / deepest back
        ("#180048",  33, 5, 0.016, -1),   # leftward drift
        ("#0a1848",  30, 6, 0.021,  1),
        ("#003040",  27, 8, 0.030,  1),   # fastest / brightest front
    ]
    _WAVE_FREQ = 2 * math.pi / 176       # one full cycle per 176 px

    # ── root background gradient (static, redraws on resize) ─────────────────
    def _start_root_bg(self):
        self._root_cv = tk.Canvas(self, bg=C["bg"], highlightthickness=0)
        self._root_cv.place(x=0, y=0, relwidth=1, relheight=1)
        self.tk.call('lower', str(self._root_cv))
        self._root_bg_due = False
        def _on_resize(e):
            if not self._root_bg_due:
                self._root_bg_due = True
                self.after(80, self._draw_root_gradient)
        self.bind("<Configure>", _on_resize)
        self.after(140, self._draw_root_gradient)

    def _draw_root_gradient(self):
        self._root_bg_due = False
        cv = self._root_cv
        w, h = self.winfo_width(), self.winfo_height()
        if w < 2: return
        cv.delete("bg")
        # top = dark purple #0f0730 → bottom = near-black #07021a
        steps = 28
        for i in range(steps):
            t = i / steps
            r = int(0x0f * (1 - t) + 0x07 * t)
            g = int(0x07 * (1 - t) + 0x02 * t)
            b = int(0x30 * (1 - t) + 0x1a * t)
            y0 = int(h * i / steps)
            y1 = int(h * (i + 1) / steps) + 1
            cv.create_rectangle(0, y0, w, y1,
                                 fill=f"#{r:02x}{g:02x}{b:02x}",
                                 outline="", tags="bg")

    # ── header wave animation ─────────────────────────────────────────────────
    def _start_hdr_waves(self):
        self._wave_offsets = [random.uniform(0, 6.28) for _ in self._WAVE_LAYERS]
        self._animate_hdr_waves()

    def _animate_hdr_waves(self):
        cv = self._hdr_cv
        w, h = cv.winfo_width(), cv.winfo_height()
        if w < 2 or h < 2:
            self.after(100, self._animate_hdr_waves)
            return
        cv.delete("wave")
        freq = self._WAVE_FREQ
        for idx, (layer, offset) in enumerate(zip(self._WAVE_LAYERS, self._wave_offsets)):
            color, y_base, amp, speed, direction = layer
            pts = [0, h]
            for x in range(0, w + 4, 2):
                y = y_base + amp * math.sin(x * freq + offset)
                pts += [x, y]
            pts += [w, h]
            cv.create_polygon(pts, fill=color, outline="", smooth=True, tags="wave")
            self._wave_offsets[idx] += speed * direction
        cv.tag_lower("wave")      # waves behind logo + text
        self.after(33, self._animate_hdr_waves)

    # ── theming ──────────────────────────────────────────────────────────────
    def _apply_style(self):
        s = ttk.Style(self)
        try: s.theme_use("clam")
        except Exception: pass

        s.configure("TNotebook",     background=C["bg"],  borderwidth=0)
        s.configure("TNotebook.Tab", background=C["tab_off"], foreground=C["muted"],
                    font=("Segoe UI", 10, "bold"), padding=(20, 9), borderwidth=0)
        s.map("TNotebook.Tab",
              background=[("selected", C["tab_on"]),  ("active", C["sel"])],
              foreground=[("selected", C["accent"]),  ("active", C["text"])],
              expand=[("selected", [1, 1, 1, 0])])

        s.configure("Dark.Horizontal.TProgressbar",
                    troughcolor=C["border"], background=C["prog"], thickness=6, borderwidth=0)
        s.configure("Treeview",
                    background=C["inp"], foreground=C["text"],
                    fieldbackground=C["inp"], rowheight=24, font=FB, borderwidth=0)
        s.configure("Treeview.Heading",
                    background=C["panel2"], foreground=C["accent"],
                    font=("Segoe UI", 9, "bold"))
        s.map("Treeview",
              background=[("selected", C["sel"])],
              foreground=[("selected", C["accent"])])
        s.configure("Vertical.TScrollbar",   background=C["border"], troughcolor=C["bg"],
                    borderwidth=0, arrowcolor=C["muted"])
        s.configure("Horizontal.TScrollbar", background=C["border"], troughcolor=C["bg"],
                    borderwidth=0, arrowcolor=C["muted"])
        s.configure("TSpinbox", fieldbackground=C["inp"], foreground=C["text"],
                    background=C["inp"], arrowcolor=C["muted"])
        s.configure("TCheckbutton", background=C["panel"], foreground=C["text"],
                    font=FB, focuscolor=C["accent"])
        s.map("TCheckbutton", background=[("active", C["panel"])])
        s.configure("TCombobox", fieldbackground=C["inp"], foreground=C["text"],
                    background=C["inp"], arrowcolor=C["muted"], selectbackground=C["sel"])

    # ── layout skeleton ───────────────────────────────────────────────────────
    def _build_ui(self):
        # ── header (Canvas – smoke animation drawn here) ──
        hdr_cv = tk.Canvas(self, bg=C["bg"], height=38, highlightthickness=0)
        hdr_cv.pack(fill=X)
        self._hdr_cv = hdr_cv
        # static content on the header canvas
        x_off = 10
        if _PIL and ICON_PATH.exists():
            try:
                _im = Image.open(str(ICON_PATH)).resize((26, 26), Image.LANCZOS)
                self._hdr_icon = ImageTk.PhotoImage(_im)
                hdr_cv.create_image(x_off + 13, 19, image=self._hdr_icon, tags="static")
                x_off += 38
            except Exception:
                pass
        hdr_cv.create_text(x_off, 19, text=APP_NAME, fill=C["accent"],
                           font=("Segoe UI", 12, "bold"), anchor="w", tags="static")
        x_off += len(APP_NAME) * 9
        hdr_cv.create_text(x_off, 21, text=f"v{APP_VERSION}", fill=C["muted"],
                           font=("Segoe UI", 9), anchor="w", tags="static")
        x_off += 36
        hdr_cv.create_text(x_off, 19,
                           text=f"  |  Opens: {SUPPORTED_IN_LABEL}"
                                f"  |  100 GB+ streaming  |  Batch · Split · Merge · Filter · Profile · Dedupe",
                           fill=C["muted"], font=("Segoe UI", 9), anchor="w", tags="static")
        # 1 px cyan accent line under header – glass edge
        tk.Frame(self, bg=C["accent"], height=1).pack(fill=X)

        # ── progress strip ──
        pstrip = tk.Frame(self, bg=C["strip"], height=38)
        pstrip.pack(fill=X); pstrip.pack_propagate(False)
        pin = tk.Frame(pstrip, bg=C["strip"])
        pin.pack(fill=X, padx=10, pady=6)

        tk.Label(pin, textvariable=self._pstat, bg=C["strip"], fg=C["accent"],
                 font=FS, width=38, anchor=W).pack(side=LEFT)
        self._pbar = ttk.Progressbar(pin, variable=self._pval, maximum=100,
                                     style="Dark.Horizontal.TProgressbar", length=380)
        self._pbar.pack(side=LEFT, padx=8)
        tk.Label(pin, textvariable=self._peta,  bg=C["strip"], fg=C["muted"], font=FS, width=16).pack(side=LEFT)
        tk.Label(pin, textvariable=self._pmeta, bg=C["strip"], fg=C["muted"], font=FS).pack(side=LEFT, padx=8)
        ttk.Checkbutton(pin, text="Pre-scan ETA", variable=self._prescan,
                        style="TCheckbutton").pack(side=LEFT, padx=8)

        self._pbtn_p = wbtn(pin, "⏸  Pause",  cmd=self._job_pause)
        self._pbtn_r = tk.Button(pin, text="▶  Resume", command=self._job_resume,
                                  bg="#007a40", fg="#ffffff",
                                  activebackground="#00a855", activeforeground="#ffffff",
                                  disabledforeground="#d0d0d0",
                                  relief="flat", bd=0, font=("Segoe UI", 10, "bold"),
                                  padx=16, pady=7, cursor="hand2",
                                  highlightthickness=1, highlightbackground="#00a855")
        self._pbtn_r.bind("<Enter>", lambda e: self._pbtn_r.configure(bg="#00a855"))
        self._pbtn_r.bind("<Leave>", lambda e: self._pbtn_r.configure(bg="#007a40"))
        self._pbtn_c = dbtn(pin, "✕  Cancel", cmd=self._job_cancel)
        self._pbtn_c.pack(side=RIGHT, padx=2)
        self._pbtn_r.pack(side=RIGHT, padx=2)
        self._pbtn_p.pack(side=RIGHT, padx=2)
        self._set_job_btns(False)

        tk.Frame(self, bg=C["border"], height=1).pack(fill=X)

        # ── body ──
        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill=BOTH, expand=True)

        self._build_sidebar(body)
        self._sidebar_visible = True

        self._nb = ttk.Notebook(body)
        self._nb.pack(side=LEFT, fill=BOTH, expand=True)

        # ── bottom status bar with Job Log toggle ────────────────────────────
        bbar = tk.Frame(self, bg=C["strip"], height=26)
        bbar.pack(fill=X, side=tk.BOTTOM); bbar.pack_propagate(False)
        self._log_toggle_btn = _btn(bbar, "📋 Job Log", cmd=self._toggle_log)
        self._log_toggle_btn.pack(side=LEFT, padx=(8,6), pady=2)
        self._bottom_status = StringVar(value=_engine_label() + f"  │  orjson: {'✓' if _ORJSON else '✗'}  │  xxhash: {'✓' if _XXHASH else '✗'}")
        tk.Label(bbar, textvariable=self._bottom_status,
                 bg=C["strip"], fg=C["muted"], font=("Segoe UI", 8)).pack(side=LEFT, padx=6)

        self._build_t1_convert()
        self._build_t2_batch()
        self._build_t3_bigviewer()
        self._build_t4_split()
        self._build_t5_merge()
        self._build_t6_profile()
        self._build_t7_filter()
        self._build_t8_dupes()
        self._build_t9_presets()
        self._build_t10_column_trim()

    # ── sidebar ───────────────────────────────────────────────────────────────
    def _build_sidebar(self, parent):
        sb = tk.Frame(parent, bg=C["panel"], width=270)
        sb.pack(side=LEFT, fill=Y, padx=(0,1)); sb.pack_propagate(False)
        self._sidebar = sb

        tk.Label(sb, text="  JOB LOG", bg=C["panel"], fg=C["muted"], font=FS).pack(anchor=W, pady=(10,2))
        tk.Frame(sb, bg=C["border"], height=1).pack(fill=X, padx=6)

        lf = _scroltext(sb, height=38, mono=True)
        lf.pack(fill=BOTH, expand=True, padx=6, pady=4)
        self._logw = lf._text
        self._logw.configure(state="disabled")

        _btn(sb, "Clear Log", cmd=self._clear_log).pack(padx=8, pady=(0,8), fill=X)

    def _toggle_log(self):
        """Toggle the Job Log sidebar visibility."""
        if self._sidebar_visible:
            self._sidebar.pack_forget()
            self._sidebar_visible = False
            self._log_toggle_btn.configure(text="📋 Show Log")
        else:
            # Re-pack sidebar before the notebook
            self._nb.pack_forget()
            self._sidebar.pack(side=LEFT, fill=Y, padx=(0,1))
            self._nb.pack(side=LEFT, fill=BOTH, expand=True)
            self._sidebar_visible = True
            self._log_toggle_btn.configure(text="📋 Job Log")

    # ── tab factory ───────────────────────────────────────────────────────────
    def _tab(self, title):
        outer = tk.Frame(self._nb, bg=C["bg"])
        self._nb.add(outer, text=f"  {title}  ")

        can = tk.Canvas(outer, bg=C["bg"], highlightthickness=0)
        vs  = ttk.Scrollbar(outer, orient=VERTICAL, command=can.yview)
        can.configure(yscrollcommand=vs.set)
        vs.pack(side=RIGHT, fill=Y)
        can.pack(side=LEFT, fill=BOTH, expand=True)

        inner = tk.Frame(can, bg=C["bg"])
        wid   = can.create_window((0,0), window=inner, anchor=NW)

        def _resize(e):
            can.configure(scrollregion=can.bbox("all"))
            can.itemconfig(wid, width=can.winfo_width())
        inner.bind("<Configure>", _resize)
        can.bind("<Configure>", lambda e: can.itemconfig(wid, width=e.width))

        # ── per-canvas mouse-wheel scrolling (no global hijack) ──────────
        def _on_enter(e):
            can.bind_all("<MouseWheel>",
                         lambda ev: can.yview_scroll(int(-1*(ev.delta/120)), "units"))
        def _on_leave(e):
            can.unbind_all("<MouseWheel>")
        can.bind("<Enter>", _on_enter)
        can.bind("<Leave>", _on_leave)
        return inner

    def _section(self, parent, title):
        # Outer wrapper with subtle border glow
        wrap = tk.Frame(parent, bg=C["border"], bd=0)
        wrap.pack(fill=X, padx=14, pady=(10, 4))
        f = tk.Frame(wrap, bg=C["panel"], bd=0)
        f.pack(fill=X, padx=1, pady=1)
        # Section header with accent left bar
        hdr = tk.Frame(f, bg=C["panel"])
        hdr.pack(fill=X, padx=0, pady=0)
        tk.Frame(hdr, bg=C["accent"], width=4).pack(side=LEFT, fill=Y, padx=(8,0), pady=(10,4))
        tk.Label(hdr, text=title, bg=C["panel"], fg=C["accent"],
                 font=FH, anchor=W).pack(side=LEFT, padx=(8,10), pady=(10,4))
        # Thin separator
        tk.Frame(f, bg=C["border"], height=1).pack(fill=X, padx=10, pady=(0,6))
        body = tk.Frame(f, bg=C["panel"])
        body.pack(fill=X, padx=6, pady=(0,10))
        return body

    # ═══════════════════════  TAB 1 — CONVERT  ═══════════════════════════════
    def _build_t1_convert(self):
        p = self._tab("⚡ Convert")

        # ═══ SECTION 1: Security Scanner ═════════════════════════════════════
        sc = self._section(p, "🛡️ Security Scanner — Content Disarm & Reconstruction (CDR)")
        tk.Label(sc, text=(
            "Scans files for embedded threats: macros · VBA · scripts · reverse shells · "
            "formula injection · encrypted payloads · binders · ActiveX · obfuscated code.\n"
            "Infected files are safely opened in isolated mode — only pure data rows are extracted. "
            "Your PC is never at risk."),
            bg=C["panel"], fg=C["muted"], font=FS, wraplength=900,
            justify=LEFT).pack(anchor=W, padx=12, pady=(4, 6))

        self.cv_in   = StringVar(); self.cv_outd = StringVar()
        self.cv_name = StringVar(); self.cv_fmt  = StringVar(value="csv")
        self.cv_trim = BooleanVar(value=True); self.cv_drop = BooleanVar(value=True)
        self.cv_sanitize = BooleanVar(value=True)

        _pathrow(sc, "Input file", self.cv_in)

        scbr = tk.Frame(sc, bg=C["panel"])
        _btn(scbr, "🔍  Scan Only", cmd=self._scan_convert).pack(side=LEFT, padx=(0, 8))
        tk.Label(scbr, text="Scan file for threats without converting (results in Job Log)",
                 bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        scbr.pack(fill=X, padx=10, pady=(4, 2))

        # Scan result display
        self._cv_scan_frame = tk.Frame(sc, bg=C["panel"])
        self._cv_scan_frame.pack(fill=X, padx=10, pady=(2, 6))
        self.cv_scan_result = StringVar(value="")
        tk.Label(self._cv_scan_frame, textvariable=self.cv_scan_result,
                 bg=C["panel"], fg=C["ok"], font=("Segoe UI", 9, "bold"),
                 wraplength=900, justify=LEFT).pack(anchor=W, padx=2)

        # ═══ SECTION 2: Convert Settings ═════════════════════════════════════
        s = self._section(p, "⚡ Convert — Clean, Sanitize & Export to Any Format")
        tk.Label(s, text=(
            "Safe files convert at full speed. Infected files are auto-sanitized — "
            "all macros, scripts, formulas, and embedded objects are stripped. "
            "Only clean data rows are written to the output file."),
            bg=C["panel"], fg=C["muted"], font=FS, wraplength=900,
            justify=LEFT).pack(anchor=W, padx=12, pady=(4, 6))

        _pathrow(s, "Output folder",  self.cv_outd, pick_file=False)

        nr = tk.Frame(s, bg=C["panel"])
        tk.Label(nr, text="Output filename (no ext)", bg=C["panel"], fg=C["muted"], font=FS, width=22, anchor=W).pack(side=LEFT)
        _ent(nr, var=self.cv_name).pack(side=LEFT, fill=X, expand=True)
        nr.pack(fill=X, padx=10, pady=3)

        _fmtrow(s, self.cv_fmt)
        _checkrow(s, ("Trim whitespace", self.cv_trim), ("Drop empty rows", self.cv_drop),
                     ("🛡️ Sanitize (strip threats)", self.cv_sanitize))

        br = tk.Frame(s, bg=C["panel"])
        abtn(br, "▶  Scan & Convert", cmd=self._run_convert).pack(side=LEFT, padx=(0,8))
        _btn(br, "👁  Preview",       cmd=self._preview_convert).pack(side=LEFT)
        br.pack(fill=X, padx=10, pady=8)

        self.cv_res = StringVar()
        tk.Label(p, textvariable=self.cv_res, bg=C["bg"], fg=C["ok"], font=FS).pack(anchor=W, padx=20, pady=4)

    # ═══════════════════════  TAB 2 — BATCH  ═════════════════════════════════
    def _build_t2_batch(self):
        p = self._tab("📂 Batch")
        s = self._section(p, "📂 Batch Directory — Scan, Sanitize & Convert All Files")
        tk.Label(s, text=(
            "Scans every file for threats before converting. Infected files are auto-sanitized — "
            "clean data extracted safely. Threat reports logged per file in Job Log."),
            bg=C["panel"], fg=C["muted"], font=FS, wraplength=900,
            justify=LEFT).pack(anchor=W, padx=12, pady=(4, 6))

        self.bt_in   = StringVar(); self.bt_out  = StringVar()
        self.bt_fmt  = StringVar(value="csv")
        self.bt_rec  = BooleanVar(value=False)
        self.bt_trim = BooleanVar(value=True)
        self.bt_drop = BooleanVar(value=True)
        self.bt_sanitize = BooleanVar(value=True)
        self.bt_wkrs = IntVar(value=max(2, min(8, os.cpu_count() or 4)))

        _pathrow(s, "Input directory",  self.bt_in,  pick_file=False)
        _pathrow(s, "Output directory", self.bt_out, pick_file=False)
        _fmtrow(s, self.bt_fmt)
        _checkrow(s, ("Recursive", self.bt_rec), ("Trim", self.bt_trim),
                     ("Drop empty rows", self.bt_drop), ("🛡️ Sanitize", self.bt_sanitize))

        wr = tk.Frame(s, bg=C["panel"])
        tk.Label(wr, text="Parallel workers", bg=C["panel"], fg=C["muted"], font=FS, width=22, anchor=W).pack(side=LEFT)
        ttk.Spinbox(wr, from_=1, to=64, textvariable=self.bt_wkrs, width=5, style="TSpinbox").pack(side=LEFT)
        tk.Label(wr, text="  (SSD: 4–16  HDD: 1–4)", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        wr.pack(fill=X, padx=10, pady=3)

        br = tk.Frame(s, bg=C["panel"])
        abtn(br, "▶  Scan & Batch Convert", cmd=self._run_batch).pack(side=LEFT)
        br.pack(fill=X, padx=10, pady=8)

        self.bt_res = StringVar()
        tk.Label(p, textvariable=self.bt_res, bg=C["bg"], fg=C["ok"], font=FS).pack(anchor=W, padx=20)

    # ═══════════════════════  TAB 3 — BIG FILE VIEWER  ═══════════════════════
    def _build_t3_bigviewer(self):
        """Professional big-file viewer: line numbers, find/replace, hex, context filter."""
        outer = tk.Frame(self._nb, bg=C["bg"])
        self._nb.add(outer, text="  🔎 Big Viewer  ")
        self._bv_outer = outer

        # ── toolbar row 1: file + navigation ─────────────────────────────────
        tb1 = tk.Frame(outer, bg=C["panel"], pady=4)
        tb1.pack(fill=X)
        tk.Label(tb1, text="FILE", bg=C["panel"], fg=C["accent"], font=("Segoe UI", 8, "bold")).pack(side=LEFT, padx=(10,6))
        self.bv_path = StringVar()
        bv_path_ent = _ent(tb1, var=self.bv_path, width=48)
        bv_path_ent.pack(side=LEFT, padx=(0,4))
        bv_path_ent.bind("<Return>", lambda e: self._bv_open_from_entry())
        abtn(tb1, "Open…", cmd=self._bv_open).pack(side=LEFT, padx=(0,8))

        tk.Frame(tb1, bg=C["border"], width=1).pack(side=LEFT, fill=Y, pady=2, padx=4)

        tk.Label(tb1, text="ROW", bg=C["panel"], fg=C["accent"], font=("Segoe UI", 8, "bold")).pack(side=LEFT, padx=(6,4))
        row_ent = _ent(tb1, var=self._big_row, width=9)
        row_ent.pack(side=LEFT)
        row_ent.bind("<Return>", lambda e: self._bv_load())
        tk.Label(tb1, text="COUNT", bg=C["panel"], fg=C["accent"], font=("Segoe UI", 8, "bold")).pack(side=LEFT, padx=(6,4))
        cnt_ent = _ent(tb1, var=self._big_chunk, width=6)
        cnt_ent.pack(side=LEFT)
        cnt_ent.bind("<Return>", lambda e: self._bv_load())

        _btn(tb1, "◀◀", cmd=self._bv_first).pack(side=LEFT, padx=(8,2))
        _btn(tb1, "◀ Prev", cmd=self._bv_prev).pack(side=LEFT, padx=2)
        abtn(tb1, "▶ Load", cmd=self._bv_load).pack(side=LEFT, padx=2)
        _btn(tb1, "Next ▶", cmd=self._bv_next).pack(side=LEFT, padx=2)
        _btn(tb1, "▶▶ Last", cmd=self._bv_last).pack(side=LEFT, padx=2)

        tk.Frame(tb1, bg=C["border"], width=1).pack(side=LEFT, fill=Y, pady=2, padx=6)

        tk.Label(tb1, text="GO TO LINE", bg=C["panel"], fg=C["accent"], font=("Segoe UI", 8, "bold")).pack(side=LEFT, padx=(4,4))
        goto_ent = _ent(tb1, var=self._big_goto_line, width=9)
        goto_ent.pack(side=LEFT)
        goto_ent.bind("<Return>", lambda e: self._bv_goto_line())
        _btn(tb1, "Go", cmd=self._bv_goto_line).pack(side=LEFT, padx=(4,4))

        tk.Frame(tb1, bg=C["border"], width=1).pack(side=LEFT, fill=Y, pady=2, padx=6)

        tk.Label(tb1, text="VIEW", bg=C["panel"], fg=C["accent"], font=("Segoe UI", 8, "bold")).pack(side=LEFT, padx=(4,4))
        view_modes = ["Table", "Raw Text", "Split", "Hex Dump", "Column Stats"]
        vmc = _combo(tb1, self._big_mode, view_modes, width=13)
        vmc.pack(side=LEFT)
        self._big_mode.trace_add("write", lambda *_: self._bv_refresh_mode())

        # ── toolbar row 2: find / replace ────────────────────────────────────
        tb2 = tk.Frame(outer, bg=C["panel2"], pady=4)
        tb2.pack(fill=X)

        tk.Label(tb2, text="FIND", bg=C["panel2"], fg=C["accent"], font=("Segoe UI", 8, "bold")).pack(side=LEFT, padx=(10,4))
        _ent(tb2, var=self._big_search, width=24).pack(side=LEFT, padx=(0,4))

        tk.Label(tb2, text="REPLACE", bg=C["panel2"], fg=C["warn"], font=("Segoe UI", 8, "bold")).pack(side=LEFT, padx=(6,4))
        _ent(tb2, var=self._big_replace, width=20).pack(side=LEFT, padx=(0,6))

        ttk.Checkbutton(tb2, text="Match Case", variable=self._big_case,
                        style="TCheckbutton").pack(side=LEFT, padx=4)
        ttk.Checkbutton(tb2, text="Regex", variable=self._big_regex,
                        style="TCheckbutton").pack(side=LEFT, padx=4)
        ttk.Checkbutton(tb2, text="Whole Word", variable=self._big_whole,
                        style="TCheckbutton").pack(side=LEFT, padx=4)

        _btn(tb2, "◀ Find Prev", cmd=self._bv_search_back).pack(side=LEFT, padx=(6,2))
        abtn(tb2, "Find Next ▶", cmd=self._bv_search_fwd).pack(side=LEFT, padx=2)
        _btn(tb2, "Count All", cmd=self._bv_count_all).pack(side=LEFT, padx=2)
        wbtn(tb2, "Replace", cmd=self._bv_replace_one).pack(side=LEFT, padx=2)
        dbtn(tb2, "Replace All", cmd=self._bv_replace_all).pack(side=LEFT, padx=2)

        tk.Label(tb2, textvariable=self._big_match_count,
                 bg=C["panel2"], fg=C["ok"], font=("Segoe UI", 9, "bold")).pack(side=LEFT, padx=8)

        # ── info bar ─────────────────────────────────────────────────────────
        info_bar = tk.Frame(outer, bg=C["panel"])
        info_bar.pack(fill=X)

        self.bv_info = StringVar(value="No file loaded — Open a file to begin.")
        tk.Label(info_bar, textvariable=self.bv_info,
                 bg=C["panel"], fg=C["text"], font=FS).pack(side=LEFT, padx=12, pady=3)

        self._big_line_count.set("")
        tk.Label(info_bar, textvariable=self._big_line_count,
                 bg=C["panel"], fg=C["ok"], font=("Segoe UI", 9, "bold")).pack(side=LEFT, padx=12)

        self.bv_idx_info = StringVar(value="")
        tk.Label(info_bar, textvariable=self.bv_idx_info,
                 bg=C["panel"], fg=C["accent"], font=FS).pack(side=RIGHT, padx=12)

        tk.Frame(outer, bg=C["accent"], height=1).pack(fill=X)

        # ── content area ─────────────────────────────────────────────────────
        self._bv_content = tk.Frame(outer, bg=C["bg"])
        self._bv_content.pack(fill=BOTH, expand=True)

        # Table pane (with line numbers)
        self._bv_tv_frame = tk.Frame(self._bv_content, bg=C["bg"])
        self._bv_tv       = None
        self._bv_tv_frame.pack(fill=BOTH, expand=True)

        # Raw pane
        self._bv_raw_frame = tk.Frame(self._bv_content, bg=C["bg"])
        rf = _scroltext(self._bv_raw_frame, height=34, mono=True)
        rf.pack(fill=BOTH, expand=True, padx=0, pady=0)
        self._bv_raw_text = rf._text
        self._bv_raw_text.configure(state="disabled")

        # Hex pane (hidden by default)
        self._bv_hex_frame = tk.Frame(self._bv_content, bg=C["bg"])
        hf = _scroltext(self._bv_hex_frame, height=34, mono=True)
        hf.pack(fill=BOTH, expand=True, padx=0, pady=0)
        self._bv_hex_text = hf._text
        self._bv_hex_text.configure(state="disabled")

        # Column stats pane (hidden by default)
        self._bv_stats_frame = tk.Frame(self._bv_content, bg=C["bg"])
        self._bv_stats_tv = None

        # ── status bar ───────────────────────────────────────────────────────
        sbar = tk.Frame(outer, bg=C["strip"])
        sbar.pack(fill=X)
        self.bv_status = StringVar(value="")
        tk.Label(sbar, textvariable=self.bv_status,
                 bg=C["strip"], fg=C["muted"], font=FS).pack(side=LEFT, padx=10, pady=2)

        # ── keyboard navigation bindings ─────────────────────────────────────
        def _bv_key(event):
            k = event.keysym
            if k == "Prior":       self._bv_prev()       # Page Up
            elif k == "Next":      self._bv_next()       # Page Down
            elif k == "Home":      self._bv_first()
            elif k == "End":       self._bv_last()
        outer.bind("<Prior>",  _bv_key)
        outer.bind("<Next>",   _bv_key)
        outer.bind("<Home>",   _bv_key)
        outer.bind("<End>",    _bv_key)

        # Mousewheel page navigation on the content area
        def _bv_wheel(event):
            if event.delta > 0:
                self._bv_prev()
            else:
                self._bv_next()

        def _bv_bind_wheel(event):
            self._bv_content.bind_all("<MouseWheel>", _bv_wheel)

        def _bv_unbind_wheel(event):
            self._bv_content.unbind_all("<MouseWheel>")

        self._bv_content.bind("<Enter>", _bv_bind_wheel)
        self._bv_content.bind("<Leave>", _bv_unbind_wheel)

        self._bv_refresh_mode()

    # ── view mode switching ──────────────────────────────────────────────────
    def _bv_refresh_mode(self):
        mode = self._big_mode.get()
        for f in (self._bv_tv_frame, self._bv_raw_frame,
                  self._bv_hex_frame, self._bv_stats_frame):
            f.pack_forget()
        if mode == "Table":
            self._bv_tv_frame.pack(fill=BOTH, expand=True)
        elif mode == "Raw Text":
            self._bv_raw_frame.pack(fill=BOTH, expand=True)
        elif mode == "Split":
            self._bv_tv_frame.pack(fill=BOTH, expand=True)
            self._bv_raw_frame.pack(fill=BOTH, expand=True)
        elif mode == "Hex Dump":
            self._bv_hex_frame.pack(fill=BOTH, expand=True)
            if self._big_engine:
                self._bv_load_hex()
        elif mode == "Column Stats":
            self._bv_stats_frame.pack(fill=BOTH, expand=True)
            if self._big_engine:
                self._bv_load_stats()

    # ── file open ────────────────────────────────────────────────────────────
    def _bv_open(self):
        p = filedialog.askopenfilename(
            title="Open large file",
            filetypes=[("All supported","*.csv *.json *.jsonl *.txt *.tsv *.xlsx *.log"),("All","*.*")]
        )
        if not p: return
        self._bv_open_path(p)

    def _bv_open_from_entry(self):
        """Open file typed/pasted into the path entry (Enter key)."""
        p = self.bv_path.get().strip()
        if p and Path(p).exists():
            self._bv_open_path(p)
        elif p:
            messagebox.showerror("Not found", p)

    def _bv_invalidate_cache(self):
        """Clear the chunk cache (called on new file open)."""
        self._bv_chunk_cache.clear()
        self._bv_cache_order.clear()

    def _bv_cache_put(self, key, value):
        """LRU insert into chunk cache."""
        if key in self._bv_chunk_cache:
            self._bv_cache_order.remove(key)
        elif len(self._bv_cache_order) >= self._bv_cache_max:
            oldest = self._bv_cache_order.pop(0)
            self._bv_chunk_cache.pop(oldest, None)
        self._bv_chunk_cache[key] = value
        self._bv_cache_order.append(key)

    def _bv_cache_get(self, key):
        """LRU lookup. Returns cached value or None."""
        if key in self._bv_chunk_cache:
            self._bv_cache_order.remove(key)
            self._bv_cache_order.append(key)
            return self._bv_chunk_cache[key]
        return None

    def _bv_open_path(self, p):
        """Open a file path in Big Viewer (called by open dialog or cross-tab)."""
        if self._big_engine:
            self._big_engine.stop_index()
            self._big_engine.stop_cache()

        self.bv_path.set(p)
        self._bv_invalidate_cache()
        self._big_engine = BigFileEngine(p)
        info = self._big_engine.file_info()
        self.bv_info.set(f"{info['name']}  |  {info['size']}  |  {info['ext']}")
        self.bv_idx_info.set("⏳ Indexing…")
        self._big_line_count.set("")
        self._big_search_offset = 0
        self._log(f"Big Viewer: Opened {info['name']} ({info['size']})")

        # Background indexer with live line count
        def _idx_prog(done, total, lines):
            pct = int(done / max(total, 1) * 100)
            self.after(0, lambda: (
                self.bv_idx_info.set(f"Indexing: {pct}%  ({lines:,} lines)"),
                self._big_line_count.set(f"Lines: {lines:,}"),
            ))
            if done >= total:
                eng = self._big_engine
                mode = "sparse" if (eng and eng._stride > 1) else "dense"
                stride_info = f" (stride {eng._stride})" if (eng and eng._stride > 1) else ""
                self.after(0, lambda: (
                    self.bv_idx_info.set(
                        f"✓ Indexed {lines:,} lines [{mode}{stride_info}] — seek ready"),
                    self._big_line_count.set(f"Total: {lines:,} lines"),
                    self._log(f"Big Viewer: Index complete — {lines:,} lines [{mode}]"),
                ))
        self._big_engine.start_index(on_progress=_idx_prog)

        # Row-data cache for XLSX/JSON
        if info["ext"] in (".xlsx", ".json"):
            def _cache_prog(rows_done, total_bytes):
                if self._big_engine and self._big_engine._row_cache_ready:
                    total = len(self._big_engine._row_cache_rows)
                    self.after(0, lambda: (
                        self.bv_idx_info.set(f"✓ Cached {total:,} rows — instant nav"),
                        self._log(f"Big Viewer: Row cache ready ({total:,} rows)"),
                    ))
                else:
                    self.after(0, lambda: self.bv_idx_info.set(
                        f"Caching rows… {rows_done:,}"
                    ))
            self._big_engine.start_row_cache(on_progress=_cache_prog)

        self._big_row.set(1)
        self._bv_load()

    # ── data loading ─────────────────────────────────────────────────────────
    def _bv_load(self):
        if not self._big_engine:
            messagebox.showinfo("No file", "Open a file first"); return
        row = max(1, self._big_row.get())
        cnt = max(1, self._big_chunk.get())
        cache_key = (row, cnt)

        # Instant display from cache
        cached = self._bv_cache_get(cache_key)
        if cached:
            headers, rows, raw_txt, elapsed = cached
            self._bv_display(headers, rows, raw_txt, row, elapsed)
            self.bv_status.set(self.bv_status.get() + "  ⚡ cached")
            self._bv_prefetch(row, cnt)
            return

        self._log(f"Big Viewer: Load rows {row:,}–{row+cnt-1:,}")

        def task():
            try:
                t0 = time.time()
                headers, rows = self._big_engine.read_rows(row, cnt)
                raw_txt = self._big_engine.read_raw_lines(row - 1, cnt + 1)
                elapsed = time.time() - t0
                self._bv_cache_put(cache_key, (headers, rows, raw_txt, elapsed))
                self.after(0, lambda: self._bv_display(headers, rows, raw_txt, row, elapsed))
                # Prefetch adjacent chunks
                self._bv_prefetch(row, cnt)
            except Exception as ex:
                self._log(f"Big Viewer: Error — {ex}")

        threading.Thread(target=task, daemon=True).start()

    def _bv_prefetch(self, current_row, cnt):
        """Background prefetch next and previous chunks for instant navigation."""
        if self._bv_prefetching or not self._big_engine:
            return
        self._bv_prefetching = True
        def _do_prefetch():
            try:
                for r in (current_row + cnt, max(1, current_row - cnt)):
                    key = (r, cnt)
                    if key not in self._bv_chunk_cache:
                        try:
                            t0 = time.time()
                            h, rows = self._big_engine.read_rows(r, cnt)
                            raw = self._big_engine.read_raw_lines(r - 1, cnt + 1)
                            el = time.time() - t0
                            self._bv_cache_put(key, (h, rows, raw, el))
                        except Exception:
                            pass
            finally:
                self._bv_prefetching = False
        threading.Thread(target=_do_prefetch, daemon=True).start()

    def _bv_display(self, headers, rows, raw_lines, start_row, elapsed=0):
        """Display loaded rows in table (with line #), raw text, and update status.

        Performance notes:
        - Treeview: batch insert with update disabled to prevent per-row redraws
        - Text widget: build one big string and insert once (1 insert vs N inserts)
        """
        # ── Table with line numbers (#) ──────────────────────────────────────
        if headers:
            display_cols = ["#"] + headers
            cur_cols = list(self._bv_tv["columns"]) if self._bv_tv else []
            if display_cols != cur_cols:
                self._bv_rebuild_tv(display_cols)
            tv = self._bv_tv
            # Batch delete + insert with canvas updates suppressed
            self.update_idletasks()
            tv.delete(*tv.get_children())
            for i, r in enumerate(rows):
                vals = [str(start_row + i)] + [str(r.get(c, "")) for c in headers]
                tv.insert("", END, values=vals, tags=("odd" if i % 2 else "even",))
            tv.tag_configure("odd",  background=C["inp"])
            tv.tag_configure("even", background=C["panel2"])

        # ── Raw text: build one string, single insert — avoids N-insert lag ──
        self._bv_raw_text.configure(state="normal")
        self._bv_raw_text.delete("1.0", END)
        if rows:
            bulk = "".join(
                f"{start_row+i:>8} │ {_fast_json_dumps(r)}\n"
                for i, r in enumerate(rows))
        else:
            bulk = "".join(
                f"{start_row+i:>8} │ {line}\n"
                for i, line in enumerate(raw_lines))
        if bulk:
            self._bv_raw_text.insert(END, bulk)
        self._bv_raw_text.configure(state="disabled")

        # ── Search highlight in visible text ─────────────────────────────────
        needle = self._big_search.get().strip()
        if needle:
            self._bv_highlight_text(needle)

        # ── Status bar ───────────────────────────────────────────────────────
        info = self._big_engine.file_info()
        idx_state = f"✓ {info['lines_indexed']:,} lines" if info["indexed"] else f"{info['lines_indexed']:,} indexed…"
        nrows = len(rows) if rows else len(raw_lines)
        speed = f"{elapsed*1000:.0f}ms" if elapsed > 0 else ""
        self.bv_status.set(
            f"Rows {start_row:,}–{start_row+nrows-1:,}  │  "
            f"{nrows:,} loaded  │  "
            f"File: {info['size']}  │  "
            f"Index: {idx_state}  │  "
            f"{speed}"
        )

    def _bv_highlight_text(self, needle):
        """Highlight search matches in the raw text pane.
        Capped at 2000 highlights to prevent GUI freeze on wildcard queries."""
        _MAX_HIGHLIGHTS = 2000
        self._bv_raw_text.configure(state="normal")
        self._bv_raw_text.tag_remove("match", "1.0", END)
        nocase = not self._big_case.get()
        start = "1.0"
        count = 0
        while count < _MAX_HIGHLIGHTS:
            idx = self._bv_raw_text.search(needle, start, stopindex=END,
                                            nocase=nocase, regexp=self._big_regex.get())
            if not idx:
                break
            end_idx = f"{idx}+{len(needle)}c"
            self._bv_raw_text.tag_add("match", idx, end_idx)
            start = end_idx
            count += 1
        self._bv_raw_text.tag_configure("match", background=C["warn"], foreground="#000")
        self._bv_raw_text.configure(state="disabled")
        if count >= _MAX_HIGHLIGHTS:
            self._big_match_count.set(f"{count}+ matches (capped — refine search)")
        elif count:
            self._big_match_count.set(f"{count} matches in view")

    # ── rebuild treeview with # column ───────────────────────────────────────
    def _bv_rebuild_tv(self, columns):
        for w in self._bv_tv_frame.winfo_children(): w.destroy()
        tv = ttk.Treeview(self._bv_tv_frame, columns=columns, show="headings")
        # Line number column
        tv.heading("#", text="#")
        tv.column("#", width=70, minwidth=50, stretch=False, anchor="e")
        col_w = max(60, min(200, 1100 // max(len(columns) - 1, 1)))
        for col in columns[1:]:
            tv.heading(col, text=col, command=lambda c=col: self._bv_sort(c))
            tv.column(col, width=col_w, minwidth=50, stretch=True)
        sy = ttk.Scrollbar(self._bv_tv_frame, orient=VERTICAL,   command=tv.yview)
        sx = ttk.Scrollbar(self._bv_tv_frame, orient=HORIZONTAL, command=tv.xview)
        tv.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=RIGHT,      fill=Y)
        sx.pack(side=tk.BOTTOM,  fill=X)
        tv.pack(fill=BOTH, expand=True)
        self._bv_tv = tv

        # Right-click context menu
        tv.bind("<Button-3>", self._bv_context_menu)

    def _bv_sort(self, col):
        if not self._bv_tv: return
        data = [(self._bv_tv.set(k, col), k) for k in self._bv_tv.get_children("")]
        try:    data.sort(key=lambda t: float(t[0]))
        except: data.sort()
        for i, (_, k) in enumerate(data):
            self._bv_tv.move(k, "", i)

    # ── hex dump view ────────────────────────────────────────────────────────
    def _bv_load_hex(self):
        if not self._big_engine: return
        row = max(0, self._big_row.get() - 1)
        offset = self._big_engine.line_to_offset(row) if self._big_engine._indexed else row * 80
        if offset < 0: offset = 0
        def task():
            try:
                hex_str = self._big_engine.read_hex(offset, 4096)
                self.after(0, lambda: self._bv_show_hex(hex_str, offset))
            except Exception as ex:
                self._log(f"Hex view error: {ex}")
        threading.Thread(target=task, daemon=True).start()

    def _bv_show_hex(self, hex_str, offset):
        self._bv_hex_text.configure(state="normal")
        self._bv_hex_text.delete("1.0", END)
        self._bv_hex_text.insert("1.0",
            f"── Hex Dump at offset 0x{offset:08x} ({offset:,} bytes) ──\n\n{hex_str}")
        self._bv_hex_text.configure(state="disabled")

    # ── column stats view ────────────────────────────────────────────────────
    def _bv_load_stats(self):
        if not self._big_engine: return
        def task():
            try:
                headers, rows = self._big_engine.read_rows(1, min(10000, self._big_chunk.get()))
                if not headers: return
                stats = []
                for col in headers:
                    vals = [str(r.get(col, "")) for r in rows]
                    non_empty = [v for v in vals if v.strip()]
                    nums = []
                    for v in non_empty:
                        try: nums.append(float(v))
                        except ValueError: pass
                    stat = {
                        "Column": col,
                        "Non-Empty": len(non_empty),
                        "Empty": len(vals) - len(non_empty),
                        "Unique": len(set(vals)),
                        "Min Len": min((len(v) for v in non_empty), default=0),
                        "Max Len": max((len(v) for v in non_empty), default=0),
                        "Numeric": len(nums),
                        "Min Val": f"{min(nums):.2f}" if nums else "",
                        "Max Val": f"{max(nums):.2f}" if nums else "",
                        "Avg": f"{sum(nums)/len(nums):.2f}" if nums else "",
                    }
                    stats.append(stat)
                self.after(0, lambda: self._bv_show_stats(stats))
            except Exception as ex:
                self._log(f"Stats error: {ex}")
        threading.Thread(target=task, daemon=True).start()

    def _bv_show_stats(self, stats):
        for w in self._bv_stats_frame.winfo_children(): w.destroy()
        if not stats: return
        cols = list(stats[0].keys())
        tv = ttk.Treeview(self._bv_stats_frame, columns=cols, show="headings")
        for c in cols:
            tv.heading(c, text=c)
            tv.column(c, width=90, minwidth=60, stretch=True)
        for i, s in enumerate(stats):
            vals = [str(s[c]) for c in cols]
            tv.insert("", END, values=vals, tags=("odd" if i % 2 else "even",))
        tv.tag_configure("odd",  background=C["inp"])
        tv.tag_configure("even", background=C["panel2"])
        sy = ttk.Scrollbar(self._bv_stats_frame, orient=VERTICAL, command=tv.yview)
        sx = ttk.Scrollbar(self._bv_stats_frame, orient=HORIZONTAL, command=tv.xview)
        tv.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=RIGHT, fill=Y)
        sx.pack(side=tk.BOTTOM, fill=X)
        tv.pack(fill=BOTH, expand=True)
        self._bv_stats_tv = tv

    # ── find / replace operations ────────────────────────────────────────────
    def _bv_get_search_needle(self):
        """Build search needle respecting Whole Word option."""
        needle = self._big_search.get().strip()
        if not needle: return None
        if self._big_whole.get() and not self._big_regex.get():
            needle = r"\b" + re.escape(needle) + r"\b"
            self._big_regex.set(True)
        return needle

    def _bv_search_fwd(self):
        needle = self._big_search.get().strip()
        if not needle or not self._big_engine: return
        case = self._big_case.get()
        use_re = self._big_regex.get()
        whole = self._big_whole.get()
        if whole and not use_re:
            search_needle = r"\b" + re.escape(needle) + r"\b"
            use_re = True
        else:
            search_needle = needle
        offset = self._big_search_offset

        def task():
            off, ln = self._big_engine.search_forward(search_needle, offset, case, use_re)
            if off >= 0:
                self._big_search_offset = off + 1
                target_row = max(1, ln + 1)  # convert 0-based line to 1-based row
                self.after(0, lambda: (
                    self._big_row.set(target_row),
                    self._bv_load(),
                    self._log(f"Found at line {target_row:,} (offset {off:,})"),
                    self._big_match_count.set(f"Found at line {target_row:,}"),
                ))
            else:
                self._big_search_offset = 0
                self.after(0, lambda: (
                    self._big_match_count.set("No more matches"),
                    self._log("Big Viewer: Search — no more matches"),
                ))
        threading.Thread(target=task, daemon=True).start()

    def _bv_search_back(self):
        needle = self._big_search.get().strip()
        if not needle or not self._big_engine: return
        case = self._big_case.get()
        offset = max(0, self._big_search_offset - 2)

        def task():
            off, ln = self._big_engine.search_backward(needle, offset, case)
            if off >= 0:
                self._big_search_offset = off
                target_row = max(1, ln + 1)
                self.after(0, lambda: (
                    self._big_row.set(target_row),
                    self._bv_load(),
                    self._log(f"Found at line {target_row:,} (offset {off:,})"),
                    self._big_match_count.set(f"Found at line {target_row:,}"),
                ))
            else:
                self.after(0, lambda: (
                    self._big_match_count.set("No previous match"),
                    self._log("Big Viewer: Search backward — no match"),
                ))
        threading.Thread(target=task, daemon=True).start()

    def _bv_count_all(self):
        needle = self._big_search.get().strip()
        if not needle or not self._big_engine: return
        self._big_match_count.set("Counting…")
        self._log("Big Viewer: Counting all matches…")
        case = self._big_case.get()
        use_re = self._big_regex.get()

        def task():
            n = self._big_engine.search_count(needle, case, use_re)
            self.after(0, lambda: (
                self._big_match_count.set(f"{n:,} total matches"),
                self._log(f"Big Viewer: Found {n:,} matches for '{needle}'"),
            ))
        threading.Thread(target=task, daemon=True).start()

    def _bv_replace_one(self):
        """Replace is not supported on huge files in-place for safety.
        Shows a message directing user to use Filter tab for transformations."""
        messagebox.showinfo("Replace",
            "Single replace on huge files is not safe in-place.\n"
            "Use Filter tab or export filtered results to a new file.")

    def _bv_replace_all(self):
        """Replace All: creates a new file with all replacements."""
        needle = self._big_search.get().strip()
        replacement = self._big_replace.get()
        if not needle or not self._big_engine:
            messagebox.showinfo("Replace All", "Enter a search term first."); return
        out = filedialog.asksaveasfilename(
            title="Save replaced file as…",
            defaultextension=self._big_engine.ext,
            filetypes=[("Same format", f"*{self._big_engine.ext}"), ("All", "*.*")]
        )
        if not out: return
        case = self._big_case.get()
        use_re = self._big_regex.get()
        self._log(f"Big Viewer: Replace All '{needle}' → '{replacement}' in {self._big_engine.path.name}")

        def task():
            count = 0
            try:
                with open(self._big_engine.path, "r", encoding=self._big_engine.encoding, errors="replace") as fin, \
                     open(out, "w", encoding="utf-8", newline="") as fout:
                    for line in fin:
                        if use_re:
                            flags = 0 if case else re.IGNORECASE
                            new_line, n = re.subn(needle, replacement, line, flags=flags)
                        else:
                            if case:
                                new_line = line.replace(needle, replacement)
                                n = line.count(needle)
                            else:
                                pat = re.compile(re.escape(needle), re.IGNORECASE)
                                new_line, n = pat.subn(replacement, line)
                        fout.write(new_line)
                        count += n
                self.after(0, lambda: (
                    self._big_match_count.set(f"Replaced {count:,} occurrences"),
                    self._log(f"Replace All done: {count:,} replacements → {out}"),
                    messagebox.showinfo("Replace All", f"Done! {count:,} replacements saved to:\n{out}"),
                ))
            except Exception as ex:
                self.after(0, lambda: (
                    self._log(f"Replace All error: {ex}"),
                    messagebox.showerror("Error", str(ex)),
                ))
        threading.Thread(target=task, daemon=True).start()

    # ── navigation ───────────────────────────────────────────────────────────
    def _bv_first(self):
        self._big_row.set(1)
        self._big_search_offset = 0
        self._bv_load()

    def _bv_prev(self):
        r = max(1, self._big_row.get() - self._big_chunk.get())
        self._big_row.set(r)
        self._bv_load()

    def _bv_next(self):
        self._big_row.set(self._big_row.get() + self._big_chunk.get())
        self._bv_load()

    def _bv_last(self):
        if not self._big_engine: return
        info = self._big_engine.file_info()
        total = info["lines_indexed"]
        if total > 0:
            r = max(1, total - self._big_chunk.get() + 1)
            self._big_row.set(r)
            self._bv_load()
        else:
            self._log("Big Viewer: Index not ready yet — can't jump to last")

    def _bv_goto_line(self):
        if not self._big_engine: return
        ln = max(1, self._big_goto_line.get())
        self._big_row.set(ln)
        self._log(f"Big Viewer: Go to line {ln:,}")
        self._bv_load()

    # ── right-click context menu ─────────────────────────────────────────────
    def _bv_context_menu(self, event):
        if not self._bv_tv: return
        row_id = self._bv_tv.identify_row(event.y)
        col_id = self._bv_tv.identify_column(event.x)
        if not row_id: return
        self._bv_tv.selection_set(row_id)

        # Get column name and cell value
        cols = self._bv_tv["columns"]
        col_idx = int(col_id.replace("#", "")) - 1 if col_id else 0
        col_name = cols[col_idx] if 0 <= col_idx < len(cols) else ""
        values = self._bv_tv.item(row_id, "values")
        cell_val = values[col_idx] if values and col_idx < len(values) else ""
        row_vals = "\t".join(str(v) for v in values) if values else ""

        menu = tk.Menu(self, tearoff=0,
                       bg=C["panel2"], fg=C["text"], activebackground=C["accent"],
                       activeforeground="#000", font=FS)

        menu.add_command(label=f"📋 Copy Cell: {str(cell_val)[:40]}",
                        command=lambda: (self.clipboard_clear(), self.clipboard_append(cell_val)))
        menu.add_command(label="📋 Copy Row",
                        command=lambda: (self.clipboard_clear(), self.clipboard_append(row_vals)))
        menu.add_separator()
        if col_name and col_name != "#":
            menu.add_command(label=f"🔍 Filter: {col_name} == '{str(cell_val)[:20]}'",
                            command=lambda: self._bv_send_to_filter(col_name, "==", cell_val))
            menu.add_command(label=f"🔍 Filter: {col_name} contains '{str(cell_val)[:20]}'",
                            command=lambda: self._bv_send_to_filter(col_name, "contains", cell_val))
            menu.add_command(label=f"🔍 Filter: {col_name} != '{str(cell_val)[:20]}'",
                            command=lambda: self._bv_send_to_filter(col_name, "!=", cell_val))
            menu.add_separator()
            menu.add_command(label=f"🔍 Filter: {col_name} > '{str(cell_val)[:20]}'",
                            command=lambda: self._bv_send_to_filter(col_name, ">", cell_val))
            menu.add_command(label=f"🔍 Filter: {col_name} < '{str(cell_val)[:20]}'",
                            command=lambda: self._bv_send_to_filter(col_name, "<", cell_val))
            menu.add_command(label=f"🔍 Filter: {col_name} is_empty",
                            command=lambda: self._bv_send_to_filter(col_name, "is_empty", ""))
            menu.add_command(label=f"🔍 Filter: {col_name} not_empty",
                            command=lambda: self._bv_send_to_filter(col_name, "not_empty", ""))
            menu.add_separator()
        menu.add_command(label="🔎 Search for this value",
                        command=lambda: (self._big_search.set(cell_val), self._bv_search_fwd()))
        menu.add_command(label="📊 View Column Stats",
                        command=lambda: (self._big_mode.set("Column Stats"),))

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    # ── send filter condition from Big Viewer to Filter tab ──────────────────
    def _bv_send_to_filter(self, col, op, val):
        """Transfer a filter condition from Big Viewer context menu to Filter tab."""
        if self._big_engine:
            self.fl_in.set(str(self._big_engine.path))
        # Add condition to filter tab
        self._fl_add_with_values(col, op, str(val))
        # Switch to Filter tab
        self._nb.select(6)  # Filter is tab index 6
        self._log(f"Filter: Added condition from Big Viewer — {col} {op} '{val}'")


    # ═══════════════════════  TAB 4 — SPLIT  ══════════════════════════════════
    def _build_t4_split(self):
        p = self._tab("✂ Split")
        s = self._section(p, "File Splitter — Split Into Multiple Output Files")
        self.sp_in   = StringVar(); self.sp_outd = StringVar()
        self.sp_rows = IntVar(value=100000); self.sp_fmt = StringVar(value="csv")
        self.sp_trim = BooleanVar(value=True); self.sp_drop = BooleanVar(value=True)

        _pathrow(s, "Input file",    self.sp_in)
        _pathrow(s, "Output folder", self.sp_outd, pick_file=False)

        rr = tk.Frame(s, bg=C["panel"])
        tk.Label(rr, text="Rows per output file", bg=C["panel"], fg=C["muted"], font=FS, width=22, anchor=W).pack(side=LEFT)
        _ent(rr, var=self.sp_rows, width=14).pack(side=LEFT)
        rr.pack(fill=X, padx=10, pady=3)

        _fmtrow(s, self.sp_fmt)
        _checkrow(s, ("Trim whitespace", self.sp_trim), ("Drop empty rows", self.sp_drop))

        br = tk.Frame(s, bg=C["panel"])
        abtn(br, "✂  Split File", cmd=self._run_split).pack(side=LEFT)
        br.pack(fill=X, padx=10, pady=8)

        self.sp_res = StringVar()
        tk.Label(p, textvariable=self.sp_res, bg=C["bg"], fg=C["ok"], font=FS).pack(anchor=W, padx=20)

    # ═══════════════════════  TAB 5 — MERGE  ══════════════════════════════════
    def _build_t5_merge(self):
        p = self._tab("🔀 Merge")
        top = self._section(p, "File Merger — Combine Multiple Files Into One")

        br = tk.Frame(top, bg=C["panel"])
        _btn(br, "+ Add Files",   cmd=self._mg_add_files).pack(side=LEFT, padx=(0,5))
        _btn(br, "+ Add Folder",  cmd=self._mg_add_folder).pack(side=LEFT, padx=(0,5))
        _btn(br, "↑ Move Up",     cmd=self._mg_up).pack(side=LEFT, padx=(0,5))
        _btn(br, "↓ Move Down",   cmd=self._mg_down).pack(side=LEFT, padx=(0,5))
        dbtn(br, "✕ Remove",      cmd=self._mg_remove).pack(side=LEFT, padx=(0,5))
        dbtn(br, "Clear All",     cmd=self._mg_clear).pack(side=LEFT)
        br.pack(fill=X, padx=10, pady=(0,6))

        lf = tk.Frame(top, bg=C["inp"])
        lf.pack(fill=BOTH, expand=True, padx=10, pady=4)
        self._mg_lb = tk.Listbox(lf, bg=C["inp"], fg=C["text"],
                                 selectbackground=C["sel"], selectforeground=C["accent"],
                                 font=FS, bd=0, relief="flat", selectmode="extended",
                                 activestyle="none", height=10)
        lbs = ttk.Scrollbar(lf, orient=VERTICAL, command=self._mg_lb.yview)
        self._mg_lb.configure(yscrollcommand=lbs.set)
        lbs.pack(side=RIGHT, fill=Y)
        self._mg_lb.pack(fill=BOTH, expand=True)
        self._mg_files = []

        self._mg_cnt = StringVar(value="0 files queued")
        tk.Label(top, textvariable=self._mg_cnt, bg=C["panel"], fg=C["muted"], font=FS).pack(anchor=W, padx=12)

        os_ = self._section(p, "Output Settings")
        self.mg_out = StringVar(); self.mg_fmt = StringVar(value="csv")
        _saverow(os_, "Output file", self.mg_out)
        _fmtrow(os_, self.mg_fmt)
        br2 = tk.Frame(os_, bg=C["panel"])
        abtn(br2, "🔀  Merge Now", cmd=self._run_merge).pack(side=LEFT)
        br2.pack(fill=X, padx=10, pady=8)

        self.mg_res = StringVar()
        tk.Label(p, textvariable=self.mg_res, bg=C["bg"], fg=C["ok"], font=FS).pack(anchor=W, padx=20)

    # ═══════════════════════  TAB 6 — PROFILE  ════════════════════════════════
    def _build_t6_profile(self):
        p = self._tab("📊 Profile")
        s = self._section(p, "Column Profiler — Type Inference & Statistics")
        self.pf_in  = StringVar(); self.pf_sample = IntVar(value=10000)

        _pathrow(s, "Input file", self.pf_in)
        sr = tk.Frame(s, bg=C["panel"])
        tk.Label(sr, text="Sample rows (0 = all)", bg=C["panel"], fg=C["muted"], font=FS, width=22, anchor=W).pack(side=LEFT)
        _ent(sr, var=self.pf_sample, width=12).pack(side=LEFT)
        sr.pack(fill=X, padx=10, pady=3)

        self.pf_exp_fmt = StringVar(value="csv")

        br = tk.Frame(s, bg=C["panel"])
        abtn(br, "📊  Run Profile", cmd=self._run_profile).pack(side=LEFT, padx=(0,8))
        _btn(br, "💾  Export",  cmd=self._export_profile).pack(side=LEFT, padx=(0,6))
        tk.Label(br, text="as", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT, padx=(0,4))
        _combo(br, self.pf_exp_fmt, SUPPORTED_OUT, width=8).pack(side=LEFT)
        br.pack(fill=X, padx=10, pady=8)

        cols = ("Column","Count","Empty %","Inferred Type","Min Len","Max Len","Unique ≤5","Sample Values")
        rf = tk.Frame(p, bg=C["bg"])
        rf.pack(fill=BOTH, expand=True, padx=14, pady=6)
        self._pf_tv = ttk.Treeview(rf, columns=cols, show="headings")
        cw = {"Column":150,"Count":70,"Empty %":75,"Inferred Type":110,
              "Min Len":70,"Max Len":70,"Unique ≤5":75,"Sample Values":280}
        for c in cols:
            self._pf_tv.heading(c, text=c)
            self._pf_tv.column(c, width=cw.get(c,100), minwidth=50, stretch=(c=="Sample Values"))
        psy = ttk.Scrollbar(rf, orient=VERTICAL,   command=self._pf_tv.yview)
        psx = ttk.Scrollbar(rf, orient=HORIZONTAL, command=self._pf_tv.xview)
        self._pf_tv.configure(yscrollcommand=psy.set, xscrollcommand=psx.set)
        psy.pack(side=RIGHT,      fill=Y)
        psx.pack(side=tk.BOTTOM,  fill=X)
        self._pf_tv.pack(fill=BOTH, expand=True)
        self._pf_data = []   # for export

    # ═══════════════════════  TAB 7 — FILTER  ════════════════════════════════
    def _build_t7_filter(self):
        p = self._tab("🔍 Filter")

        # ═══ SECTION 1: Query Builder ════════════════════════════════════════
        s = self._section(p, "🔍 Query Builder — Stream Rows Matching Conditions (AND logic)")
        self.fl_in   = StringVar(); self.fl_out  = StringVar()
        self.fl_fmt  = StringVar(value="csv"); self._fl_conds = []

        _pathrow(s, "Input file",  self.fl_in)
        _saverow(s, "Output file", self.fl_out)
        _fmtrow(s, self.fl_fmt)

        # conditions header
        ch = tk.Frame(s, bg=C["panel"])
        tk.Label(ch, text="Filter conditions  (ALL must match)",
                 bg=C["panel"], fg=C["text"], font=("Segoe UI", 9, "bold")).pack(side=LEFT)
        ch.pack(fill=X, padx=10, pady=(8, 2))

        self._fl_cond_box = tk.Frame(s, bg=C["panel"])
        self._fl_cond_box.pack(fill=X, padx=10, pady=4)

        cb = tk.Frame(s, bg=C["panel"])
        _btn(cb, "+ Add Condition", cmd=self._fl_add).pack(side=LEFT, padx=(0,6))
        dbtn(cb, "Clear All",       cmd=self._fl_clear).pack(side=LEFT)
        cb.pack(fill=X, padx=10, pady=4)

        br = tk.Frame(s, bg=C["panel"])
        abtn(br, "▶  Filter & Save",    cmd=self._run_filter).pack(side=LEFT, padx=(0,8))
        _btn(br, "👁  Preview 200 rows", cmd=self._preview_filter).pack(side=LEFT)
        br.pack(fill=X, padx=10, pady=(4, 6))

        self.fl_res = StringVar()
        tk.Label(s, textvariable=self.fl_res, bg=C["panel"], fg=C["ok"],
                 font=("Segoe UI", 9, "bold")).pack(anchor=W, padx=12, pady=(0, 2))

        # Query preview area
        qpf = _scroltext(p, height=10, mono=True)
        qpf.pack(fill=X, padx=16, pady=(2, 8))
        self._fl_prev_text = qpf._text
        self._fl_prev_text.configure(state="disabled")

        # ═══ SECTION 2: File Match Analyzer ══════════════════════════════════
        ms = self._section(p, "📊 File Match Analyzer — Compare Files for Similarity")
        self.fm_folder = StringVar()
        self.fm_threshold = IntVar(value=50)

        fmr = tk.Frame(ms, bg=C["panel"])
        tk.Label(fmr, text="Folder", bg=C["panel"], fg=C["muted"],
                 font=FS, width=14, anchor=W).pack(side=LEFT)
        _ent(fmr, var=self.fm_folder, width=48).pack(side=LEFT, padx=(0,4))
        _btn(fmr, "Browse…", cmd=lambda: self.fm_folder.set(
            filedialog.askdirectory(title="Select folder to analyze") or self.fm_folder.get()
        )).pack(side=LEFT)
        fmr.pack(fill=X, padx=10, pady=4)

        fmr2 = tk.Frame(ms, bg=C["panel"])
        tk.Label(fmr2, text="Min match %", bg=C["panel"], fg=C["muted"],
                 font=FS, width=14, anchor=W).pack(side=LEFT)
        _ent(fmr2, var=self.fm_threshold, width=6).pack(side=LEFT, padx=(0,8))
        tk.Label(fmr2, text="(show pairs above this similarity threshold)",
                 bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        fmr2.pack(fill=X, padx=10, pady=4)

        fmbr = tk.Frame(ms, bg=C["panel"])
        abtn(fmbr, "▶  Analyze & Compare", cmd=self._run_file_match).pack(side=LEFT, padx=(0,8))
        fmbr.pack(fill=X, padx=10, pady=(4, 6))

        self._fm_result = StringVar(value="")
        tk.Label(ms, textvariable=self._fm_result, bg=C["panel"], fg=C["ok"],
                 font=("Segoe UI", 9, "bold")).pack(anchor=W, padx=12, pady=(0, 2))

        # Report table
        self._fm_tv_frame = tk.Frame(p, bg=C["bg"])
        self._fm_tv_frame.pack(fill=X, padx=16, pady=(2, 8))
        self._fm_tv = None

        # ═══ SECTION 3: Keep Unique ══════════════════════════════════════════
        us = self._section(p, "🔑 Keep Unique — Deduplicate & Merge Unique Lines Into One File")
        self.ku_files   = StringVar()
        self.ku_mode    = StringVar(value="email")
        self.ku_custom  = StringVar()
        self.ku_out     = StringVar()
        self.ku_fmt     = StringVar(value="csv")
        self.ku_case    = BooleanVar(value=False)
        self.ku_strip   = BooleanVar(value=True)
        self.ku_skip_empty = BooleanVar(value=True)
        self.ku_merge_all  = BooleanVar(value=True)

        # Input row
        kur = tk.Frame(us, bg=C["panel"])
        tk.Label(kur, text="Input files", bg=C["panel"], fg=C["muted"],
                 font=FS, width=14, anchor=W).pack(side=LEFT)
        _ent(kur, var=self.ku_files, width=48).pack(side=LEFT, padx=(0,4))
        _btn(kur, "Browse…", cmd=self._ku_browse).pack(side=LEFT)
        kur.pack(fill=X, padx=10, pady=4)
        tk.Label(us, text="Select multiple files (semicolon-separated). All unique entries merged into one output.",
                 bg=C["panel"], fg=C["muted"], font=("Segoe UI", 8)).pack(anchor=W, padx=26, pady=(0,4))

        # Unique-by mode row
        kumr = tk.Frame(us, bg=C["panel"])
        tk.Label(kumr, text="Unique by", bg=C["panel"], fg=C["muted"],
                 font=FS, width=14, anchor=W).pack(side=LEFT)
        _combo(kumr, self.ku_mode,
               ["email", "phone", "entire_line", "column", "regex_pattern",
                "custom_columns", "first_n_chars", "json_key"],
               width=16).pack(side=LEFT, padx=(0,8))
        tk.Label(kumr, text="Column / Pattern / Key:",
                 bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT, padx=(8,4))
        _ent(kumr, var=self.ku_custom, width=22).pack(side=LEFT)
        kumr.pack(fill=X, padx=10, pady=4)

        # Custom config help
        ku_help = tk.Frame(us, bg=C["panel"])
        tk.Label(ku_help, bg=C["panel"], fg=C["muted"], font=("Segoe UI", 8),
                 text="email / phone: auto-detect  |  column: header name  |  "
                      "regex_pattern: your regex  |  custom_columns: col1,col2  |  "
                      "first_n_chars: number  |  json_key: nested.key.path"
                 ).pack(anchor=W, padx=4)
        ku_help.pack(fill=X, padx=26, pady=(0, 6))

        # Options row
        kor = tk.Frame(us, bg=C["panel"])
        tk.Label(kor, text="Options", bg=C["panel"], fg=C["muted"],
                 font=FS, width=14, anchor=W).pack(side=LEFT)
        ttk.Checkbutton(kor, text="Case insensitive", variable=self.ku_case,
                        style="TCheckbutton").pack(side=LEFT, padx=(0,12))
        ttk.Checkbutton(kor, text="Strip whitespace", variable=self.ku_strip,
                        style="TCheckbutton").pack(side=LEFT, padx=(0,12))
        ttk.Checkbutton(kor, text="Skip empty", variable=self.ku_skip_empty,
                        style="TCheckbutton").pack(side=LEFT, padx=(0,12))
        ttk.Checkbutton(kor, text="Merge all into one", variable=self.ku_merge_all,
                        style="TCheckbutton").pack(side=LEFT)
        kor.pack(fill=X, padx=10, pady=4)

        # Output
        _saverow(us, "Output file", self.ku_out)
        _fmtrow(us, self.ku_fmt)

        # Action buttons
        kubr = tk.Frame(us, bg=C["panel"])
        abtn(kubr, "▶  Extract Unique & Save", cmd=self._run_keep_unique).pack(side=LEFT, padx=(0,8))
        _btn(kubr, "👁  Preview 200", cmd=self._preview_keep_unique).pack(side=LEFT)
        kubr.pack(fill=X, padx=10, pady=(4, 6))

        self._ku_result = StringVar(value="")
        tk.Label(us, textvariable=self._ku_result, bg=C["panel"], fg=C["ok"],
                 font=("Segoe UI", 9, "bold")).pack(anchor=W, padx=12, pady=(0, 2))

        # Keep Unique preview area
        kupf = _scroltext(p, height=8, mono=True)
        kupf.pack(fill=X, padx=16, pady=(2, 8))
        self._ku_prev_text = kupf._text
        self._ku_prev_text.configure(state="disabled")

    FILTER_OPS = ["==","!=","contains","not_contains","starts","ends",
                  ">","<",">=","<=","is_empty","not_empty","regex"]

    def _fl_add(self):
        rf = tk.Frame(self._fl_cond_box, bg=C["panel"], pady=2)
        rf.pack(fill=X, pady=1)
        cv = StringVar(); ov = StringVar(value="contains"); vv = StringVar()
        tk.Label(rf, text="Col:", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        _ent(rf, var=cv, width=14).pack(side=LEFT, padx=(3,8))
        tk.Label(rf, text="Op:", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        _combo(rf, ov, self.FILTER_OPS, width=14).pack(side=LEFT, padx=(3,8))
        tk.Label(rf, text="Value:", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        _ent(rf, var=vv, width=22).pack(side=LEFT, padx=(3,8))
        cond = {"col_var":cv,"op_var":ov,"val_var":vv,"frame":rf}
        def rm():
            self._fl_conds.remove(cond)
            rf.destroy()
        dbtn(rf, "✕", cmd=rm).pack(side=LEFT)
        self._fl_conds.append(cond)

    def _fl_clear(self):
        for c in self._fl_conds: c["frame"].destroy()
        self._fl_conds.clear()

    def _fl_add_with_values(self, col, op, val):
        """Add a pre-populated condition (called from Big Viewer context menu)."""
        rf = tk.Frame(self._fl_cond_box, bg=C["panel"], pady=2)
        rf.pack(fill=X, pady=1)
        cv = StringVar(value=col); ov = StringVar(value=op); vv = StringVar(value=val)
        tk.Label(rf, text="Col:", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        _ent(rf, var=cv, width=14).pack(side=LEFT, padx=(3,8))
        tk.Label(rf, text="Op:", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        _combo(rf, ov, self.FILTER_OPS, width=14).pack(side=LEFT, padx=(3,8))
        tk.Label(rf, text="Value:", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        _ent(rf, var=vv, width=22).pack(side=LEFT, padx=(3,8))
        cond = {"col_var":cv,"op_var":ov,"val_var":vv,"frame":rf}
        def rm():
            self._fl_conds.remove(cond)
            rf.destroy()
        dbtn(rf, "✕", cmd=rm).pack(side=LEFT)
        self._fl_conds.append(cond)

    def _fl_get_conds(self):
        return [{"col": c["col_var"].get().strip(),
                 "op":  c["op_var"].get(),
                 "val": c["val_var"].get()}
                for c in self._fl_conds if c["col_var"].get().strip()]

    # ── File Match Analyzer backend ──────────────────────────────────────────
    _EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
    _PHONE_RE = re.compile(r"\+?\d[\d\-\s().]{7,}\d")

    @staticmethod
    def _file_text_lines(path):
        """Read all text lines from a supported file, return list of stripped strings."""
        ext = Path(path).suffix.lower()
        lines = []
        try:
            if ext in (".csv", ".tsv", ".txt", ".log", ".jsonl"):
                with open(path, "r", encoding="utf-8", errors="replace") as f:
                    lines = [l.rstrip("\n\r") for l in f]
            elif ext == ".json":
                with open(path, "r", encoding="utf-8", errors="replace") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    for item in data:
                        lines.append(_fast_json_dumps(item))
                else:
                    lines = [_fast_json_dumps(data)]
            elif ext == ".xlsx" and _OPENPYXL:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join(str(c) if c is not None else "" for c in row))
                wb.close()
            else:
                with open(path, "r", encoding="utf-8", errors="replace") as f:
                    lines = [l.rstrip("\n\r") for l in f]
        except Exception:
            pass
        return lines

    def _run_file_match(self):
        folder = self.fm_folder.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showerror("Error", "Select a valid folder"); return
        threshold = max(0, min(100, self.fm_threshold.get()))
        self._fm_result.set("⏳ Scanning files…")
        self._log(f"File Match: Analyzing folder {folder}")

        def task():
            try:
                # Gather all supported files
                supported = {".csv", ".tsv", ".txt", ".log", ".jsonl", ".json", ".xlsx"}
                files = [f for f in Path(folder).iterdir()
                         if f.is_file() and f.suffix.lower() in supported]
                if len(files) < 2:
                    self.after(0, lambda: (
                        self._fm_result.set("Need at least 2 files to compare"),
                        self._log("File Match: Less than 2 files found"),
                    ))
                    return

                self.after(0, lambda: self._fm_result.set(
                    f"⏳ Reading {len(files)} files…"))

                # Read all files into sets of lines (parallel I/O)
                file_data = {}
                def _read_one(fp):
                    lines = self._file_text_lines(str(fp))
                    # Use xxhash for fast content fingerprint
                    content_hash = _fast_hash("\n".join(lines)) if lines else ""
                    return fp.name, {
                        "path": str(fp),
                        "lines": lines,
                        "line_set": set(lines),
                        "count": len(lines),
                        "size": fp.stat().st_size,
                        "hash": content_hash,
                    }
                with ThreadPoolExecutor(max_workers=min(_HW["workers"], len(files))) as pool:
                    for name, data in pool.map(_read_one, files):
                        file_data[name] = data

                self.after(0, lambda: self._fm_result.set(
                    f"⏳ Comparing {len(files)} files ({len(files)*(len(files)-1)//2} pairs)…"))

                # Pairwise comparison
                results = []
                names = sorted(file_data.keys())
                for i in range(len(names)):
                    for j in range(i + 1, len(names)):
                        a_name, b_name = names[i], names[j]
                        a, b = file_data[a_name], file_data[b_name]

                        # Fast path: xxhash content fingerprint → instant identical check
                        exact = (a["hash"] == b["hash"] and a["hash"] != "")

                        if exact:
                            pct = 100.0
                        else:
                            a_set, b_set = a["line_set"], b["line_set"]
                            if not a_set and not b_set:
                                pct = 100.0
                            elif not a_set or not b_set:
                                pct = 0.0
                            else:
                                common = len(a_set & b_set)
                                union  = len(a_set | b_set)
                                pct = (common / union) * 100 if union else 0

                        if pct >= threshold or exact:
                            status = "🟢 Identical" if exact else (
                                "🟡 Very Similar" if pct >= 90 else (
                                "🟠 Similar" if pct >= 50 else "🔴 Different"
                            ))
                            results.append({
                                "File A": a_name,
                                "File B": b_name,
                                "Match %": f"{pct:.1f}%",
                                "Common Lines": len(a_set & b_set),
                                "A Lines": a["count"],
                                "B Lines": b["count"],
                                "Status": status,
                            })

                results.sort(key=lambda r: float(r["Match %"].rstrip("%")), reverse=True)

                self.after(0, lambda: self._show_file_match_results(results, len(files)))
                self._log(f"File Match: {len(results)} pairs found above {threshold}% threshold")
            except Exception as ex:
                self.after(0, lambda: (
                    self._fm_result.set(f"Error: {ex}"),
                    self._log(f"File Match error: {ex}"),
                ))

        threading.Thread(target=task, daemon=True).start()

    def _show_file_match_results(self, results, total_files):
        """Display file match results in the report table."""
        # Clear old table
        for w in self._fm_tv_frame.winfo_children():
            w.destroy()

        if not results:
            self._fm_result.set(f"No similar pairs found among {total_files} files.")
            return

        cols = list(results[0].keys())
        tv = ttk.Treeview(self._fm_tv_frame, columns=cols, show="headings", height=min(12, len(results)))
        for c in cols:
            w = 130 if c in ("File A", "File B") else (80 if c == "Match %" else 100)
            tv.heading(c, text=c)
            tv.column(c, width=w, minwidth=60, stretch=True)
        for i, r in enumerate(results):
            vals = [str(r[c]) for c in cols]
            tv.insert("", END, values=vals, tags=("odd" if i % 2 else "even",))
        tv.tag_configure("odd",  background=C["inp"])
        tv.tag_configure("even", background=C["panel2"])
        sy = ttk.Scrollbar(self._fm_tv_frame, orient=VERTICAL, command=tv.yview)
        sx = ttk.Scrollbar(self._fm_tv_frame, orient=HORIZONTAL, command=tv.xview)
        tv.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sy.pack(side=RIGHT, fill=Y)
        sx.pack(side=tk.BOTTOM, fill=X)
        tv.pack(fill=BOTH, expand=True)
        self._fm_tv = tv

        identical = sum(1 for r in results if "Identical" in r["Status"])
        similar   = sum(1 for r in results if "Similar" in r["Status"] and "Identical" not in r["Status"])
        self._fm_result.set(
            f"✓ {total_files} files scanned  │  "
            f"{len(results)} pairs shown  │  "
            f"{identical} identical  │  {similar} similar"
        )

    # ── Keep Unique backend ──────────────────────────────────────────────────
    def _ku_browse(self):
        files = filedialog.askopenfilenames(
            title="Select files for deduplication",
            filetypes=[("All supported", "*.csv *.json *.jsonl *.txt *.tsv *.xlsx *.log"), ("All", "*.*")]
        )
        if files:
            self.ku_files.set(";".join(files))

    def _ku_extract_key(self, line, mode, custom="", case_insensitive=False,
                        strip_ws=True):
        """Extract the uniqueness key from a line based on mode."""
        if strip_ws:
            line = line.strip()
        if mode == "email":
            m = self._EMAIL_RE.search(line)
            key = m.group(0).lower() if m else None
        elif mode == "phone":
            m = self._PHONE_RE.search(line)
            key = re.sub(r"[\s\-().]+", "", m.group(0)) if m else None
        elif mode == "entire_line":
            key = line
        elif mode == "column":
            key = custom  # handled in main loop via CSV DictReader
        elif mode == "regex_pattern":
            if custom:
                try:
                    m = re.search(custom, line)
                    key = m.group(0) if m else None
                except re.error:
                    key = None
            else:
                key = None
        elif mode == "first_n_chars":
            try:
                n = int(custom) if custom else 10
            except ValueError:
                n = 10
            key = line[:n] if line else None
        elif mode == "json_key":
            # Extract a nested key from a JSON line, e.g. "user.email"
            try:
                obj = _fast_json_loads(line)
                parts = custom.split(".") if custom else []
                val = obj
                for part in parts:
                    if isinstance(val, dict):
                        val = val.get(part)
                    else:
                        val = None
                        break
                key = str(val) if val is not None else None
            except Exception:
                key = None
        else:
            key = line
        if key is not None and case_insensitive:
            key = key.lower()
        return key

    def _ku_extract_key_dict(self, row, mode, custom="", case_insensitive=False):
        """Extract uniqueness key from a dict row (CSV/JSONL with headers)."""
        if mode == "column":
            val = str(row.get(custom, "")).strip()
        elif mode == "custom_columns":
            cols = [c.strip() for c in custom.split(",") if c.strip()]
            val = "|".join(str(row.get(c, "")).strip() for c in cols)
        elif mode == "email":
            combined = " ".join(str(v) for v in row.values())
            m = self._EMAIL_RE.search(combined)
            val = m.group(0).lower() if m else None
            return val
        elif mode == "phone":
            combined = " ".join(str(v) for v in row.values())
            m = self._PHONE_RE.search(combined)
            val = re.sub(r"[\s\-().]+", "", m.group(0)) if m else None
            return val
        elif mode == "json_key":
            parts = custom.split(".") if custom else []
            val = row
            for part in parts:
                if isinstance(val, dict):
                    val = val.get(part)
                else:
                    return None
            val = str(val) if val is not None else None
        elif mode == "first_n_chars":
            try:
                n = int(custom) if custom else 10
            except ValueError:
                n = 10
            combined = " ".join(str(v) for v in row.values())
            val = combined[:n]
        elif mode == "entire_line":
            val = "|".join(str(v) for v in row.values())
        elif mode == "regex_pattern":
            combined = " ".join(str(v) for v in row.values())
            try:
                m = re.search(custom, combined) if custom else None
                val = m.group(0) if m else None
            except re.error:
                val = None
        else:
            val = "|".join(str(v) for v in row.values())
        if val is not None and case_insensitive:
            val = val.lower()
        return val if val else None

    def _run_keep_unique(self):
        raw = self.ku_files.get().strip()
        out = self.ku_out.get().strip()
        if not raw:
            messagebox.showerror("Error", "Select input files"); return
        if not out:
            messagebox.showerror("Error", "Specify output file"); return
        files = [f.strip() for f in raw.split(";") if f.strip() and Path(f.strip()).exists()]
        if not files:
            messagebox.showerror("Error", "No valid files found"); return

        mode = self.ku_mode.get()
        custom = self.ku_custom.get().strip()
        fmt = self.ku_fmt.get()
        case_i = self.ku_case.get()
        strip_ws = self.ku_strip.get()
        skip_empty = self.ku_skip_empty.get()
        self._ku_result.set("⏳ Processing…")
        self._log(f"Keep Unique: {len(files)} files, mode={mode}, case_insensitive={case_i}")

        def task():
            try:
                seen_keys = set()
                unique_rows = []     # list of dict or str
                total_lines = 0
                skipped = 0
                is_dict_mode = False
                all_headers = []

                for fpath in files:
                    ext = Path(fpath).suffix.lower()

                    # ── Structured (CSV/TSV/JSONL with headers) ──────────
                    if ext in (".csv", ".tsv", ".txt", ".log"):
                        is_dict_mode = True
                        delim = DataProcessor._detect_delim(Path(fpath))
                        with open(fpath, "r", encoding="utf-8", errors="replace",
                                  newline="", buffering=_IO_BUFFER) as f:
                            reader = csv.DictReader(f, delimiter=delim)
                            if reader.fieldnames:
                                for h in reader.fieldnames:
                                    if h not in all_headers:
                                        all_headers.append(h)
                            for row in reader:
                                total_lines += 1
                                if skip_empty and all(
                                        not str(v).strip() for v in row.values()):
                                    skipped += 1
                                    continue
                                key = self._ku_extract_key_dict(
                                    row, mode, custom, case_i)
                                if key is None:
                                    skipped += 1
                                    continue
                                if strip_ws:
                                    key = key.strip()
                                h = _fast_hash(key)
                                if h not in seen_keys:
                                    seen_keys.add(h)
                                    unique_rows.append(row)
                    elif ext == ".jsonl":
                        is_dict_mode = True
                        with open(fpath, "rb", buffering=_IO_BUFFER) as f:
                            for raw_line in f:
                                raw_line = raw_line.strip()
                                if not raw_line:
                                    continue
                                total_lines += 1
                                try:
                                    obj = _fast_json_loads(raw_line)
                                except Exception:
                                    skipped += 1
                                    continue
                                if not isinstance(obj, dict):
                                    skipped += 1
                                    continue
                                for k in obj:
                                    if k not in all_headers:
                                        all_headers.append(k)
                                key = self._ku_extract_key_dict(
                                    obj, mode, custom, case_i)
                                if key is None:
                                    skipped += 1
                                    continue
                                h = _fast_hash(key)
                                if h not in seen_keys:
                                    seen_keys.add(h)
                                    unique_rows.append(obj)
                    else:
                        # ── Plain text line-based ────────────────────────
                        lines = self._file_text_lines(fpath)
                        for line in lines:
                            total_lines += 1
                            if skip_empty and not line.strip():
                                skipped += 1
                                continue
                            key = self._ku_extract_key(
                                line, mode, custom, case_i, strip_ws)
                            if key is None:
                                skipped += 1
                                continue
                            h = _fast_hash(key)
                            if h not in seen_keys:
                                seen_keys.add(h)
                                unique_rows.append(line)

                # ── Write all unique to single output file ───────────────
                n = 0
                if unique_rows:
                    if is_dict_mode and isinstance(unique_rows[0], dict):
                        n = DataProcessor.write_rows(
                            iter(unique_rows), out, fmt, log=self._log)
                    else:
                        with open(out, "w", encoding="utf-8",
                                  newline="", buffering=_IO_BUFFER) as f:
                            for line in unique_rows:
                                f.write(str(line) + "\n")
                                n += 1

                dups = total_lines - len(seen_keys) - skipped
                self.after(0, lambda: (
                    self._ku_result.set(
                        f"✓ {total_lines:,} lines  │  "
                        f"{len(seen_keys):,} unique  │  "
                        f"{dups:,} dups removed  │  "
                        f"{skipped:,} skipped  │  → {Path(out).name}"
                    ),
                    self._log(f"Keep Unique: {n:,} unique → {out}"),
                ))
            except Exception as ex:
                self.after(0, lambda: (
                    self._ku_result.set(f"Error: {ex}"),
                    self._log(f"Keep Unique error: {ex}"),
                ))

        threading.Thread(target=task, daemon=True).start()

    def _preview_keep_unique(self):
        raw = self.ku_files.get().strip()
        if not raw:
            messagebox.showerror("Error", "Select input files"); return
        files = [f.strip() for f in raw.split(";") if f.strip() and Path(f.strip()).exists()]
        if not files:
            messagebox.showerror("Error", "No valid files found"); return
        mode = self.ku_mode.get()
        custom = self.ku_custom.get().strip()
        case_i = self.ku_case.get()
        strip_ws = self.ku_strip.get()

        def task():
            seen_keys = set()
            preview = []
            for fpath in files:
                ext = Path(fpath).suffix.lower()
                if ext in (".csv", ".tsv", ".txt", ".log"):
                    delim = DataProcessor._detect_delim(Path(fpath))
                    with open(fpath, "r", encoding="utf-8", errors="replace",
                              newline="") as f:
                        reader = csv.DictReader(f, delimiter=delim)
                        for row in reader:
                            key = self._ku_extract_key_dict(
                                row, mode, custom, case_i)
                            if key:
                                h = _fast_hash(key)
                                if h not in seen_keys:
                                    seen_keys.add(h)
                                    preview.append(_fast_json_dumps(row))
                            if len(preview) >= 200: break
                elif ext == ".jsonl":
                    with open(fpath, "rb") as f:
                        for raw_line in f:
                            raw_line = raw_line.strip()
                            if not raw_line: continue
                            try:
                                obj = _fast_json_loads(raw_line)
                            except Exception:
                                continue
                            if isinstance(obj, dict):
                                key = self._ku_extract_key_dict(
                                    obj, mode, custom, case_i)
                                if key:
                                    h = _fast_hash(key)
                                    if h not in seen_keys:
                                        seen_keys.add(h)
                                        preview.append(_fast_json_dumps(obj))
                            if len(preview) >= 200: break
                else:
                    lines = self._file_text_lines(fpath)
                    for line in lines:
                        key = self._ku_extract_key(
                            line, mode, custom, case_i, strip_ws)
                        if key:
                            h = _fast_hash(key)
                            if h not in seen_keys:
                                seen_keys.add(h)
                                preview.append(line)
                        if len(preview) >= 200: break
                if len(preview) >= 200: break

            def update():
                self._ku_prev_text.configure(state="normal")
                self._ku_prev_text.delete("1.0", END)
                self._ku_prev_text.insert("1.0",
                    "\n".join(preview) if preview else "No unique lines found")
                self._ku_prev_text.configure(state="disabled")
            self.after(0, update)
            self._log(f"Keep Unique preview: {len(preview)} rows")

        threading.Thread(target=task, daemon=True).start()

    # ═══════════════════════  TAB 8 — DUPLICATES  ════════════════════════════
    def _build_t8_dupes(self):
        p = self._tab("♊ Duplicates")
        s = self._section(p, "Duplicate Finder — Find & Extract Duplicates by Key Columns")
        self.du_in   = StringVar(); self.du_keys = StringVar()
        self.du_odup = StringVar(); self.du_ouni = StringVar()
        self.du_fmt  = StringVar(value="csv")

        _pathrow(s, "Input file", self.du_in)

        kr = tk.Frame(s, bg=C["panel"])
        tk.Label(kr, text="Key columns (comma-sep)", bg=C["panel"], fg=C["muted"], font=FS, width=24, anchor=W).pack(side=LEFT)
        _ent(kr, var=self.du_keys, width=40).pack(side=LEFT)
        kr.pack(fill=X, padx=10, pady=3)

        _saverow(s, "Save duplicates →", self.du_odup)
        _saverow(s, "Save unique rows →", self.du_ouni)
        _fmtrow(s, self.du_fmt)

        br = tk.Frame(s, bg=C["panel"])
        abtn(br, "🔍  Scan",               cmd=self._run_scan_dupes).pack(side=LEFT, padx=(0,8))
        _btn(br, "💾  Export Duplicates",  cmd=lambda: self._export_dupes("dup")).pack(side=LEFT, padx=(0,8))
        _btn(br, "💾  Export Unique",      cmd=lambda: self._export_dupes("uni")).pack(side=LEFT)
        br.pack(fill=X, padx=10, pady=8)

        df = tk.Frame(p, bg=C["bg"])
        df.pack(fill=BOTH, expand=True, padx=14, pady=6)
        dc = ("Key Values", "Occurrences")
        self._du_tv = ttk.Treeview(df, columns=dc, show="headings")
        self._du_tv.heading("Key Values",   text="Key Values")
        self._du_tv.heading("Occurrences",  text="Occurrences")
        self._du_tv.column("Key Values",    width=560, minwidth=100)
        self._du_tv.column("Occurrences",   width=100, minwidth=60)
        dsy = ttk.Scrollbar(df, orient=VERTICAL,   command=self._du_tv.yview)
        dsx = ttk.Scrollbar(df, orient=HORIZONTAL, command=self._du_tv.xview)
        self._du_tv.configure(yscrollcommand=dsy.set, xscrollcommand=dsx.set)
        dsy.pack(side=RIGHT,     fill=Y)
        dsx.pack(side=tk.BOTTOM, fill=X)
        self._du_tv.pack(fill=BOTH, expand=True)

        self.du_res = StringVar()
        tk.Label(p, textvariable=self.du_res, bg=C["bg"], fg=C["ok"], font=FS).pack(anchor=W, padx=20)
        self._du_seen     = {}
        self._du_dup_keys = set()
        self._du_key_cols = []

    # ═══════════════════════  TAB 9 — PRESETS  ════════════════════════════════
    def _build_t9_presets(self):
        p = self._tab("💾 Presets")

        left  = tk.Frame(p, bg=C["bg"])
        left.pack(side=LEFT, fill=BOTH, expand=True, padx=(14,6), pady=10)
        right = tk.Frame(p, bg=C["bg"])
        right.pack(side=LEFT, fill=Y, padx=(0,14), pady=10, ipadx=10)

        # Left: preset list
        tk.Label(left, text="SAVED PRESETS", bg=C["bg"], fg=C["muted"], font=FS).pack(anchor=W, pady=(0,4))
        tk.Frame(left, bg=C["border"], height=1).pack(fill=X, pady=(0,6))
        lf = tk.Frame(left, bg=C["inp"])
        lf.pack(fill=BOTH, expand=True)
        self._pr_lb = tk.Listbox(lf, bg=C["inp"], fg=C["text"],
                                 selectbackground=C["sel"], selectforeground=C["accent"],
                                 font=FB, bd=0, relief="flat", activestyle="none")
        prs = ttk.Scrollbar(lf, orient=VERTICAL, command=self._pr_lb.yview)
        self._pr_lb.configure(yscrollcommand=prs.set)
        prs.pack(side=RIGHT, fill=Y)
        self._pr_lb.pack(fill=BOTH, expand=True)

        pb = tk.Frame(left, bg=C["bg"])
        _btn(pb,  "Load Selected",  cmd=self._pr_load).pack(side=LEFT, padx=(0,6))
        dbtn(pb,  "Delete",         cmd=self._pr_delete).pack(side=LEFT)
        pb.pack(fill=X, pady=8)
        self._pr_refresh()

        # Right: save form
        tk.Label(right, text="SAVE PRESET", bg=C["bg"], fg=C["muted"], font=FS).pack(anchor=W, pady=(0,4))
        tk.Frame(right, bg=C["border"], height=1).pack(fill=X, pady=(0,8))

        self.pr_name = StringVar(); self.pr_jtype = StringVar(value="convert")
        self.pr_in   = StringVar(); self.pr_out   = StringVar()
        self.pr_fmt  = StringVar(value="csv")

        def _row(label, var, w=26):
            r = tk.Frame(right, bg=C["bg"])
            tk.Label(r, text=label, bg=C["bg"], fg=C["muted"], font=FS, width=16, anchor=W).pack(side=LEFT)
            _ent(r, var=var, width=w).pack(side=LEFT)
            r.pack(fill=X, pady=3)

        _row("Preset name",   self.pr_name)

        jr = tk.Frame(right, bg=C["bg"])
        tk.Label(jr, text="Job type", bg=C["bg"], fg=C["muted"], font=FS, width=16, anchor=W).pack(side=LEFT)
        _combo(jr, self.pr_jtype, ["convert","batch","split","merge","filter"], width=14).pack(side=LEFT)
        jr.pack(fill=X, pady=3)

        ir = tk.Frame(right, bg=C["bg"])
        tk.Label(ir, text="Input path", bg=C["bg"], fg=C["muted"], font=FS, width=16, anchor=W).pack(side=LEFT)
        _ent(ir, var=self.pr_in, width=28).pack(side=LEFT, padx=(0,4))
        _btn(ir, "…", cmd=lambda: self.pr_in.set(filedialog.askopenfilename() or self.pr_in.get())).pack(side=LEFT)
        ir.pack(fill=X, pady=3)

        or_ = tk.Frame(right, bg=C["bg"])
        tk.Label(or_, text="Output path", bg=C["bg"], fg=C["muted"], font=FS, width=16, anchor=W).pack(side=LEFT)
        _ent(or_, var=self.pr_out, width=28).pack(side=LEFT, padx=(0,4))
        _btn(or_, "…", cmd=lambda: self.pr_out.set(filedialog.asksaveasfilename() or self.pr_out.get())).pack(side=LEFT)
        or_.pack(fill=X, pady=3)

        fr = tk.Frame(right, bg=C["bg"])
        tk.Label(fr, text="Format", bg=C["bg"], fg=C["muted"], font=FS, width=16, anchor=W).pack(side=LEFT)
        _combo(fr, self.pr_fmt, SUPPORTED_OUT, width=10).pack(side=LEFT)
        fr.pack(fill=X, pady=3)

        abtn(right, "💾  Save Preset",  cmd=self._pr_save).pack(fill=X, pady=10)
        abtn(right, "▶  Run Preset",    cmd=self._pr_run).pack(fill=X)


    # ═══════════════════════  TAB 10 — COLUMN TRIM  ══════════════════════════
    def _build_t10_column_trim(self):
        """Line-level keyword filter for any file format — single file or batch folder."""
        p = self._tab("🔤 Column Trim")

        self.ct_mode    = StringVar(value="file")
        self.ct_file    = StringVar()
        self.ct_folder  = StringVar()
        self.ct_out_fmt = StringVar(value="csv")
        self._ct_keywords = []   # list of {"var": StringVar, "frame": Frame}

        # ── SECTION 1: Source ─────────────────────────────────────────────────
        ss = self._section(p, "📂 Source — Select a Single File or a Folder for Batch Processing")

        # Mode radio row
        mf = tk.Frame(ss, bg=C["panel"])
        mf.pack(fill=X, padx=10, pady=(4, 6))
        tk.Label(mf, text="Input mode:", bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT, padx=(0, 12))
        ttk.Radiobutton(mf, text="Single File  (any format)",
                        variable=self.ct_mode, value="file",
                        style="TCheckbutton").pack(side=LEFT, padx=(0, 24))
        ttk.Radiobutton(mf, text="Folder — batch all files inside (any format)",
                        variable=self.ct_mode, value="folder",
                        style="TCheckbutton").pack(side=LEFT)

        # Container that switches between single-file row and folder row
        src_cont = tk.Frame(ss, bg=C["panel"])
        src_cont.pack(fill=X, padx=10, pady=2)

        # Single-file row
        self._ct_file_row = tk.Frame(src_cont, bg=C["panel"])
        tk.Label(self._ct_file_row, text="File",
                 bg=C["panel"], fg=C["muted"], font=FS, width=14, anchor=W).pack(side=LEFT)
        _ent(self._ct_file_row, var=self.ct_file, width=52).pack(side=LEFT, padx=(0, 4))
        _btn(self._ct_file_row, "Browse…",
             cmd=lambda: self.ct_file.set(
                 filedialog.askopenfilename(
                     title="Select file to filter",
                     filetypes=[("All files", "*.*")]
                 ) or self.ct_file.get()
             )).pack(side=LEFT)
        self._ct_file_row.pack(fill=X, pady=3)

        # Folder row (initially hidden)
        self._ct_folder_row = tk.Frame(src_cont, bg=C["panel"])
        tk.Label(self._ct_folder_row, text="Folder",
                 bg=C["panel"], fg=C["muted"], font=FS, width=14, anchor=W).pack(side=LEFT)
        _ent(self._ct_folder_row, var=self.ct_folder, width=52).pack(side=LEFT, padx=(0, 4))
        _btn(self._ct_folder_row, "Browse…",
             cmd=lambda: self.ct_folder.set(
                 filedialog.askdirectory(title="Select folder with files to filter") or self.ct_folder.get()
             )).pack(side=LEFT)
        self._ct_folder_row.pack_forget()   # hidden until "folder" mode selected

        def _ct_mode_changed(*_):
            if self.ct_mode.get() == "file":
                self._ct_folder_row.pack_forget()
                self._ct_file_row.pack(fill=X, pady=3)
            else:
                self._ct_file_row.pack_forget()
                self._ct_folder_row.pack(fill=X, pady=3)

        self.ct_mode.trace_add("write", _ct_mode_changed)

        # ── SECTION 2: Match Keywords ──────────────────────────────────────────
        ks = self._section(p, "🔤 Match Keywords — Any Line Containing ANY Keyword Is Kept (Case-Insensitive)")

        tk.Label(ks, bg=C["panel"], fg=C["muted"], font=FS, wraplength=860, justify=LEFT,
                 text=("Add one keyword per box.  Any line (row) that contains any of "
                       "the keywords — regardless of upper / lower case — is included in "
                       "the output.\n"
                       "Example: keyword  'US'  also matches  'BUSINESS', 'business', "
                       "'United States', 'us company', etc.")
                 ).pack(anchor=W, padx=12, pady=(0, 8))

        self._ct_kw_box = tk.Frame(ks, bg=C["panel"])
        self._ct_kw_box.pack(fill=X, padx=10, pady=4)

        # Start with 3 empty keyword boxes
        for _ in range(3):
            self._ct_add_keyword()

        kwbr = tk.Frame(ks, bg=C["panel"])
        _btn(kwbr, "+ Add Keyword", cmd=self._ct_add_keyword).pack(side=LEFT, padx=(0, 6))
        dbtn(kwbr, "Clear All",     cmd=self._ct_clear_keywords).pack(side=LEFT)
        kwbr.pack(fill=X, padx=10, pady=(2, 6))

        # ── SECTION 3: Output Format & Run ────────────────────────────────────
        os_ = self._section(p, "💾 Output Format & Run")

        fmr = tk.Frame(os_, bg=C["panel"])
        tk.Label(fmr, text="Output format",
                 bg=C["panel"], fg=C["muted"], font=FS, width=18, anchor=W).pack(side=LEFT)
        _combo(fmr, self.ct_out_fmt, list(SUPPORTED_OUT), width=10).pack(side=LEFT, padx=(0, 12))
        tk.Label(fmr,
                 text="(output filename gets an auto-incremented number:  name_1.csv, name_2.csv …)",
                 bg=C["panel"], fg=C["muted"], font=FS).pack(side=LEFT)
        fmr.pack(fill=X, padx=10, pady=6)

        abr = tk.Frame(os_, bg=C["panel"])
        abtn(abr, "▶  Start Filter & Save", cmd=self._run_column_trim).pack(side=LEFT, padx=(0, 8))
        abr.pack(fill=X, padx=10, pady=(4, 8))

        self._ct_status = StringVar(
            value="Ready — select source and keywords, then click ▶ Start Filter & Save")
        tk.Label(os_, textvariable=self._ct_status,
                 bg=C["panel"], fg=C["ok"],
                 font=("Segoe UI", 9, "bold"),
                 wraplength=880, justify=LEFT).pack(anchor=W, padx=12, pady=(0, 4))

        self._ct_pval = DoubleVar(value=0)
        ttk.Progressbar(os_, variable=self._ct_pval, maximum=100,
                        style="Dark.Horizontal.TProgressbar").pack(fill=X, padx=12, pady=(0, 6))

        # ── Live stats row ────────────────────────────────────────────────────
        self._ct_lines_var = StringVar(value="Scanned: –   │   Matched: –")
        self._ct_speed_var = StringVar(value="⚡  – lines/sec   │   – MB/s")
        self._ct_eta_var   = StringVar(value="⏱ ETA: –")
        gpu_label = ("🟢 GPU (cudf) ready" if _CUDF else
                     f"🔵 CPU {_HW['cpus']} cores · {_HW['tier']} ({_HW['avail_gb']} GB free RAM)")
        self._ct_gpu_var   = StringVar(value=gpu_label)

        sf = tk.Frame(os_, bg=C["panel"])
        sf.pack(fill=X, padx=10, pady=(2, 8))
        tk.Label(sf, textvariable=self._ct_gpu_var,
                 bg=C["panel"], fg=C["accent"],
                 font=("Consolas", 9, "bold"), width=24, anchor=W).pack(side=LEFT, padx=(0, 10))
        tk.Label(sf, textvariable=self._ct_speed_var,
                 bg=C["panel"], fg=C["warn"],
                 font=("Consolas", 9, "bold"), width=34, anchor=W).pack(side=LEFT, padx=(0, 10))
        tk.Label(sf, textvariable=self._ct_lines_var,
                 bg=C["panel"], fg=C["ok"],
                 font=("Consolas", 9, "bold"), width=34, anchor=W).pack(side=LEFT, padx=(0, 10))
        tk.Label(sf, textvariable=self._ct_eta_var,
                 bg=C["panel"], fg=C["muted"],
                 font=("Consolas", 9, "bold"), width=14, anchor=W).pack(side=LEFT)

        # ── Log area ──────────────────────────────────────────────────────────
        lf = _scroltext(p, height=10, mono=True)
        lf.pack(fill=X, padx=16, pady=(2, 12))
        self._ct_log_text = lf._text
        self._ct_log_text.configure(state="disabled")

    # ── keyword box helpers ───────────────────────────────────────────────────
    def _ct_add_keyword(self):
        """Append a new keyword input row to the Column Trim keyword list."""
        rf = tk.Frame(self._ct_kw_box, bg=C["panel"], pady=2)
        rf.pack(fill=X, pady=1)
        kv = StringVar()
        tk.Label(rf, text="Keyword:",
                 bg=C["panel"], fg=C["muted"], font=FS, width=10, anchor=W).pack(side=LEFT)
        _ent(rf, var=kv, width=46).pack(side=LEFT, padx=(3, 8))
        entry_data = {"var": kv, "frame": rf}
        def _rm():
            if entry_data in self._ct_keywords:
                self._ct_keywords.remove(entry_data)
            rf.destroy()
        dbtn(rf, "✕", cmd=_rm).pack(side=LEFT)
        self._ct_keywords.append(entry_data)

    def _ct_clear_keywords(self):
        """Remove all keyword rows."""
        for kw in list(self._ct_keywords):
            kw["frame"].destroy()
        self._ct_keywords.clear()

    # ── Column Trim log helper (thread-safe) ──────────────────────────────────
    def _ct_log_msg(self, msg):
        def _do():
            self._ct_log_text.configure(state="normal")
            self._ct_log_text.insert(END, f"[{time.strftime('%H:%M:%S')}] {msg}\n")
            self._ct_log_text.see(END)
            self._ct_log_text.configure(state="disabled")
        self.after(0, _do)

    # ── numbered output path helper ───────────────────────────────────────────
    @staticmethod
    def _ct_numbered_path(path_str, fmt=None):
        """Return first available: stem_1.ext, stem_2.ext, stem_3.ext …
        Strips any existing trailing _N from stem before numbering."""
        p = Path(path_str)
        ext = f".{fmt}" if fmt else p.suffix
        base = re.sub(r'_\d+$', '', p.stem)
        parent = p.parent
        parent.mkdir(parents=True, exist_ok=True)
        n = 1
        while True:
            candidate = parent / f"{base}_{n}{ext}"
            if not candidate.exists():
                return str(candidate)
            n += 1

    # ── core filter engine (ultra-fast) ──────────────────────────────────────
    @staticmethod
    def _ct_filter_file_fast(src_path, out_path, keywords, fmt,
                              on_stats=None, job_ctrl=None):
        """
        Ultra-fast parallel line filter with live stats.

        Strategy auto-selected by hardware:
          GPU  — NVIDIA RAPIDS cudf  (100+ GB/s, zero Python loop)
          CPU  — mmap + compiled OR-regex + all cores (~500 MB/s)
          CPUs — single-thread buffered fallback for structured formats

        on_stats(scanned, matched, bytes_done, elapsed) called every 50 K lines.
        Returns (matched_count, elapsed_seconds).
        """
        src_path = Path(src_path)
        ext      = src_path.suffix.lower()
        try:
            file_size = src_path.stat().st_size
        except OSError:
            file_size = 0

        # Compile all keywords into ONE OR-pattern — single pass per line
        escaped = [re.escape(k) for k in keywords if k.strip()]
        if not escaped:
            return 0, 0.0
        pattern_str = "|".join(escaped)
        pat_bytes   = re.compile(pattern_str.encode("utf-8", errors="replace"), re.IGNORECASE)
        pat_text    = re.compile(pattern_str, re.IGNORECASE)

        t0           = time.time()
        scanned      = [0]
        matched_c    = [0]
        bytes_done   = [0]
        REPORT_EVERY = 50_000
        READ_CHUNK   = _HW["read_chunk"]    # adaptive: 4 MB (LOW) / 32 MB (MID) / 128 MB (HIGH)
        WRITE_BUF    = _HW["write_buf"]     # adaptive: 512 KB (LOW) / 8 MB (MID) / 32 MB (HIGH)

        def _report():
            if on_stats:
                on_stats(scanned[0], matched_c[0], bytes_done[0], time.time() - t0)

        _STRUCTURED = {".csv", ".tsv", ".json", ".jsonl", ".xlsx"}

        # ══════════════════════════════════════════════════════════════════════
        # PATH 1 — GPU via RAPIDS cudf  (CSV / TSV only)
        # ══════════════════════════════════════════════════════════════════════
        if _CUDF and ext in {".csv", ".tsv"}:
            try:
                delim   = "\t" if ext == ".tsv" else ","
                gdf     = _cudf_mod.read_csv(str(src_path), delimiter=delim,
                                             dtype=str, low_memory=False)
                combined = gdf.iloc[:, 0].astype(str)
                for ci in range(1, len(gdf.columns)):
                    combined = combined + " " + gdf.iloc[:, ci].astype(str)
                mask       = combined.str.contains(pattern_str, case=False, regex=True)
                result_gdf = gdf[mask]
                n          = int(len(result_gdf))
                scanned[0] = int(len(gdf))
                matched_c[0] = n
                bytes_done[0] = file_size
                out_p = Path(out_path)
                out_p.parent.mkdir(parents=True, exist_ok=True)
                sep_out = "\t" if fmt == "tsv" else ","
                if fmt in ("csv", "tsv"):
                    result_gdf.to_csv(str(out_p), index=False, sep=sep_out)
                elif fmt == "jsonl":
                    with out_p.open("w", encoding="utf-8") as fh:
                        for row in result_gdf.to_pandas().to_dict(orient="records"):
                            fh.write(_fast_json_dumps(row) + "\n")
                elif fmt == "xlsx" and _OPENPYXL:
                    result_gdf.to_pandas().to_excel(str(out_p), index=False)
                else:
                    result_gdf.to_csv(str(out_p), index=False)
                elapsed = time.time() - t0
                if on_stats: on_stats(scanned[0], n, file_size, elapsed)
                return n, elapsed
            except Exception:
                pass   # fall through to CPU

        # ══════════════════════════════════════════════════════════════════════
        # PATH 2 — CPU structured (CSV / JSONL / JSON / XLSX)
        # Compiled text regex on joined row values, streaming write
        # ══════════════════════════════════════════════════════════════════════
        if ext in _STRUCTURED:
            try:
                out_p = Path(out_path)
                out_p.parent.mkdir(parents=True, exist_ok=True)

                def _stream():
                    for row in DataProcessor.iter_rows(src_path):
                        if job_ctrl and job_ctrl.is_cancelled(): return
                        if job_ctrl: job_ctrl.wait_if_paused()
                        scanned[0] += 1
                        search_str = " ".join(str(v) for v in row.values())
                        if pat_text.search(search_str):
                            matched_c[0] += 1
                            yield row
                        if scanned[0] % REPORT_EVERY == 0:
                            _report()

                DataProcessor.write_rows(_stream(), out_p, fmt)
                elapsed = time.time() - t0
                if on_stats: on_stats(scanned[0], matched_c[0], file_size, elapsed)
                return matched_c[0], elapsed
            except Exception:
                pass   # fall through to raw-text

        # ══════════════════════════════════════════════════════════════════════
        # PATH 3 — Raw text: mmap + compiled bytes regex + ALL CPU cores
        # Handles TXT, LOG, DB, any unknown format, and fallback
        # ══════════════════════════════════════════════════════════════════════
        out_p = Path(out_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)

        if file_size == 0:
            out_p.write_bytes(b"")
            return 0, 0.0

        N_WORKERS = _HW["workers"]   # adaptive: 1-2 (LOW) / all cores (MID/HIGH)

        def _split_chunks(size, n):
            """Split file into N chunks aligned on line boundaries."""
            if size <= READ_CHUNK or n == 1:
                return [(0, size)]
            approx = size // n
            chunks = []
            try:
                with open(src_path, "rb") as _f:
                    pos = 0
                    for _ in range(n - 1):
                        end = min(pos + approx, size)
                        _f.seek(end)
                        peek = _f.read(min(8192, size - end))
                        nl = peek.find(b"\n")
                        end = end + nl + 1 if nl >= 0 else end
                        chunks.append((pos, min(end, size)))
                        pos = min(end, size)
                        if pos >= size:
                            break
                    if pos < size:
                        chunks.append((pos, size))
            except Exception:
                return [(0, size)]
            return chunks if chunks else [(0, size)]

        chunks        = _split_chunks(file_size, N_WORKERS)
        n_chunks      = len(chunks)
        results       = [None]  * n_chunks
        chunk_scanned = [0]     * n_chunks
        chunk_bytes   = [0]     * n_chunks

        def _process_chunk(idx, c_start, c_end):
            local_matched  = []
            local_scanned  = 0
            try:
                with open(src_path, "rb") as _f:
                    _f.seek(c_start)
                    leftover = b""
                    pos = c_start
                    while pos < c_end:
                        if job_ctrl and job_ctrl.is_cancelled():
                            break
                        to_read = min(READ_CHUNK, c_end - pos)
                        block   = _f.read(to_read)
                        if not block:
                            break
                        data = leftover + block
                        pos += len(block)
                        if pos < c_end:
                            last_nl  = data.rfind(b"\n")
                            leftover = data[last_nl + 1:] if last_nl >= 0 else data
                            data     = data[:last_nl + 1] if last_nl >= 0 else b""
                        else:
                            leftover = b""
                        for line in data.split(b"\n"):
                            s = line.rstrip(b"\r")
                            if s:
                                local_scanned += 1
                                if pat_bytes.search(s):
                                    local_matched.append(
                                        s.decode("utf-8", errors="replace"))
                    # tail
                    if leftover:
                        s = leftover.rstrip(b"\r")
                        if s:
                            local_scanned += 1
                            if pat_bytes.search(s):
                                local_matched.append(
                                    s.decode("utf-8", errors="replace"))
            except Exception:
                pass
            results[idx]       = local_matched
            chunk_scanned[idx] = local_scanned
            chunk_bytes[idx]   = c_end - c_start
            scanned[0]    = sum(chunk_scanned)
            matched_c[0]  = sum(len(r) for r in results if r is not None)
            bytes_done[0] = sum(chunk_bytes)
            _report()

        with ThreadPoolExecutor(max_workers=N_WORKERS) as pool:
            futs = [pool.submit(_process_chunk, i, s, e)
                    for i, (s, e) in enumerate(chunks)]
            for f in as_completed(futs):
                try: f.result()
                except Exception: pass

        # Collect results in original order and write with large buffer
        n = sum(len(r) for r in results if r)
        delim_out = "\t" if fmt == "tsv" else ","
        if fmt == "txt":
            with out_p.open("w", encoding="utf-8", buffering=WRITE_BUF) as fh:
                for r in results:
                    if r:
                        fh.write("\n".join(r))
                        fh.write("\n")
        elif fmt in ("csv", "tsv"):
            with out_p.open("w", encoding="utf-8", newline="",
                            buffering=WRITE_BUF) as fh:
                w = csv.writer(fh, delimiter=delim_out)
                w.writerow(["line"])
                for r in results:
                    if r:
                        w.writerows([ln] for ln in r)
        elif fmt == "jsonl":
            with out_p.open("w", encoding="utf-8", buffering=WRITE_BUF) as fh:
                for r in results:
                    if r:
                        for ln in r:
                            fh.write(_fast_json_dumps({"line": ln}) + "\n")
        elif fmt == "xlsx" and _OPENPYXL:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("data")
            ws.append(["line"])
            for r in results:
                if r:
                    for ln in r:
                        ws.append([ln])
            wb.save(str(out_p))
        else:
            with out_p.open("w", encoding="utf-8", buffering=WRITE_BUF) as fh:
                for r in results:
                    if r:
                        fh.write("\n".join(r))
                        fh.write("\n")

        elapsed = time.time() - t0
        if on_stats: on_stats(scanned[0], n, file_size, elapsed)
        return n, elapsed

    # ── run action ────────────────────────────────────────────────────────────
    def _run_column_trim(self):
        """Triggered by ▶ Start Filter & Save button."""
        keywords = [k["var"].get().strip() for k in self._ct_keywords]
        keywords = [k for k in keywords if k]

        if not keywords:
            messagebox.showerror("No Keywords",
                "Enter at least one keyword to match.\n"
                "Lines containing any keyword will be kept.")
            return

        fmt  = self.ct_out_fmt.get()
        mode = self.ct_mode.get()

        # Reset stats display
        self._ct_lines_var.set("Scanned: –   │   Matched: –")
        self._ct_speed_var.set("⚡  – lines/sec   │   – MB/s")
        self._ct_eta_var.set("⏱ ETA: –")

        # Show which engine will be used
        if _CUDF and mode == "file" and Path(self.ct_file.get().strip()).suffix.lower() in {".csv", ".tsv"}:
            self._ct_gpu_var.set("🟢 GPU (cudf) — active")
        else:
            self._ct_gpu_var.set(f"🔵 CPU  {_CPU_COUNT} cores — active")

        def _make_on_stats(file_sz):
            """Returns a thread-safe stats callback for one file."""
            def _on_stats(sc, mt, bd, elapsed):
                if elapsed < 0.05:
                    return
                lps = sc / elapsed
                mbs = bd / (1024 * 1024 * max(elapsed, 1e-6))
                if file_sz > 0 and bd > 0:
                    pct    = min(98.0, bd / file_sz * 100)
                    remain = max(0, file_sz - bd)
                    eta_s  = int(remain / max(bd / elapsed, 1))
                else:
                    pct   = 0.0
                    eta_s = 0
                eta_str = (
                    f"{eta_s//3600}h{(eta_s%3600)//60}m{eta_s%60}s" if eta_s >= 3600
                    else f"{eta_s//60}m{eta_s%60}s" if eta_s >= 60
                    else f"{eta_s}s" if eta_s > 0 else "–"
                )
                self.after(0, lambda sc=sc, mt=mt, lps=lps, mbs=mbs,
                                         pct=pct, eta=eta_str: (
                    self._ct_lines_var.set(f"Scanned: {sc:,}   │   Matched: {mt:,}"),
                    self._ct_speed_var.set(f"⚡  {lps:,.0f} ln/s   │   {mbs:.1f} MB/s"),
                    self._ct_eta_var.set(f"⏱ {eta}"),
                    self._ct_pval.set(pct),
                    self._ct_status.set(
                        f"⏳  {sc:,} lines scanned  ·  {mt:,} matched  ·  {mbs:.1f} MB/s"),
                ))
            return _on_stats

        # ── Single file mode ──────────────────────────────────────────────────
        if mode == "file":
            src = self.ct_file.get().strip()
            if not src:
                messagebox.showerror("No File", "Select an input file."); return
            if not Path(src).exists():
                messagebox.showerror("Not Found", f"File not found:\n{src}"); return

            stem = Path(src).stem
            filetypes = [(f"{fmt.upper()} files", f"*.{fmt}"), ("All files", "*.*")]
            raw_out = filedialog.asksaveasfilename(
                title="Save filtered output as…",
                initialfile=f"{stem}_1.{fmt}",
                defaultextension=f".{fmt}",
                filetypes=filetypes
            )
            if not raw_out:
                return

            out_path  = self._ct_numbered_path(raw_out, fmt)
            file_size = Path(src).stat().st_size
            on_stats  = _make_on_stats(file_size)

            if not self._start_job("Column Trim"):
                return

            self._ct_pval.set(2)
            self._ct_status.set("⏳ Starting filter engine…")
            self._ct_log_msg("▶  Column Trim — Single File")
            self._ct_log_msg(f"   Source:   {src}")
            self._ct_log_msg(f"   Size:     {file_size/1024/1024:.1f} MB")
            self._ct_log_msg(f"   Keywords: {', '.join(repr(k) for k in keywords)}")
            self._ct_log_msg(f"   Output:   {out_path}")
            self._ct_log_msg(f"   Engine:   {'GPU (cudf)' if _CUDF else f'CPU {_CPU_COUNT} cores + 32 MB chunks'}")
            self._log(f"Column Trim: {Path(src).name}  keywords={keywords}")

            def task():
                try:
                    n, elapsed = self._ct_filter_file_fast(
                        src, out_path, keywords, fmt,
                        on_stats=on_stats, job_ctrl=self._job_ctrl)
                    out_name = Path(out_path).name
                    mb_s = file_size / (1024*1024*max(elapsed, 0.001))
                    summary = (
                        f"✓ Done — {n:,} rows matched  │  "
                        f"{mb_s:.1f} MB/s avg  │  {elapsed:.1f}s total"
                    )
                    self.after(0, lambda: (
                        self._ct_pval.set(100),
                        self._ct_status.set(summary),
                        self._ct_lines_var.set(
                            f"Scanned: complete   │   Matched: {n:,}"),
                        self._ct_eta_var.set("⏱ Done"),
                        self._ct_log_msg(summary),
                        self._ct_log_msg(f"   Saved → {out_path}"),
                        self._finish_job(f"Column Trim done: {n:,} rows"),
                        messagebox.showinfo("Column Trim Complete",
                            f"Filter complete!\n\n"
                            f"Matched rows : {n:,}\n"
                            f"Avg speed    : {mb_s:.1f} MB/s\n"
                            f"Time         : {elapsed:.1f}s\n"
                            f"Saved to     : {out_path}"),
                    ))
                except Exception as ex:
                    err = str(ex)
                    self.after(0, lambda: (
                        self._ct_status.set(f"Error: {err}"),
                        self._ct_log_msg(f"✗ Error: {err}"),
                        self._finish_job("Column Trim: error", False),
                        messagebox.showerror("Column Trim Error", err),
                    ))

            self._run_bg(task)

        # ── Batch / folder mode ───────────────────────────────────────────────
        else:
            folder = self.ct_folder.get().strip()
            if not folder:
                messagebox.showerror("No Folder", "Select an input folder."); return
            if not Path(folder).is_dir():
                messagebox.showerror("Invalid Folder",
                    f"Folder not found:\n{folder}"); return

            out_folder_str = filedialog.askdirectory(
                title="Select destination folder to save filtered files")
            if not out_folder_str:
                return
            out_folder = Path(out_folder_str)

            if not self._start_job("Column Trim Batch"):
                return

            self._ct_pval.set(0)
            self._ct_status.set("⏳ Batch filtering…  scanning folder")
            self._ct_log_msg("▶  Column Trim — Batch Folder Mode")
            self._ct_log_msg(f"   Source folder: {folder}")
            self._ct_log_msg(f"   Dest folder:   {out_folder}")
            self._ct_log_msg(f"   Keywords: {', '.join(repr(k) for k in keywords)}")
            self._ct_log_msg(f"   Engine:   CPU {_CPU_COUNT} cores per file")
            self._log(f"Column Trim Batch: {folder}  keywords={keywords}")

            def task():
                try:
                    files = sorted(fp for fp in Path(folder).iterdir() if fp.is_file())
                    total = len(files)

                    if total == 0:
                        self.after(0, lambda: (
                            self._ct_status.set("No files found in the selected folder."),
                            self._ct_log_msg("No files found — nothing to process."),
                            self._finish_job("Column Trim batch: no files"),
                        ))
                        return

                    total_size = sum(fp.stat().st_size for fp in files)
                    self._ct_log_msg(
                        f"   Found {total} file(s)  │  "
                        f"{total_size/1024/1024:.1f} MB total")
                    total_rows    = 0
                    total_elapsed = 0.0
                    done          = 0

                    for fp in files:
                        if self._job_ctrl.is_cancelled():
                            self._ct_log_msg("✕ Cancelled by user")
                            break
                        self._job_ctrl.wait_if_paused()

                        file_idx  = done + 1
                        fname     = fp.name
                        fsize     = fp.stat().st_size
                        on_stats  = _make_on_stats(fsize)
                        self._ct_log_msg(
                            f"  [{file_idx}/{total}]  {fname}  "
                            f"({fsize/1024/1024:.1f} MB)")

                        out_path = self._ct_numbered_path(
                            str(out_folder / fp.stem), fmt)

                        n, elapsed = self._ct_filter_file_fast(
                            str(fp), out_path, keywords, fmt,
                            on_stats=on_stats, job_ctrl=self._job_ctrl)
                        total_rows    += n
                        total_elapsed += elapsed
                        done          += 1

                        file_mb_s = fsize / (1024*1024*max(elapsed, 0.001))
                        pct       = int(done / total * 100)
                        out_name  = Path(out_path).name
                        self._ct_log_msg(
                            f"        → {n:,} rows kept  "
                            f"│  {file_mb_s:.1f} MB/s  "
                            f"│  {elapsed:.1f}s  →  {out_name}")
                        self.after(0, lambda pv=pct, d=done, nm=fname: (
                            self._ct_pval.set(pv),
                            self._ct_status.set(
                                f"[{d}/{total}]  {nm}  —  {total_rows:,} rows matched so far"),
                        ))

                    avg_mb_s = total_size / (1024*1024*max(total_elapsed, 0.001))
                    summary  = (
                        f"✓ Batch done — {done}/{total} files  │  "
                        f"{total_rows:,} matching rows  │  "
                        f"{avg_mb_s:.1f} MB/s avg  │  {total_elapsed:.1f}s total"
                    )
                    out_folder_disp = str(out_folder)
                    self.after(0, lambda: (
                        self._ct_pval.set(100),
                        self._ct_status.set(summary),
                        self._ct_lines_var.set(
                            f"Scanned: complete   │   Matched: {total_rows:,}"),
                        self._ct_eta_var.set("⏱ Done"),
                        self._ct_log_msg(summary),
                        self._finish_job(
                            f"Column Trim batch: {done} files, {total_rows:,} rows"),
                        messagebox.showinfo("Column Trim Batch Complete",
                            f"Batch filter complete!\n\n"
                            f"Files processed : {done} / {total}\n"
                            f"Total rows saved: {total_rows:,}\n"
                            f"Avg speed       : {avg_mb_s:.1f} MB/s\n"
                            f"Total time      : {total_elapsed:.1f}s\n"
                            f"Saved to folder : {out_folder_disp}"),
                    ))
                except Exception as ex:
                    err = str(ex)
                    self.after(0, lambda: (
                        self._ct_status.set(f"Error: {err}"),
                        self._ct_log_msg(f"✗ Error: {err}"),
                        self._finish_job("Column Trim batch: error", False),
                        messagebox.showerror("Column Trim Error", err),
                    ))

            self._run_bg(task)


    # ═════════════════════════════════════════════════════════════════════════
    #  JOB HELPERS
    # ═════════════════════════════════════════════════════════════════════════
    def _set_job_btns(self, on):
        st = "normal" if on else "disabled"
        for b in (self._pbtn_p, self._pbtn_r, self._pbtn_c): b.configure(state=st)

    def _start_job(self, name):
        if self._job_running:
            messagebox.showwarning("Busy","A job is already running. Cancel it first.")
            return False
        self._job_ctrl = JobControl()
        self._job_running = True
        self._pval.set(0); self._pstat.set(f"{name}…")
        self._peta.set(""); self._pmeta.set("starting")
        self._set_job_btns(True)
        return True

    def _finish_job(self, msg, ok=True):
        self._job_running = False
        self._pstat.set(msg); self._peta.set(""); self._pmeta.set("")
        self._set_job_btns(False)
        if ok: self._pval.set(100)

    def _job_pause(self):
        if self._job_running: self._job_ctrl.pause(); self._pmeta.set("⏸ Paused")
    def _job_resume(self):
        if self._job_running: self._job_ctrl.resume(); self._pmeta.set("▶ Running")
    def _job_cancel(self):
        if self._job_running: self._job_ctrl.cancel(); self._pmeta.set("✕ Cancelling…")

    def _uprog(self, label, cur, total, elapsed, meta=""):
        if total and total > 0:
            pct = min(100.0, cur / total * 100)
            self._pval.set(pct)
            if elapsed and cur > 0:
                eta = (total - cur) / (cur / max(elapsed, 1e-9))
                self._peta.set(self._eta_str(int(eta)))
            self._pstat.set(f"{label}: {cur:,} / {total:,}")
        else:
            self._pval.set((self._pval.get() + 2) % 100)
            self._pstat.set(f"{label}: {cur:,}")
            self._peta.set("estimating…")
        if meta: self._pmeta.set(meta)

    @staticmethod
    def _eta_str(s):
        h, r = divmod(s, 3600); m, s = divmod(r, 60)
        return f"ETA {h}h{m}m{s}s" if h else (f"ETA {m}m{s}s" if m else f"ETA {s}s")

    def _run_bg(self, fn):
        threading.Thread(target=fn, daemon=True).start()

    # ─── log ─────────────────────────────────────────────────────────────────
    def _log(self, msg):
        self._log_q.put(f"[{time.strftime('%H:%M:%S')}] {msg}")

    def _poll_log(self):
        while True:
            try:
                msg = self._log_q.get_nowait()
            except queue.Empty:
                break
            self._logw.configure(state="normal")
            self._logw.insert(END, msg + "\n")
            self._logw.see(END)
            self._logw.configure(state="disabled")
        self.after(100, self._poll_log)

    def _clear_log(self):
        self._logw.configure(state="normal")
        self._logw.delete("1.0", END)
        self._logw.configure(state="disabled")

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 1 — CONVERT ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _scan_convert(self):
        """Scan-only: check file for threats and report in Job Log + UI."""
        src = self.cv_in.get().strip()
        if not src or not Path(src).exists():
            messagebox.showerror("Not found", "Select a valid input file"); return
        self._log("═" * 60)
        self._log(f"🛡️ SECURITY SCAN: {Path(src).name}")
        self._log("═" * 60)
        scan = DataProcessor.scan_file(src, log=self._log)
        sev = scan["severity"]
        sev_icons = {"clean": "✅", "low": "⚠️", "medium": "🔶", "high": "🔴"}
        icon = sev_icons.get(sev, "?")
        sz = scan["size"]
        if sz < 1024: szl = f"{sz} B"
        elif sz < 1024**2: szl = f"{sz/1024:.1f} KB"
        elif sz < 1024**3: szl = f"{sz/1024**2:.1f} MB"
        else: szl = f"{sz/1024**3:.2f} GB"
        self._log(f"  File: {Path(src).name} ({szl})")
        self._log(f"  Verdict: {icon} {sev.upper()} — {scan['threat_count']} threat(s)")
        if scan["threats"]:
            self._log("  ─── Threats Found ───")
            for t in scan["threats"]:
                lvl = t["severity"].upper()
                self._log(f"    [{lvl}] {t['type']}: {t['detail']}")
            self._log("  ─── Recommendation ───")
            if sev == "high":
                self._log("    🔴 HIGH RISK — Enable sanitize to safely extract data")
                self._log("    ⚠️  Contains: macros/scripts/shells/injections")
                self._log("    ✓  Safe convert will strip ALL active content")
            elif sev == "medium":
                self._log("    🔶 MEDIUM RISK — Sanitize recommended")
                self._log("    ✓  Active content will be neutralized on convert")
            else:
                self._log("    ⚠️  LOW RISK — Minor suspicious patterns")
        else:
            self._log("  ✅ No threats detected — file is clean")
            self._log("  ✓  Safe to convert at full speed")
        self._log("═" * 60)

        # Update UI scan result
        if sev == "clean":
            self.cv_scan_result.set(f"✅ CLEAN — No threats detected. Safe to convert.")
        else:
            self.cv_scan_result.set(
                f"{icon} {sev.upper()} — {scan['threat_count']} threat(s) found. "
                f"Enable 🛡️ Sanitize to safely extract clean data.")

    def _run_convert(self):
        src = self.cv_in.get().strip()
        outd= self.cv_outd.get().strip()
        if not src or not outd:
            messagebox.showerror("Missing","Input file and output folder required"); return
        if not Path(src).exists():
            messagebox.showerror("Not found", src); return
        if not self._start_job("Convert"): return
        fmt  = self.cv_fmt.get()
        name = self.cv_name.get().strip() or f"{Path(src).stem}_clean"
        out  = Path(outd) / f"{name}.{fmt}"
        sanitize = self.cv_sanitize.get()

        def task():
            try:
                total = None
                if self._prescan.get():
                    self._log("Pre-scanning row count…")
                    total = DataProcessor.count_rows(src, self._log, self._job_ctrl)
                    self._log(f"Pre-scan: {total:,} rows")

                def prog(proc, kept, tot, el):
                    self.after(0, lambda: self._uprog("Convert", proc, total, el, f"kept {kept:,}"))

                n, scan = DataProcessor.convert_safe(
                    src, out, fmt,
                    trim=self.cv_trim.get(),
                    drop_empty=self.cv_drop.get(),
                    sanitize=sanitize,
                    log=self._log,
                    control=self._job_ctrl,
                    on_progress=prog, total=total)
                sev = scan["severity"]
                sev_icons = {"clean": "✅", "low": "⚠️", "medium": "🔶", "high": "🔴"}
                icon = sev_icons.get(sev, "")
                threat_msg = f" │ {icon} {scan['threat_count']} threats stripped" if scan["threat_count"] else " │ ✅ clean"
                self._log(f"Done: {n:,} rows → {out}{threat_msg}")
                self.after(0, lambda: (
                    self._finish_job("Convert: done"),
                    self.cv_res.set(f"✓  {n:,} rows → {out}{threat_msg}"),
                    self.cv_scan_result.set(
                        f"{icon} {sev.upper()} — {scan['threat_count']} threat(s) "
                        f"{'stripped' if sanitize else 'detected'}"
                        if scan['threat_count'] else "✅ CLEAN — No threats")))
            except Exception as ex:
                self._log(f"ERROR: {ex}")
                self.after(0, lambda: self._finish_job("Convert: error", False))

        self._run_bg(task)

    def _preview_convert(self):
        src = self.cv_in.get().strip()
        if not src: return
        self.bv_path.set(src)
        self._big_row.set(1); self._big_chunk.set(500)
        self._nb.select(2)
        self._bv_open_path(src) if not self._big_engine or str(self._big_engine.path) != src else self._bv_load()

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 2 — BATCH ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _run_batch(self):
        src = self.bt_in.get().strip(); outd = self.bt_out.get().strip()
        if not src or not outd:
            messagebox.showerror("Missing","Input and output directories required"); return
        if not Path(src).is_dir():
            messagebox.showerror("Not found","Input directory not found"); return
        if not self._start_job("Batch"): return
        fmt   = self.bt_fmt.get()
        wkrs  = max(1, self.bt_wkrs.get())
        rec   = self.bt_rec.get()
        sanitize = self.bt_sanitize.get()

        def task():
            try:
                src_p = Path(src)
                files = list(src_p.rglob("*") if rec else src_p.glob("*"))
                files = [f for f in files if f.is_file() and f.suffix.lower() in SUPPORTED_IN]
                if not files:
                    self._log("No supported files found")
                    self.after(0, lambda: self._finish_job("Batch: no files", False)); return
                self._log(f"Batch: {len(files)} files, {wkrs} workers, sanitize={'ON' if sanitize else 'OFF'}")
                self._log("═" * 60)
                t0 = time.time(); done = success = 0
                total_threats = 0

                def do_one(fp):
                    op = Path(outd) / f"{fp.stem}_clean.{fmt}"
                    try:
                        n, scan = DataProcessor.convert_safe(
                            fp, op, fmt,
                            trim=self.bt_trim.get(),
                            drop_empty=self.bt_drop.get(),
                            sanitize=sanitize,
                            log=self._log,
                            control=self._job_ctrl)
                        return fp, True, None, scan
                    except Exception as ex:
                        return fp, False, ex, None

                with ThreadPoolExecutor(max_workers=wkrs) as ex:
                    futs = {ex.submit(do_one, fp): fp for fp in files}
                    for fut in as_completed(futs):
                        fp, ok, err, scan = fut.result()
                        done += 1; success += ok
                        if scan and scan["threat_count"]:
                            total_threats += scan["threat_count"]
                            sev = scan["severity"]
                            sev_icons = {"clean": "✅", "low": "⚠️", "medium": "🔶", "high": "🔴"}
                            self._log(f"  {sev_icons.get(sev, '?')} {fp.name}: "
                                      f"{scan['threat_count']} threats [{sev.upper()}] → sanitized")
                        elif ok:
                            self._log(f"  ✅ {fp.name}: clean → converted")
                        else:
                            self._log(f"  ✗ {fp.name}: {err}")
                        el = time.time() - t0
                        rate = done / max(el, 0.001)
                        self.after(0, lambda d=done, t=len(files), e=el, r=rate, s=success:
                                   self._uprog("Batch", d, t, e, f"{s}/{len(files)} ok  {r:.1f}/s"))

                self._log("═" * 60)
                threat_note = f" │ {total_threats} threats stripped" if total_threats else " │ all clean"
                msg = f"Batch done: {success}/{len(files)}{threat_note}"
                self._log(msg)
                self.after(0, lambda: (self._finish_job("Batch: done"),
                                       self.bt_res.set(f"✓  {msg}")))
            except Exception as ex:
                self._log(f"ERROR: {ex}")
                self.after(0, lambda: self._finish_job("Batch: error", False))

        self._run_bg(task)

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 4 — SPLIT ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _run_split(self):
        src = self.sp_in.get().strip(); outd = self.sp_outd.get().strip()
        if not src or not outd:
            messagebox.showerror("Missing","Input file and output folder required"); return
        if not Path(src).exists():
            messagebox.showerror("Not found", src); return
        if not self._start_job("Split"): return
        fmt = self.sp_fmt.get(); rps = max(1, self.sp_rows.get())

        def task():
            try:
                total = None
                if self._prescan.get():
                    total = DataProcessor.count_rows(src, self._log, self._job_ctrl)
                    self._log(f"Pre-scan: {total:,} rows")

                def prog(proc, kept, tot, el):
                    self.after(0, lambda: self._uprog("Split", proc, total, el))

                proc, files = DataProcessor.split(src, outd, rps, fmt,
                                                   trim=self.sp_trim.get(),
                                                   drop_empty=self.sp_drop.get(),
                                                   log=self._log,
                                                   control=self._job_ctrl,
                                                   on_progress=prog, total=total)
                msg = f"{proc:,} rows → {len(files)} files"
                self._log(f"Split done: {msg}")
                self.after(0, lambda: (self._finish_job("Split: done"),
                                       self.sp_res.set(f"✓  {msg}")))
            except Exception as ex:
                self._log(f"ERROR: {ex}")
                self.after(0, lambda: self._finish_job("Split: error", False))

        self._run_bg(task)

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 5 — MERGE ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _mg_add_files(self):
        ps = filedialog.askopenfilenames(
            filetypes=[("Supported","*.csv *.json *.jsonl *.txt *.xlsx"),("All","*.*")])
        for p in ps:
            if p not in self._mg_files:
                self._mg_files.append(p); self._mg_lb.insert(END, p)
        self._mg_cnt.set(f"{len(self._mg_files)} files queued")

    def _mg_add_folder(self):
        folder = filedialog.askdirectory()
        if not folder: return
        added = 0
        for fp in sorted(Path(folder).glob("*")):
            if fp.is_file() and fp.suffix.lower() in SUPPORTED_IN:
                s = str(fp)
                if s not in self._mg_files:
                    self._mg_files.append(s); self._mg_lb.insert(END, s); added += 1
        self._log(f"Added {added} files"); self._mg_cnt.set(f"{len(self._mg_files)} files queued")

    def _mg_up(self):
        sel = list(self._mg_lb.curselection())
        if not sel or sel[0] == 0: return
        for i in sel:
            self._mg_files[i-1], self._mg_files[i] = self._mg_files[i], self._mg_files[i-1]
        self._mg_lb.delete(0, END)
        for f in self._mg_files: self._mg_lb.insert(END, f)

    def _mg_down(self):
        sel = list(self._mg_lb.curselection())
        if not sel or sel[-1] >= len(self._mg_files)-1: return
        for i in reversed(sel):
            self._mg_files[i], self._mg_files[i+1] = self._mg_files[i+1], self._mg_files[i]
        self._mg_lb.delete(0, END)
        for f in self._mg_files: self._mg_lb.insert(END, f)

    def _mg_remove(self):
        for i in reversed(list(self._mg_lb.curselection())):
            self._mg_lb.delete(i); self._mg_files.pop(i)
        self._mg_cnt.set(f"{len(self._mg_files)} files queued")

    def _mg_clear(self):
        self._mg_files = []; self._mg_lb.delete(0, END)
        self._mg_cnt.set("0 files queued")

    def _run_merge(self):
        if not self._mg_files:
            messagebox.showerror("Empty","Add files to merge list"); return
        out = self.mg_out.get().strip()
        if not out:
            messagebox.showerror("Missing","Select output file"); return
        if not self._start_job("Merge"): return
        fmt = self.mg_fmt.get(); files = list(self._mg_files)

        def task():
            try:
                total = None
                if self._prescan.get():
                    total = sum(DataProcessor.count_rows(f, self._log, self._job_ctrl) for f in files)
                    self._log(f"Pre-scan total: {total:,} rows across {len(files)} files")

                def prog(proc, kept, tot, el):
                    self.after(0, lambda: self._uprog("Merge", proc, total, el))

                n = DataProcessor.merge(files, out, fmt, log=self._log,
                                        control=self._job_ctrl, on_progress=prog, total=total)
                self._log(f"Merge done: {n:,} rows → {out}")
                self.after(0, lambda: (self._finish_job("Merge: done"),
                                       self.mg_res.set(f"✓  {n:,} rows → {out}")))
            except Exception as ex:
                self._log(f"ERROR: {ex}")
                self.after(0, lambda: self._finish_job("Merge: error", False))

        self._run_bg(task)

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 6 — PROFILE ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _run_profile(self):
        src = self.pf_in.get().strip()
        if not src or not Path(src).exists():
            messagebox.showerror("Not found","Input file not found"); return
        sample = self.pf_sample.get() or 0
        if sample == 0: sample = int(1e12)
        self._log(f"Profiling {Path(src).name} sample={sample:,}")

        def task():
            try:
                stats = DataProcessor.profile_columns(src, sample=sample, log=self._log)
                rows = []
                for col, s in stats.items():
                    tot = s["count"]; ep = f"{100*s['empty']/max(tot,1):.1f}%"
                    types = {k: s[k] for k in ("int","float","date","bool","text")}
                    inf = max(types, key=types.get) if any(types.values()) else "empty"
                    uniq = len(s["sample"])
                    samp = ", ".join(sorted(s["sample"])[:5])
                    rows.append((col, tot, ep, inf,
                                 s["min_len"] if s["min_len"] is not None else 0,
                                 s["max_len"], uniq, samp))
                self._pf_data = rows

                def update():
                    self._pf_tv.delete(*self._pf_tv.get_children())
                    for r in rows: self._pf_tv.insert("", END, values=r)
                    self._log(f"Profile: {len(rows)} columns")

                self.after(0, update)
            except Exception as ex:
                self._log(f"Profile error: {ex}")

        self._run_bg(task)

    def _export_profile(self):
        if not self._pf_data:
            messagebox.showinfo("Empty","Run profile first"); return
        fmt = self.pf_exp_fmt.get()
        ext = fmt if fmt != "jsonl" else "jsonl"
        ftypes = [(fmt.upper(), f"*.{ext}"), ("All", "*.*")]
        p = filedialog.asksaveasfilename(defaultextension=f".{ext}",
                filetypes=ftypes)
        if not p: return
        try:
            headers = ["Column","Count","Empty %","Inferred Type",
                       "Min Len","Max Len","Unique ≤5","Sample Values"]
            rows = [{headers[j]: v for j, v in enumerate(row)}
                    for row in self._pf_data]
            n = DataProcessor.write_rows(iter(rows), p, fmt, log=self._log)
            self._log(f"Profile exported: {n} rows → {p}")
        except Exception as ex:
            self._log(f"Export error: {ex}")

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 7 — FILTER ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _run_filter(self):
        src = self.fl_in.get().strip(); out = self.fl_out.get().strip()
        if not src or not out:
            messagebox.showerror("Missing","Input and output files required"); return
        if not Path(src).exists():
            messagebox.showerror("Not found", src); return
        conds = self._fl_get_conds()
        if not conds:
            messagebox.showerror("No conditions","Add at least one filter condition"); return
        if not self._start_job("Filter"): return
        fmt = self.fl_fmt.get()

        def task():
            try:
                total = None
                if self._prescan.get():
                    total = DataProcessor.count_rows(src, self._log, self._job_ctrl)
                gen = DataProcessor.filter_rows(src, conds, log=self._log, control=self._job_ctrl)
                n   = DataProcessor.write_rows(gen, out, fmt, log=self._log)
                self._log(f"Filter done: {n:,} rows → {out}")
                self.after(0, lambda: (self._finish_job("Filter: done"),
                                       self.fl_res.set(f"✓  {n:,} matching rows → {out}")))
            except Exception as ex:
                self._log(f"Filter error: {ex}")
                self.after(0, lambda: self._finish_job("Filter: error", False))

        self._run_bg(task)

    def _preview_filter(self):
        src = self.fl_in.get().strip()
        if not src or not Path(src).exists():
            messagebox.showerror("Not found","Input file not found"); return
        conds = self._fl_get_conds()
        if not conds:
            messagebox.showerror("No conditions","Add at least one condition"); return

        def task():
            rows = []
            for row in DataProcessor.filter_rows(src, conds, log=self._log):
                rows.append(row)
                if len(rows) >= 200: break
            lines = [_fast_json_dumps(r) for r in rows]
            def update():
                self._fl_prev_text.configure(state="normal")
                self._fl_prev_text.delete("1.0", END)
                self._fl_prev_text.insert("1.0", "\n".join(lines) or "No matching rows")
                self._fl_prev_text.configure(state="disabled")
            self.after(0, update)
            self._log(f"Filter preview: {len(rows)} rows")

        self._run_bg(task)

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 8 — DUPLICATES ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _get_keys(self):
        return [c.strip() for c in self.du_keys.get().split(",") if c.strip()]

    def _run_scan_dupes(self):
        src = self.du_in.get().strip()
        if not src or not Path(src).exists():
            messagebox.showerror("Not found","Input file not found"); return
        keys = self._get_keys()
        if not keys:
            messagebox.showerror("No keys","Enter key column names"); return
        self._log(f"Scanning duplicates on: {keys}")

        def task():
            try:
                def prog(tot, dupes, _, el):
                    self.after(0, lambda: self._uprog("Dup Scan", tot, None, el, f"{dupes} dup groups"))

                seen, dup_keys, unique, total = DataProcessor.find_duplicates(
                    src, keys, log=self._log, on_progress=prog)

                self._du_seen     = seen
                self._du_dup_keys = dup_keys
                self._du_key_cols = keys

                dup_items = sorted([(k, v) for k, v in seen.items() if v > 1],
                                   key=lambda x: -x[1])

                def update():
                    self._du_tv.delete(*self._du_tv.get_children())
                    for k, v in dup_items[:5000]:
                        self._du_tv.insert("", END, values=(" | ".join(k), v))
                    self.du_res.set(
                        f"✓  Total rows: {total:,}   Unique: {unique:,}   "
                        f"Duplicate groups: {len(dup_keys):,}"
                    )
                    self._log(f"Dup scan done. {len(dup_keys)} dup groups in {total:,} rows")

                self.after(0, update)
            except Exception as ex:
                self._log(f"Dup error: {ex}")

        self._run_bg(task)

    def _export_dupes(self, mode):
        src = self.du_in.get().strip()
        if not src or not Path(src).exists():
            messagebox.showerror("Not found","Input file not found"); return
        keys = self._get_keys()
        if not keys:
            messagebox.showerror("No keys","Enter key columns first"); return
        if not self._du_seen:
            messagebox.showinfo("Scan first","Run Scan first to build the duplicate index"); return
        out = (self.du_odup if mode == "dup" else self.du_ouni).get().strip()
        if not out:
            messagebox.showerror("Missing","Select an output file path"); return
        if not self._start_job("Export"): return
        fmt       = self.du_fmt.get()
        dup_keys  = self._du_dup_keys
        key_cols  = self._du_key_cols

        def task():
            try:
                if mode == "dup":
                    gen = (row for row in DataProcessor.iter_rows(src)
                           if _fast_hash("|".join(str(row.get(c,"")) for c in key_cols)) in dup_keys)
                else:
                    emitted: set = set()
                    def _uni():
                        for row in DataProcessor.iter_rows(src):
                            k = _fast_hash("|".join(str(row.get(c,"")) for c in key_cols))
                            if k not in dup_keys and k not in emitted:
                                emitted.add(k)
                                yield row
                    gen = _uni()

                n = DataProcessor.write_rows(gen, out, fmt, log=self._log)
                label = "duplicates" if mode == "dup" else "unique rows"
                self._log(f"Export {label}: {n:,} rows → {out}")
                self.after(0, lambda: self._finish_job("Export: done"))
            except Exception as ex:
                self._log(f"Export error: {ex}")
                self.after(0, lambda: self._finish_job("Export: error", False))

        self._run_bg(task)

    # ═════════════════════════════════════════════════════════════════════════
    #  TAB 9 — PRESETS ACTIONS
    # ═════════════════════════════════════════════════════════════════════════
    def _pr_refresh(self):
        self._pr_lb.delete(0, END)
        for name in sorted(self._presets): self._pr_lb.insert(END, name)

    def _pr_save(self):
        name = self.pr_name.get().strip()
        if not name:
            messagebox.showerror("Missing","Enter a preset name"); return
        self._presets[name] = {
            "job_type": self.pr_jtype.get(),
            "input":    self.pr_in.get().strip(),
            "output":   self.pr_out.get().strip(),
            "format":   self.pr_fmt.get(),
        }
        save_presets(self._presets); self._pr_refresh()
        self._log(f"Preset saved: {name}")

    def _pr_load(self):
        sel = self._pr_lb.curselection()
        if not sel: return
        name = self._pr_lb.get(sel[0])
        p = self._presets.get(name, {})
        self.pr_name.set(name); self.pr_jtype.set(p.get("job_type","convert"))
        self.pr_in.set(p.get("input","")); self.pr_out.set(p.get("output",""))
        self.pr_fmt.set(p.get("format","csv")); self._log(f"Preset loaded: {name}")

    def _pr_delete(self):
        sel = self._pr_lb.curselection()
        if not sel: return
        name = self._pr_lb.get(sel[0])
        if messagebox.askyesno("Delete", f"Delete preset '{name}'?"):
            self._presets.pop(name, None); save_presets(self._presets)
            self._pr_refresh(); self._log(f"Preset deleted: {name}")

    def _pr_run(self):
        name = self.pr_name.get().strip()
        if not name or name not in self._presets:
            messagebox.showerror("Not found","Load a preset first"); return
        p   = self._presets[name]
        jt  = p.get("job_type","convert")
        src = p.get("input","")
        out = p.get("output","")
        fmt = p.get("format","csv")
        if not src or not Path(src).exists():
            messagebox.showerror("Not found",f"Input not found:\n{src}"); return
        if not self._start_job(f"Preset:{name}"): return

        def task():
            try:
                if jt == "convert":
                    n = DataProcessor.convert(src, out, fmt, log=self._log, control=self._job_ctrl)
                    self._log(f"Preset convert done: {n:,} rows")
                elif jt == "split":
                    proc, files = DataProcessor.split(src, str(Path(out).parent), 100000, fmt,
                                                      log=self._log, control=self._job_ctrl)
                    self._log(f"Preset split done: {proc:,} rows → {len(files)} files")
                elif jt == "batch":
                    files = [f for f in Path(src).glob("*")
                             if f.is_file() and f.suffix.lower() in SUPPORTED_IN]
                    for fp in files:
                        DataProcessor.convert(fp, Path(out)/f"{fp.stem}_clean.{fmt}",
                                              fmt, log=self._log, control=self._job_ctrl)
                    self._log(f"Preset batch done: {len(files)} files")
                else:
                    self._log(f"Job type '{jt}' not supported in preset runner")
                self.after(0, lambda: self._finish_job(f"Preset:{name} done"))
            except Exception as ex:
                self._log(f"Preset error: {ex}")
                self.after(0, lambda: self._finish_job("Preset: error", False))

        self._run_bg(task)


# ═════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════
def main():
    # Fix Windows taskbar icon – replaces Python logo with our custom icon
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            "OMIO.UltraDBStudio.v4"
        )
    except Exception:
        pass

    try:
        app = UltraCSVStudio()
        app.mainloop()
    except Exception as _startup_err:
        import traceback as _tb
        _msg = _tb.format_exc()
        try:
            import tkinter.messagebox as _mb
            _mb.showerror(
                "OMIO Ultra DB Studio — Startup Error",
                f"The application encountered an error during startup:\n\n{_startup_err}\n\n"
                f"Please check requirements.txt and reinstall dependencies.\n\n"
                f"Details:\n{_msg[:800]}"
            )
        except Exception:
            pass
        raise

if __name__ == "__main__":
    main()

